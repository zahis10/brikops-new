"""Live HTTP + local Mongo probe for BATCH #586 spare tiles + matrix.

This intentionally targets the already-running development server. It does not
use ASGITransport and never prints minted bearer tokens.
"""

import json
import os
import sys
import uuid
from io import BytesIO
from datetime import datetime, timedelta, timezone

import requests
from openpyxl import load_workbook
from pymongo import MongoClient


os.environ["MONGO_URL"] = "mongodb://127.0.0.1:27017"
os.environ["DB_NAME"] = "contractor_ops"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from contractor_ops.router import _create_token  # noqa: E402


BASE_URL = "http://127.0.0.1:5000"
RESULTS = []


def check(name, condition, detail=""):
    if not condition:
        raise AssertionError(f"{name}: {detail}")
    RESULTS.append(name)
    print(f"[PASS] {name}" + (f" — {detail}" if detail else ""))


def call(method, path, headers, expected, **kwargs):
    response = requests.request(
        method,
        f"{BASE_URL}{path}",
        headers=headers,
        timeout=15,
        **kwargs,
    )
    check(
        f"{method} {path} -> {expected}",
        response.status_code == expected,
        f"status={response.status_code}",
    )
    return response


def unit_document(db, unit_id):
    return db.units.find_one({"id": unit_id}, {"_id": 0})


def emit_documents(label, before, after):
    """Print review-ready, synthetic probe documents without auth headers."""
    print(f"\n[V7 JSON] {label} BEFORE")
    print(json.dumps(before, ensure_ascii=False, sort_keys=True, indent=2))
    print(f"[V7 JSON] {label} AFTER")
    print(json.dumps(after, ensure_ascii=False, sort_keys=True, indent=2))


def only_profile_changed(before, after, expected_profile_id):
    expected = dict(before)
    expected["spare_profile_id"] = expected_profile_id
    return after == expected


def overview_spare_values(payload, unit_id):
    for building in payload["buildings"]:
        for floor in building["floors"]:
            for unit in floor["units"]:
                if unit["unit_id"] == unit_id:
                    return {
                        "spare_tiles_count": unit["spare_tiles_count"],
                        "spare_tiles": unit["spare_tiles"],
                    }
    raise AssertionError(f"unit missing from handover overview: {unit_id}")


def exported_spare_values(content):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["ליקויים"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    spare_headers = [header for header in headers if str(header).startswith("ספייר:")]
    return {header: values[headers.index(header)] for header in spare_headers}


def execution_matrix_export_values(content):
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["מטריצת ביצוע"]
    headers = [cell.value for cell in sheet[1]]
    spare_col = (
        headers.index("ריצוף ספייר") + 1
        if "ריצוף ספייר" in headers
        else None
    )
    rows = {}
    if spare_col:
        for row in range(2, sheet.max_row + 1):
            unit_no = str(sheet.cell(row=row, column=3).value)
            cell = sheet.cell(row=row, column=spare_col)
            rows[unit_no] = {
                "value": cell.value,
                "fill": cell.fill.fgColor.rgb[-6:],
            }
    return {"headers": headers, "rows": rows}


def main():
    client = MongoClient(os.environ["MONGO_URL"], serverSelectionTimeoutMS=3000)
    db = client[os.environ["DB_NAME"]]
    tag = uuid.uuid4().hex[:10]
    org_id = f"probe-spare-org-{tag}"
    project_id = f"probe-spare-project-{tag}"
    other_project_id = f"probe-spare-other-project-{tag}"
    pm_id = f"probe-spare-pm-{tag}"
    owner_id = f"probe-spare-owner-{tag}"
    management_id = f"probe-spare-management-{tag}"
    viewer_id = f"probe-spare-viewer-{tag}"
    building_id = f"probe-spare-building-main-{tag}"
    other_building_id = f"probe-spare-building-first-{tag}"
    floor_low_id = f"probe-spare-floor-low-{tag}"
    floor_high_id = f"probe-spare-floor-high-{tag}"
    unit_one_id = f"probe-spare-unit-1-{tag}"
    unit_two_id = f"probe-spare-unit-2-{tag}"
    unit_three_id = f"probe-spare-unit-3-{tag}"
    cross_unit_id = f"probe-spare-cross-unit-{tag}"
    profile_a = str(uuid.uuid4())
    profile_b = str(uuid.uuid4())
    task_id = f"probe-spare-task-{tag}"
    legacy_spare_tiles = [
        {"type": "ריצוף יבש", "count": 8, "notes": "אריחים"},
        {"type": "ריצוף מרפסות", "count": 3, "notes": "קרטון"},
        {"type": "סוג מותאם", "count": 2, "notes": ""},
        {"type": "חיפוי מטבח", "count": 0, "notes": "נבדק"},
    ]
    legacy_count = 13
    legacy_notes = "ריצוף יבש: אריחים, ריצוף מרפסות: קרטון, חיפוי מטבח: נבדק"

    user_ids = [pm_id, owner_id, management_id, viewer_id]
    project_ids = [project_id, other_project_id]
    building_ids = [building_id, other_building_id]
    floor_ids = [floor_low_id, floor_high_id]
    unit_ids = [unit_one_id, unit_two_id, unit_three_id, cross_unit_id]

    try:
        client.admin.command("ping")
        health = requests.get(f"{BASE_URL}/api/health", timeout=10)
        check("live server reachable without ASGITransport", health.status_code < 500,
              f"status={health.status_code}")

        now = datetime.now(timezone.utc)
        db.organizations.insert_one({
            "id": org_id,
            "name": f"Spare Probe {tag}",
            "owner_user_id": owner_id,
        })
        db.subscriptions.insert_one({
            "id": f"probe-spare-sub-{tag}",
            "org_id": org_id,
            "status": "active",
            "paid_until": (now + timedelta(days=30)).isoformat(),
        })
        db.projects.insert_many([
            {
                "id": project_id,
                "name": f"Spare Probe Project {tag}",
                "code": tag,
                "org_id": org_id,
                "status": "active",
                "archived": False,
            },
            {
                "id": other_project_id,
                "name": f"Spare Probe Other {tag}",
                "org_id": org_id,
                "status": "active",
                "archived": False,
            },
        ])
        db.users.insert_many([
            {
                "id": pm_id,
                "role": "project_manager",
                "full_name": "מנהל בדיקת ספייר",
                "user_status": "active",
                "session_version": 0,
            },
            {
                "id": owner_id,
                "role": "owner",
                "full_name": "בעלים בדיקת ספייר",
                "user_status": "active",
                "session_version": 0,
            },
            {
                "id": viewer_id,
                "role": "viewer",
                "full_name": "צופה בדיקת ספייר",
                "user_status": "active",
                "session_version": 0,
            },
            {
                "id": management_id,
                "role": "management_team",
                "full_name": "צוות ניהול בדיקת ספייר",
                "user_status": "active",
                "session_version": 0,
            },
        ])
        db.organization_memberships.insert_many([
            {"id": f"probe-om-pm-{tag}", "org_id": org_id, "user_id": pm_id, "role": "member"},
            {"id": f"probe-om-owner-{tag}", "org_id": org_id, "user_id": owner_id, "role": "owner"},
            {"id": f"probe-om-management-{tag}", "org_id": org_id,
             "user_id": management_id, "role": "member"},
            {"id": f"probe-om-view-{tag}", "org_id": org_id, "user_id": viewer_id, "role": "member"},
        ])
        db.project_memberships.insert_many([
            {"id": f"probe-pm-pm-{tag}", "project_id": project_id, "user_id": pm_id,
             "role": "project_manager"},
            {"id": f"probe-pm-owner-{tag}", "project_id": project_id, "user_id": owner_id,
             "role": "owner"},
            {"id": f"probe-pm-management-{tag}", "project_id": project_id,
             "user_id": management_id, "role": "management_team"},
            {"id": f"probe-pm-view-{tag}", "project_id": project_id, "user_id": viewer_id,
             "role": "viewer"},
            {"id": f"probe-pm-other-{tag}", "project_id": other_project_id, "user_id": pm_id,
             "role": "project_manager"},
        ])
        # Insert deliberately out of display order.
        db.buildings.insert_many([
            {"id": building_id, "project_id": project_id, "name": "בניין 20",
             "sort_index": 20, "archived": False},
            {"id": other_building_id, "project_id": project_id, "name": "בניין 2",
             "sort_index": 10, "archived": False},
        ])
        db.floors.insert_many([
            {"id": floor_high_id, "project_id": project_id, "building_id": building_id,
             "name": "קומה 8", "floor_number": 8, "sort_index": 8000, "archived": False},
            {"id": floor_low_id, "project_id": project_id, "building_id": building_id,
             "name": "קומה 1", "floor_number": 1, "sort_index": 1000, "archived": False},
        ])
        db.units.insert_many([
            {
                "id": unit_two_id, "project_id": project_id, "building_id": building_id,
                "floor_id": floor_low_id, "unit_no": "102", "display_label": "דירה 102",
                "sort_index": 20, "status": "available", "archived": False,
            },
            {
                "id": unit_one_id, "project_id": project_id, "building_id": building_id,
                "floor_id": floor_low_id, "unit_no": "101", "display_label": "דירה 101",
                "sort_index": 10, "status": "available", "archived": False,
                "spare_profile_id": None,
                "spare_tiles": legacy_spare_tiles,
                "spare_tiles_count": legacy_count,
                "spare_tiles_notes": legacy_notes,
            },
            {
                "id": unit_three_id, "project_id": project_id, "building_id": building_id,
                "floor_id": floor_high_id, "unit_no": "801", "display_label": "דירה 801",
                "sort_index": 5, "status": "available", "archived": False,
            },
            {
                "id": cross_unit_id, "project_id": other_project_id,
                "building_id": f"probe-cross-building-{tag}",
                "floor_id": f"probe-cross-floor-{tag}", "unit_no": "X1",
                "sort_index": 1, "status": "available", "archived": False,
                "spare_tiles": legacy_spare_tiles,
                "spare_tiles_count": legacy_count,
                "spare_tiles_notes": legacy_notes,
            },
        ])
        db.tasks.insert_one({
            "id": task_id,
            "project_id": project_id,
            "building_id": building_id,
            "floor_id": floor_low_id,
            "unit_id": unit_one_id,
            "display_number": f"SPARE-{tag}",
            "title": "בדיקת שימור ספייר",
            "description": "",
            "category": "tiling",
            "status": "open",
            "created_at": now.isoformat(),
            "updated_at": now.isoformat(),
            "archived": False,
        })

        pm_headers = {"Authorization": f"Bearer {_create_token(pm_id, 'project_manager')}"}
        owner_headers = {"Authorization": f"Bearer {_create_token(owner_id, 'owner')}"}
        management_headers = {
            "Authorization": f"Bearer {_create_token(management_id, 'management_team')}"
        }
        viewer_headers = {"Authorization": f"Bearer {_create_token(viewer_id, 'viewer')}"}

        # V3 baseline before any project settings are persisted.
        baseline_response = call("GET", f"/api/units/{unit_one_id}", viewer_headers, 200)
        baseline = baseline_response.json()
        check("V3 absent settings has no profiles", baseline["spare_profiles_exist"] is False)
        check("V3 absent settings status is no_profile",
              baseline["spare_status"]["overall"] == "no_profile")
        additive_top = {
            "spare_settings",
            "spare_profiles_exist",
            "spare_status",
            "spare_can_write",
            "spare_can_assign",
        }
        baseline_top_keys = set(baseline) - additive_top
        baseline_unit_keys = set(baseline["unit"]) - {"spare_profile_id"}
        check("V3 baseline existing top-level keys captured", bool(baseline_top_keys))
        check("V3 baseline existing unit keys captured", bool(baseline_unit_keys))
        matrix_path = f"/api/execution-matrix/{project_id}"
        matrix_export_path = f"{matrix_path}/export.xlsx"
        matrix_before = call("GET", matrix_path, pm_headers, 200).json()
        matrix_keys_before = set(matrix_before) - {"spare"}
        check("V2 matrix without settings disables spare",
              matrix_before["spare"] == {"enabled": False, "by_unit": {}})
        matrix_export_before = execution_matrix_export_values(
            call("POST", matrix_export_path, pm_headers, 200, json={}).content
        )
        check("V2 matrix export without settings has no spare header",
              "ריצוף ספייר" not in matrix_export_before["headers"])
        legacy_before = unit_document(db, unit_one_id)
        check("V7 seeded legacy spare_tiles exact array",
              legacy_before["spare_tiles"] == legacy_spare_tiles)
        check("V7 seeded legacy aggregate fields exact",
              legacy_before["spare_tiles_count"] == legacy_count
              and legacy_before["spare_tiles_notes"] == legacy_notes)

        overview_path = f"/api/projects/{project_id}/handover/overview"
        overview_before = overview_spare_values(
            call("GET", overview_path, pm_headers, 200).json(), unit_one_id
        )
        export_request = {
            "scope": "unit",
            "unit_id": unit_one_id,
            "format": "excel",
            "filters": {},
        }
        export_before = exported_spare_values(
            call("POST", "/api/defects/export", pm_headers, 200, json=export_request).content
        )

        settings_path = f"/api/projects/{project_id}/spare-settings"
        assignments_path = f"/api/projects/{project_id}/spare-assignments"
        profile_payload = {
            "categories": [
                {"name": "ריצוף יבש", "measure": "tiles"},
                {"name": "ריצוף מרפסות", "measure": "tiles"},
                {"name": "חיפוי מטבח", "measure": "sqm"},
            ],
            "profiles": [
                {"id": profile_a, "name": "3 חדרים",
                 "targets": {"ריצוף יבש": 10, "ריצוף מרפסות": 3}},
                {"id": profile_b, "name": "4 חדרים",
                 "targets": {"ריצוף יבש": 12, "ריצוף מרפסות": 3}},
            ],
            "margin_pct": 10,
        }

        viewer_get = call("GET", settings_path, viewer_headers, 200).json()
        check("V2 viewer GET settings allowed and read-only",
              viewer_get["can_write"] is False)
        management_get = call("GET", settings_path, management_headers, 200).json()
        check("V2 management team GET settings is read-only",
              management_get["can_write"] is False)
        call("GET", assignments_path, viewer_headers, 200)
        call("PUT", settings_path, viewer_headers, 403, json=profile_payload)
        management_put = call(
            "PUT", settings_path, management_headers, 403, json=profile_payload
        ).json()
        check("V2 management team settings rejection message is exact",
              management_put["detail"]
              == "רק מנהל הפרויקט יכול לשנות הגדרות ושיוך ריצוף ספייר")
        call("PATCH", f"{settings_path.rsplit('/', 1)[0]}/spare-profiles/{profile_a}/units",
             viewer_headers, 403, json={"add": [unit_one_id], "remove": []})
        call("PATCH", f"{settings_path.rsplit('/', 1)[0]}/spare-profiles/{profile_a}/units",
             management_headers, 403, json={"add": [unit_one_id], "remove": []})

        pm_saved = call("PUT", settings_path, pm_headers, 200, json=profile_payload).json()
        check("V2 PM saved two profiles",
              [p["id"] for p in pm_saved["profiles"]] == [profile_a, profile_b])
        after_settings_save = unit_document(db, unit_one_id)
        emit_documents("settings save", legacy_before, after_settings_save)
        check("V7 whole unit unchanged after settings/profile save",
              after_settings_save == legacy_before)
        pm_version = pm_saved["updated_at"]

        owner_payload = {
            **profile_payload,
            "margin_pct": 11,
            "updated_at": pm_version,
        }
        owner_saved = call("PUT", settings_path, owner_headers, 200, json=owner_payload).json()
        check("V2 owner project-scoped write allowed",
              owner_saved["can_write"] is True and owner_saved["margin_pct"] == 11)
        owner_version = owner_saved["updated_at"]

        cartons_payload = {
            **profile_payload,
            "categories": [
                {"name": "ריצוף יבש", "measure": "cartons"},
                *profile_payload["categories"][1:],
            ],
            "updated_at": owner_version,
        }
        cartons_saved = call(
            "PUT", settings_path, pm_headers, 200, json=cartons_payload
        ).json()
        check("V3 cartons measure coerced to tiles",
              cartons_saved["categories"][0]["measure"] == "tiles"
              and db.projects.find_one(
                  {"id": project_id}, {"_id": 0, "spare_settings": 1}
              )["spare_settings"]["categories"][0]["measure"] == "tiles")
        owner_version = cartons_saved["updated_at"]

        stale_payload = {**profile_payload, "margin_pct": 12, "updated_at": pm_version}
        call("PUT", settings_path, pm_headers, 409, json=stale_payload)

        assignment_payload = call(
            "GET", f"{assignments_path}?building_id={building_id}", viewer_headers, 200
        ).json()
        check("V2 buildings preserve structure ordering",
              [b["id"] for b in assignment_payload["buildings"]]
              == [other_building_id, building_id])
        check("V2 floors preserve sort_index ordering",
              [f["id"] for f in assignment_payload["floors"]]
              == [floor_low_id, floor_high_id])
        check("V2 units preserve sort_index ordering",
              [u["id"] for u in assignment_payload["floors"][0]["units"]]
              == [unit_one_id, unit_two_id])
        expected_unit_keys = {"id", "unit_no", "display_label", "spare_profile_id"}
        check("V2 assignment unit payload is exact",
              set(assignment_payload["floors"][0]["units"][0]) == expected_unit_keys)

        management_unit = call(
            "GET", f"/api/units/{unit_one_id}", management_headers, 200
        ).json()
        check("V3 management team can update inventory but cannot assign profiles",
              management_unit["spare_can_write"] is True
              and management_unit["spare_can_assign"] is False)
        pm_unit = call("GET", f"/api/units/{unit_one_id}", pm_headers, 200).json()
        check("V3 PM can update inventory and assign profiles",
              pm_unit["spare_can_write"] is True
              and pm_unit["spare_can_assign"] is True)

        management_inventory_before = unit_document(db, unit_one_id)
        call("PATCH", f"/api/units/{unit_one_id}/spare-tiles", management_headers, 200,
             json={"spare_tiles": legacy_spare_tiles})
        management_inventory_after = unit_document(db, unit_one_id)
        check("V3 management team inventory no-op preserves whole unit",
              management_inventory_after == management_inventory_before)

        profile_base = f"/api/projects/{project_id}/spare-profiles"
        added = call("PATCH", f"{profile_base}/{profile_a}/units", pm_headers, 200,
                     json={"add": [unit_one_id, unit_two_id], "remove": []}).json()
        check("V2 assignment add exact count", added == {"added": 2, "removed": 0})
        after_assign = unit_document(db, unit_one_id)
        emit_documents("profile assignment", legacy_before, after_assign)
        check("V7 assignment changes only spare_profile_id",
              only_profile_changed(legacy_before, after_assign, profile_a))
        moved = call("PATCH", f"{profile_base}/{profile_b}/units", owner_headers, 200,
                     json={"add": [unit_one_id], "remove": []}).json()
        check("V2 assignment move exact count", moved == {"added": 1, "removed": 0})
        removed = call("PATCH", f"{profile_base}/{profile_b}/units", pm_headers, 200,
                       json={"add": [], "remove": [unit_one_id]}).json()
        check("V2 assignment remove exact count", removed == {"added": 0, "removed": 1})
        after_remove = unit_document(db, unit_one_id)
        emit_documents("profile removal", after_assign, after_remove)
        check("V7 removal restores whole unit with spare_profile_id None",
              after_remove == legacy_before)
        removed_again = call("PATCH", f"{profile_base}/{profile_b}/units", pm_headers, 200,
                             json={"add": [], "remove": [unit_one_id]}).json()
        check("V2 no-op removal reports zero", removed_again == {"added": 0, "removed": 0})

        delete_assigned = {
            "categories": owner_saved["categories"],
            "profiles": [p for p in owner_saved["profiles"] if p["id"] != profile_a],
            "margin_pct": owner_saved["margin_pct"],
            "updated_at": owner_version,
        }
        call("PUT", settings_path, pm_headers, 409, json=delete_assigned)
        call("PATCH", f"{profile_base}/{profile_b}/units", pm_headers, 422,
             json={"add": [cross_unit_id], "remove": []})

        unchanged_save_before = unit_document(db, unit_one_id)
        call("PATCH", f"/api/units/{unit_one_id}/spare-tiles", pm_headers, 200,
             json={"spare_tiles": legacy_spare_tiles})
        unchanged_save_after = unit_document(db, unit_one_id)
        emit_documents("profile-aware unchanged inventory save",
                       unchanged_save_before, unchanged_save_after)
        check("V7 profile-aware unchanged save preserves whole unit",
              unchanged_save_after == unchanged_save_before)

        call("PATCH", f"{profile_base}/{profile_a}/units", pm_headers, 200,
             json={"add": [unit_one_id], "remove": []})
        confirmed_zero_payload = [
            {"type": "ריצוף יבש", "count": 0, "notes": "", "entered": True},
            {"type": "ריצוף מרפסות", "count": 3, "notes": "קרטון"},
            {"type": "חיפוי מטבח", "count": 0, "notes": "", "entered": True},
        ]
        call("PATCH", f"/api/units/{unit_one_id}/spare-tiles", pm_headers, 200,
             json={"spare_tiles": confirmed_zero_payload})
        confirmed_zero_doc = unit_document(db, unit_one_id)
        emit_documents("explicit confirmed zero", unchanged_save_after, confirmed_zero_doc)
        check("V1 confirmed zero stored exactly",
              confirmed_zero_doc["spare_tiles"] == confirmed_zero_payload)
        confirmed_response = call(
            "GET", f"/api/units/{unit_one_id}", viewer_headers, 200
        ).json()
        confirmed_dry_row = next(
            row for row in confirmed_response["spare_status"]["categories"]
            if row["name"] == "ריצוף יבש"
        )
        check("V1 confirmed zero is short with full target missing",
              confirmed_dry_row["actual"] == 0
              and confirmed_dry_row["entered"] is True
              and confirmed_dry_row["status"] == "short"
              and confirmed_dry_row["missing"] == confirmed_dry_row["target"] == 10)
        confirmed_kitchen_row = next(
            row for row in confirmed_response["spare_status"]["categories"]
            if row["name"] == "חיפוי מטבח"
        )
        check("V1 confirmed zero without target is short with unknown missing",
              confirmed_kitchen_row["actual"] == 0
              and confirmed_kitchen_row["entered"] is True
              and confirmed_kitchen_row["status"] == "short"
              and confirmed_kitchen_row["target"] is None
              and confirmed_kitchen_row["missing"] is None)
        confirmed_matrix = call("GET", matrix_path, pm_headers, 200).json()
        confirmed_summary = confirmed_matrix["spare"]["by_unit"][unit_one_id]
        check("V2 matrix includes confirmed zero without target as short",
              confirmed_summary["overall"] == "short"
              and any(
                  row["name"] == "חיפוי מטבח" and row["missing"] is None
                  for row in confirmed_summary["short"]
              ))
        confirmed_matrix_export = execution_matrix_export_values(
            call("POST", matrix_export_path, pm_headers, 200, json={}).content
        )
        confirmed_export_value = confirmed_matrix_export["rows"]["101"]["value"]
        check("V2 matrix export prints confirmed zero without target honestly",
              confirmed_export_value.startswith("חסר — להזמין:")
              and "חיפוי מטבח (אין ספייר)" in confirmed_export_value
              and "ריצוף יבש 10" in confirmed_export_value)

        call("PATCH", f"/api/units/{unit_one_id}/spare-tiles", pm_headers, 200,
             json={"spare_tiles": legacy_spare_tiles})
        restored_legacy = unit_document(db, unit_one_id)
        emit_documents("legacy array after confirmed zero", confirmed_zero_doc, restored_legacy)
        check("V7 legacy array restored without entered keys",
              restored_legacy["spare_tiles"] == legacy_spare_tiles
              and all("entered" not in row for row in restored_legacy["spare_tiles"]))
        call("PATCH", f"{profile_base}/{profile_a}/units", pm_headers, 200,
             json={"add": [], "remove": [unit_one_id]})

        category_delete_payload = {
            **owner_payload,
            "categories": [
                category for category in owner_saved["categories"]
                if category["name"] != "חיפוי מטבח"
            ],
            "updated_at": owner_version,
        }
        category_saved = call(
            "PUT", settings_path, owner_headers, 200, json=category_delete_payload
        ).json()
        check("V7 category deletion persisted",
              all(item["name"] != "חיפוי מטבח" for item in category_saved["categories"]))
        category_save_before = unit_document(db, unit_one_id)
        call("PATCH", f"/api/units/{unit_one_id}/spare-tiles", pm_headers, 200,
             json={"spare_tiles": legacy_spare_tiles})
        category_save_after = unit_document(db, unit_one_id)
        emit_documents("save after category deletion",
                       category_save_before, category_save_after)
        check("V7 deleted configured category remains exact custom entry",
              category_save_after == category_save_before
              and category_save_after["spare_tiles"][-1]
              == {"type": "חיפוי מטבח", "count": 0, "notes": "נבדק"})

        legacy_branch_before = unit_document(db, cross_unit_id)
        call("PATCH", f"/api/units/{cross_unit_id}/spare-tiles", pm_headers, 200,
             json={"spare_tiles": legacy_spare_tiles})
        legacy_branch_after = unit_document(db, cross_unit_id)
        emit_documents("legacy no-profile unchanged save",
                       legacy_branch_before, legacy_branch_after)
        check("V7 no-profile legacy branch preserves whole unit",
              legacy_branch_after == legacy_branch_before)

        final_add = call("PATCH", f"{profile_base}/{profile_a}/units", pm_headers, 200,
                         json={"add": [unit_one_id], "remove": []}).json()
        check("V3 final profile assignment exact count",
              final_add == {"added": 1, "removed": 0})
        after = call("GET", f"/api/units/{unit_one_id}", viewer_headers, 200).json()
        check("V3 existing top-level keys unchanged",
              set(after) - additive_top == baseline_top_keys)
        check("V3 existing nested unit keys unchanged",
              set(after["unit"]) - {"spare_profile_id"} == baseline_unit_keys)
        check("V3 assigned profile returned",
              after["spare_status"]["profile"]["id"] == profile_a)
        dry_row = next(row for row in after["spare_status"]["categories"]
                       if row["name"] == "ריצוף יבש")
        check("V3 status reflects actual 8 against target 10",
              after["spare_status"]["overall"] == "short"
              and dry_row["actual"] == 8 and dry_row["target"] == 10
              and dry_row["missing"] == 2)
        check("V3 viewer unit capabilities are false",
              after["spare_can_write"] is False
              and after["spare_can_assign"] is False)

        matrix_after = call("GET", matrix_path, pm_headers, 200).json()
        check("V2 matrix existing key set unchanged",
              set(matrix_after) - {"spare"} == matrix_keys_before)
        matrix_spare = matrix_after["spare"]
        expected_short = [{
            "name": "ריצוף יבש",
            "missing": 2,
            "measure": "tiles",
        }]
        check("V2 matrix unit A short summary",
              matrix_spare["enabled"] is True
              and matrix_spare["by_unit"][unit_one_id]["overall"] == "short"
              and matrix_spare["by_unit"][unit_one_id]["short"] == expected_short
              and matrix_spare["by_unit"][unit_one_id]["borderline"]
              == ["ריצוף מרפסות"]
              and matrix_spare["by_unit"][unit_one_id]["missing_total"] == 2)
        check("V2 matrix unit B not entered",
              matrix_spare["by_unit"][unit_two_id]["overall"] == "not_entered")
        check("V2 matrix unit C no profile",
              matrix_spare["by_unit"][unit_three_id]["overall"] == "no_profile")
        matrix_export_after = execution_matrix_export_values(
            call("POST", matrix_export_path, pm_headers, 200, json={}).content
        )
        check("V2 matrix export spare header is last",
              matrix_export_after["headers"][-1] == "ריצוף ספייר")
        check("V2 matrix export short text and fill",
              matrix_export_after["rows"]["101"]["value"]
              == "חסר — להזמין: ריצוף יבש 2 · גבולי: ריצוף מרפסות"
              and matrix_export_after["rows"]["101"]["fill"] == "FEE2E2")
        check("V2 matrix export not-entered and no-profile values",
              matrix_export_after["rows"]["102"]["value"]
              == "לא הוזן: ריצוף יבש, ריצוף מרפסות"
              and matrix_export_after["rows"]["801"]["value"] == "אחר")
        print("\n[V2 JSON] execution matrix spare")
        print(json.dumps(matrix_spare, ensure_ascii=False, sort_keys=True, indent=2))

        overview_after = overview_spare_values(
            call("GET", overview_path, pm_headers, 200).json(), unit_one_id
        )
        export_after = exported_spare_values(
            call("POST", "/api/defects/export", pm_headers, 200, json=export_request).content
        )
        print("\n[V7 JSON] handover/export BEFORE")
        print(json.dumps({"handover": overview_before, "export": export_before},
                         ensure_ascii=False, sort_keys=True, indent=2))
        print("[V7 JSON] handover/export AFTER")
        print(json.dumps({"handover": overview_after, "export": export_after},
                         ensure_ascii=False, sort_keys=True, indent=2))
        check("V7 handover overview spare values unchanged",
              overview_after == overview_before)
        check("V7 export spare columns unchanged", export_after == export_before)

        audit_actions = {
            row["action"] for row in db.audit_events.find(
                {"entity_type": "project", "entity_id": project_id},
                {"_id": 0, "action": 1},
            )
        }
        check("V2 settings audit action recorded",
              "spare_settings_updated" in audit_actions)
        check("V2 assignment audit action recorded",
              "spare_profile_units_updated" in audit_actions)
        print(f"\nAll {len(RESULTS)} live HTTP checks passed.")
    finally:
        db.audit_events.delete_many({"entity_id": {"$in": project_ids}})
        db.execution_matrix.delete_many({"project_id": {"$in": project_ids}})
        db.tasks.delete_many({"id": task_id})
        db.units.delete_many({"id": {"$in": unit_ids}})
        db.floors.delete_many({"id": {"$in": floor_ids}})
        db.buildings.delete_many({"id": {"$in": building_ids}})
        db.project_memberships.delete_many({
            "$or": [{"project_id": {"$in": project_ids}}, {"user_id": {"$in": user_ids}}]
        })
        db.organization_memberships.delete_many({
            "$or": [{"org_id": org_id}, {"user_id": {"$in": user_ids}}]
        })
        db.subscriptions.delete_many({"org_id": org_id})
        db.projects.delete_many({"id": {"$in": project_ids}})
        db.users.delete_many({"id": {"$in": user_ids}})
        db.organizations.delete_many({"id": org_id})
        client.close()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\n[FAIL] {type(exc).__name__}: {exc}", file=sys.stderr)
        raise