"""Tests for Execution Matrix .xlsx export (Batch #502, 2026-05-05).

T1 build_matrix_xlsx structure (RTL, frozen, sheet name).
T2 status label translation + fill colors.
T3 note attached as Comment + whitespace strip + control-char strip
   + author fallback.
T4 endpoint RBAC (contractor 403, PM 200).
T5 filename sanitization + 2000-unit cap.
"""
import asyncio
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from contractor_ops import execution_matrix_router as emr
from contractor_ops.execution_matrix_export import (
    build_matrix_xlsx,
    STATUS_LABELS,
    STATUS_FILLS,
)


# =====================================================================
# Cursor / DB mocks (mirror test_execution_matrix.py)
# =====================================================================

class _AsyncCursor:
    def __init__(self, items):
        self._items = list(items)

    def __aiter__(self):
        async def gen():
            for it in self._items:
                yield it
        return gen()

    async def to_list(self, _n=None):
        return list(self._items)


def _mk_db_for_export(*, project=None, units=None, floors=None,
                      buildings=None, cells=None, matrix_config=None,
                      unit_count=None):
    db = MagicMock()
    db.projects.find_one = AsyncMock(
        return_value=project or {"id": "p1", "name": "Demo"}
    )
    db.units.count_documents = AsyncMock(
        return_value=(unit_count if unit_count is not None else len(units or []))
    )
    db.units.find = MagicMock(return_value=_AsyncCursor(units or []))
    db.floors.find = MagicMock(return_value=_AsyncCursor(floors or []))
    db.buildings.find = MagicMock(return_value=_AsyncCursor(buildings or []))
    db.execution_matrix_cells.find = MagicMock(return_value=_AsyncCursor(cells or []))
    db.execution_matrix.find_one = AsyncMock(return_value=matrix_config or {
        "id": "m1", "project_id": "p1",
        "custom_stages": [], "base_stages_removed": [], "deletedAt": None,
    })
    db.execution_matrix.insert_one = AsyncMock()
    db.execution_matrix.update_one = AsyncMock(return_value=MagicMock(matched_count=1))
    db.project_approvers.find = MagicMock(return_value=_AsyncCursor([]))
    return db


def _read_workbook(buf):
    buf.seek(0)
    return load_workbook(buf)


# =====================================================================
# T1 — build_matrix_xlsx structure
# =====================================================================

def test_build_xlsx_basic_structure():
    """Sheet name, RTL view, header row, frozen panes."""
    project = {"id": "p1", "name": "פרויקט הדמו"}
    stages = [
        {"id": "s1", "title": "ריצוף", "type": "status", "order": 10},
        {"id": "s2", "title": "צבע", "type": "tag", "order": 20},
        {"id": "s3", "title": "חשמל", "type": "status", "order": 30},
    ]
    units = [
        {"id": "u1", "building_id": "b1", "floor_id": "f1",
         "unit_no": "1", "room_count": 4},
        {"id": "u2", "building_id": "b1", "floor_id": "f1",
         "unit_no": "2", "room_count": 3},
    ]
    cells = [
        {"unit_id": "u1", "stage_id": "s1", "status": "completed",
         "text_value": None, "note": None, "last_actor_name": None},
        {"unit_id": "u1", "stage_id": "s2", "status": None,
         "text_value": "פרקט", "note": None, "last_actor_name": None},
        {"unit_id": "u2", "stage_id": "s3", "status": "not_done",
         "text_value": None, "note": None, "last_actor_name": None},
        # one cell intentionally absent → empty
    ]
    buildings = {"b1": {"id": "b1", "name": "מגדל א"}}
    floors = {"f1": {"id": "f1", "floor_number": 1}}

    buf = build_matrix_xlsx(project, units, stages, cells, buildings, floors)
    assert isinstance(buf, BytesIO)

    wb = _read_workbook(buf)
    assert wb.sheetnames == ["מטריצת ביצוע"]
    ws = wb.active
    assert ws.sheet_view.rightToLeft is True
    assert ws.freeze_panes == "E2"

    # Header row
    expected_header = ["בניין", "קומה", "דירה", "מס׳ חדרים", "ריצוף", "צבע", "חשמל"]
    actual_header = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert actual_header == expected_header
    # Header bold
    assert ws.cell(row=1, column=1).font.bold is True

    # Data rows count = 2
    assert ws.max_row == 3  # header + 2 units


# =====================================================================
# T2 — status label translation + fills
# =====================================================================

def test_status_label_translation():
    """status → Hebrew label + status-specific fill."""
    project = {"id": "p1", "name": "Demo"}
    stages = [
        {"id": "s1", "title": "ריצוף", "type": "status", "order": 10},
        {"id": "s2", "title": "חשמל", "type": "status", "order": 20},
        {"id": "s3", "title": "אינסטלציה", "type": "status", "order": 30},
    ]
    units = [{"id": "u1", "building_id": "b1", "floor_id": "f1", "unit_no": "1"}]
    cells = [
        {"unit_id": "u1", "stage_id": "s1", "status": "completed",
         "text_value": None, "note": None, "last_actor_name": None},
        {"unit_id": "u1", "stage_id": "s2", "status": "not_done",
         "text_value": None, "note": None, "last_actor_name": None},
        # s3 has no cell → blank
    ]
    buildings = {"b1": {"name": "מגדל"}}
    floors = {"f1": {"floor_number": 1}}

    buf = build_matrix_xlsx(project, units, stages, cells, buildings, floors)
    wb = _read_workbook(buf)
    ws = wb.active

    # row 2 col 5 = ריצוף = completed
    c_completed = ws.cell(row=2, column=5)
    assert c_completed.value == "הושלם"
    assert c_completed.fill.fgColor.rgb.upper().endswith("D1FAE5")

    # row 2 col 6 = חשמל = not_done
    c_notdone = ws.cell(row=2, column=6)
    assert c_notdone.value == "לא בוצע"
    assert c_notdone.fill.fgColor.rgb.upper().endswith("FEE2E2")

    # row 2 col 7 = אינסטלציה = no cell → blank
    c_blank = ws.cell(row=2, column=7)
    assert c_blank.value is None


# =====================================================================
# T3 — note as Comment + whitespace + control chars + author
# =====================================================================

def test_note_attached_as_comment_with_author_and_whitespace_strip():
    """Note Comment author = last_actor_name; whitespace/None = no Comment;
    fallback author = 'BrikOps'; control chars stripped (no IllegalCharacterError)."""
    project = {"id": "p1", "name": "Demo"}
    stages = [
        {"id": "s1", "title": "ריצוף", "type": "status", "order": 10},
        {"id": "s2", "title": "חשמל", "type": "status", "order": 20},
        {"id": "s3", "title": "צבע", "type": "status", "order": 30},
        {"id": "s4", "title": "אינסטלציה", "type": "status", "order": 40},
    ]
    units = [{"id": "u1", "building_id": "b1", "floor_id": "f1", "unit_no": "1"}]
    cells = [
        # Has note + last_actor_name
        {"unit_id": "u1", "stage_id": "s1", "status": "not_done",
         "text_value": None, "note": "לא תקין",
         "last_actor_name": "ישראל ישראלי"},
        # Whitespace-only note → no Comment
        {"unit_id": "u1", "stage_id": "s2", "status": "completed",
         "text_value": None, "note": "   ",
         "last_actor_name": "פלוני"},
        # None note → no Comment
        {"unit_id": "u1", "stage_id": "s3", "status": "completed",
         "text_value": None, "note": None, "last_actor_name": "אלמוני"},
        # Note with control chars + no last_actor_name → Comment, author=BrikOps
        {"unit_id": "u1", "stage_id": "s4", "status": "partial",
         "text_value": None, "note": "חלקי\x00מאוד\x0bכאן",
         "last_actor_name": None},
    ]
    buildings = {"b1": {"name": "מגדל"}}
    floors = {"f1": {"floor_number": 1}}

    buf = build_matrix_xlsx(project, units, stages, cells, buildings, floors)
    wb = _read_workbook(buf)
    ws = wb.active

    c_with_note = ws.cell(row=2, column=5)
    assert c_with_note.comment is not None
    assert c_with_note.comment.author == "ישראל ישראלי"
    assert "לא תקין" in c_with_note.comment.text

    c_whitespace = ws.cell(row=2, column=6)
    assert c_whitespace.comment is None, "whitespace-only note must NOT add Comment"

    c_none = ws.cell(row=2, column=7)
    assert c_none.comment is None, "None note must NOT add Comment"

    c_ctrl = ws.cell(row=2, column=8)
    assert c_ctrl.comment is not None, "control-char note still adds Comment"
    assert c_ctrl.comment.author == "BrikOps"
    # Control chars stripped, content preserved
    assert "\x00" not in c_ctrl.comment.text
    assert "\x0b" not in c_ctrl.comment.text
    assert "חלקי" in c_ctrl.comment.text
    assert "מאוד" in c_ctrl.comment.text
    assert "כאן" in c_ctrl.comment.text


# =====================================================================
# T4 — endpoint RBAC (contractor 403, PM 200)
# =====================================================================

def test_export_endpoint_403_for_contractor():
    """Contractor blocked at _check_matrix_view; PM gets StreamingResponse."""

    # --- 4a: contractor → 403 ---
    async def run_contractor():
        db = _mk_db_for_export()
        with patch.object(emr, "get_db", return_value=db), \
             patch.object(emr, "_get_project_role",
                          new=AsyncMock(return_value="contractor")), \
             patch.object(emr, "_is_super_admin", return_value=False):
            with pytest.raises(HTTPException) as exc_info:
                await emr.export_matrix_xlsx(
                    "p1", payload={}, user={"id": "c1"}
                )
            assert exc_info.value.status_code == 403
    asyncio.run(run_contractor())

    # --- 4b: PM → 200 + StreamingResponse ---
    async def run_pm():
        project = {"id": "p1", "name": "Demo"}
        units = [{"id": "u1", "project_id": "p1", "building_id": "b1",
                  "floor_id": "f1", "unit_no": "1"}]
        floors = [{"id": "f1", "floor_number": 1}]
        buildings = [{"id": "b1", "name": "מגדל"}]
        cells = [{"unit_id": "u1", "stage_id": "s1", "status": "completed",
                  "audit": [{"actor_name": "PM", "ts": 0}]}]
        tpl = {"id": "tpl1", "stages": [
            {"id": "s1", "title": "ריצוף", "scope": "unit", "order": 10},
        ]}
        db = _mk_db_for_export(
            project=project, units=units, floors=floors,
            buildings=buildings, cells=cells,
        )
        with patch.object(emr, "get_db", return_value=db), \
             patch.object(emr, "_get_project_role",
                          new=AsyncMock(return_value="project_manager")), \
             patch.object(emr, "_is_super_admin", return_value=False), \
             patch.object(emr, "_get_template", new=AsyncMock(return_value=tpl)):
            res = await emr.export_matrix_xlsx(
                "p1", payload={}, user={"id": "pm1"}
            )
        # Verify it's a StreamingResponse with xlsx mime type
        assert res.media_type == (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        assert "Content-Disposition" in res.headers
        assert "attachment" in res.headers["Content-Disposition"]
    asyncio.run(run_pm())


# =====================================================================
# T5 — filename sanitization + 2000-unit cap
# =====================================================================

def test_filename_sanitizes_illegal_chars_and_caps_count():
    """Project name illegal chars → underscores; >2000 units → 400."""

    # --- 5a: sanitize illegal chars in filename ---
    async def run_sanitize():
        project = {"id": "p1", "name": "רחוב הרצל / חיפה 123:test"}
        units = [{"id": "u1", "project_id": "p1", "building_id": "b1",
                  "floor_id": "f1", "unit_no": "1"}]
        floors = [{"id": "f1", "floor_number": 1}]
        buildings = [{"id": "b1", "name": "מגדל"}]
        tpl = {"id": "tpl1", "stages": [
            {"id": "s1", "title": "ריצוף", "scope": "unit", "order": 10},
        ]}
        db = _mk_db_for_export(
            project=project, units=units, floors=floors, buildings=buildings,
        )
        with patch.object(emr, "get_db", return_value=db), \
             patch.object(emr, "_get_project_role",
                          new=AsyncMock(return_value="project_manager")), \
             patch.object(emr, "_is_super_admin", return_value=False), \
             patch.object(emr, "_get_template", new=AsyncMock(return_value=tpl)):
            res = await emr.export_matrix_xlsx(
                "p1", payload={}, user={"id": "pm1"}
            )
        cd = res.headers["Content-Disposition"]
        # `/` and `:` stripped from project name in encoded filename
        # (encoded filename is URL-encoded so check for the underscore presence
        # by decoding the relevant slice)
        from urllib.parse import unquote
        # Pull the filename* value
        import re as _re
        m = _re.search(r"filename\*=UTF-8''([^;]+)", cd)
        assert m, f"no filename* in Content-Disposition: {cd}"
        decoded = unquote(m.group(1))
        # Original "/" and ":" must be replaced with "_"
        assert "/" not in decoded.split(".xlsx")[0], f"slash leaked: {decoded}"
        assert ":" not in decoded.split(".xlsx")[0], f"colon leaked: {decoded}"
        assert "_" in decoded, f"sanitization marker missing: {decoded}"
    asyncio.run(run_sanitize())

    # --- 5b: 2001 units → 400 with Hebrew "2000 דירות" ---
    async def run_cap_exceeded():
        db = _mk_db_for_export(unit_count=2001)
        with patch.object(emr, "get_db", return_value=db), \
             patch.object(emr, "_get_project_role",
                          new=AsyncMock(return_value="project_manager")), \
             patch.object(emr, "_is_super_admin", return_value=False):
            with pytest.raises(HTTPException) as exc_info:
                await emr.export_matrix_xlsx(
                    "p1", payload={}, user={"id": "pm1"}
                )
            assert exc_info.value.status_code == 400
            assert "2000 דירות" in exc_info.value.detail
    asyncio.run(run_cap_exceeded())

    # --- 5c: 1500 units → 200 ---
    async def run_under_cap():
        project = {"id": "p1", "name": "Demo"}
        # Just need 1 actual unit doc — count is mocked separately
        units = [{"id": "u1", "project_id": "p1", "building_id": "b1",
                  "floor_id": "f1", "unit_no": "1"}]
        floors = [{"id": "f1", "floor_number": 1}]
        buildings = [{"id": "b1", "name": "מגדל"}]
        tpl = {"id": "tpl1", "stages": [
            {"id": "s1", "title": "ריצוף", "scope": "unit", "order": 10},
        ]}
        db = _mk_db_for_export(
            project=project, units=units, floors=floors,
            buildings=buildings, unit_count=1500,
        )
        with patch.object(emr, "get_db", return_value=db), \
             patch.object(emr, "_get_project_role",
                          new=AsyncMock(return_value="project_manager")), \
             patch.object(emr, "_is_super_admin", return_value=False), \
             patch.object(emr, "_get_template", new=AsyncMock(return_value=tpl)):
            res = await emr.export_matrix_xlsx(
                "p1", payload={}, user={"id": "pm1"}
            )
        assert res.status_code == 200
    asyncio.run(run_under_cap())
