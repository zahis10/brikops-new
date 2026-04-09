import io
import logging
from collections import Counter
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from starlette.responses import StreamingResponse

from contractor_ops.router import get_db, get_current_user, _check_project_read_access, _get_project_role, _audit
from models import CATEGORIES
from services.object_storage import generate_url

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")

CATEGORY_LABEL = {c['value']: c['label'] for c in CATEGORIES}

STATUS_LABEL = {
    'open': 'פתוח',
    'assigned': 'שויך',
    'in_progress': 'בביצוע',
    'waiting_verify': 'ממתין לאימות',
    'pending_contractor_proof': 'ממתין להוכחת קבלן',
    'pending_manager_approval': 'ממתין לאישור מנהל',
    'returned_to_contractor': 'הוחזר לקבלן',
    'closed': 'סגור',
    'reopened': 'נפתח מחדש',
}

STATUS_COLORS = {
    'open': '#EF4444',
    'assigned': '#F59E0B',
    'in_progress': '#3B82F6',
    'waiting_verify': '#8B5CF6',
    'pending_contractor_proof': '#F59E0B',
    'pending_manager_approval': '#F59E0B',
    'returned_to_contractor': '#EF4444',
    'closed': '#22C55E',
    'reopened': '#EF4444',
}

OPEN_STATUSES_APARTMENT = {'open', 'assigned', 'reopened'}
OPEN_STATUSES_BUILDING = {'open', 'assigned', 'reopened', 'in_progress',
                          'pending_contractor_proof', 'pending_manager_approval',
                          'returned_to_contractor', 'waiting_verify'}
IN_PROGRESS_STATUSES = {'in_progress', 'pending_contractor_proof',
                        'pending_manager_approval', 'returned_to_contractor',
                        'waiting_verify'}
CLOSED_STATUSES = {'closed'}
BLOCKING_STATUSES = {'open', 'in_progress'}


class ExportFilters(BaseModel):
    status: Optional[str] = 'all'
    category: Optional[str] = 'all'
    company: Optional[str] = 'all'
    assignee: Optional[str] = 'all'
    created_by: Optional[str] = 'all'
    floor: Optional[str] = 'all'
    unit: Optional[str] = 'all'
    search: Optional[str] = ''


VALID_EXPORT_FORMATS = {'excel', 'pdf'}


class ExportRequest(BaseModel):
    scope: str
    unit_id: Optional[str] = None
    building_id: Optional[str] = None
    project_id: Optional[str] = None
    filters: ExportFilters = ExportFilters()
    format: Optional[str] = 'excel'


def _apply_filters(tasks, filters: ExportFilters, scope: str = 'unit'):
    result = []
    search_lower = (filters.search or '').strip().lower()
    open_statuses = OPEN_STATUSES_BUILDING if scope == 'building' else OPEN_STATUSES_APARTMENT
    for t in tasks:
        s = t.get('status', 'open')
        if filters.status and filters.status != 'all':
            if filters.status == 'open' and s not in open_statuses:
                continue
            elif filters.status == 'in_progress' and s not in IN_PROGRESS_STATUSES:
                continue
            elif filters.status == 'closed' and s not in CLOSED_STATUSES:
                continue
            elif filters.status == 'blocking' and s not in BLOCKING_STATUSES:
                continue
        if filters.category and filters.category != 'all':
            if t.get('category') != filters.category:
                continue
        if filters.company and filters.company != 'all':
            if t.get('company_id') != filters.company:
                continue
        if filters.assignee and filters.assignee != 'all':
            if t.get('assignee_id') != filters.assignee:
                continue
        if filters.created_by and filters.created_by != 'all':
            if t.get('created_by') != filters.created_by:
                continue
        if filters.floor and filters.floor != 'all':
            if t.get('floor_id') != filters.floor:
                continue
        if filters.unit and filters.unit != 'all':
            if t.get('unit_id') != filters.unit:
                continue
        if search_lower:
            title = (t.get('title') or '').lower()
            desc = (t.get('description') or '').lower()
            if search_lower not in title and search_lower not in desc:
                continue
        result.append(t)
    return result


_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.heic', '.heif', '.tiff', '.tif'}


def _is_image_ref(url, content_type=None):
    if content_type and content_type.startswith('image/'):
        return True
    import os
    _, ext = os.path.splitext(url.split('?')[0])
    return ext.lower() in _IMAGE_EXTENSIONS


async def _resolve_image_links(db, task):
    links = []
    seen = set()

    try:
        task_id = task.get('id') or task.get('task_id') or ''
        if task_id:
            att_updates = await db.task_updates.find({
                'task_id': task_id,
                'update_type': 'attachment',
                'deletedAt': {'$exists': False},
            }, {'_id': 0, 'attachment_url': 1, 'content_type': 1}).to_list(200)

            for att in att_updates:
                url = att.get('attachment_url') or ''
                if not url:
                    continue
                ct = att.get('content_type') or ''
                if not _is_image_ref(url, ct):
                    continue
                resolved = generate_url(url) if url.startswith('s3://') else url
                if resolved not in seen:
                    seen.add(resolved)
                    links.append(resolved)

        for att in (task.get('attachments') or []):
            url = att.get('file_url') or att.get('url') or ''
            if url and _is_image_ref(url):
                resolved = generate_url(url) if url.startswith('s3://') else url
                if resolved not in seen:
                    seen.add(resolved)
                    links.append(resolved)

        proof_urls = task.get('proof_urls') or []
        for url in proof_urls:
            if isinstance(url, str) and url and _is_image_ref(url):
                resolved = generate_url(url) if url.startswith('s3://') else url
                if resolved not in seen:
                    seen.add(resolved)
                    links.append(resolved)
    except Exception as e:
        logger.warning(f'[EXPORT] Image link resolution failed for task {task.get("id","?")}: {e}')

    return links


def _format_datetime(dt_str):
    if not dt_str:
        return ''
    try:
        if isinstance(dt_str, datetime):
            return dt_str.strftime('%Y-%m-%d %H:%M')
        dt = datetime.fromisoformat(str(dt_str).replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M')
    except Exception:
        return str(dt_str)[:16] if dt_str else ''


async def _build_lookup_maps(db, tasks):
    user_ids = set()
    company_ids = set()
    for t in tasks:
        if t.get('assignee_id'):
            user_ids.add(t['assignee_id'])
        if t.get('created_by'):
            user_ids.add(t['created_by'])
        if t.get('company_id'):
            company_ids.add(t['company_id'])

    user_map = {}
    if user_ids:
        users = await db.users.find({'id': {'$in': list(user_ids)}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(1000)
        user_map = {u['id']: u.get('name', '') for u in users}

    company_map = {}
    if company_ids:
        companies = await db.companies.find({'id': {'$in': list(company_ids)}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(1000)
        company_map = {c['id']: c.get('name', '') for c in companies}

    return user_map, company_map


async def _build_location_maps(db, tasks):
    floor_ids = set()
    unit_ids = set()
    building_ids = set()
    for t in tasks:
        if t.get('floor_id'):
            floor_ids.add(t['floor_id'])
        if t.get('unit_id'):
            unit_ids.add(t['unit_id'])
        if t.get('building_id'):
            building_ids.add(t['building_id'])

    floor_map = {}
    if floor_ids:
        floors = await db.floors.find({'id': {'$in': list(floor_ids)}}, {'_id': 0, 'id': 1, 'name': 1, 'display_label': 1}).to_list(1000)
        floor_map = {f['id']: f.get('display_label') or f.get('name', '') for f in floors}

    unit_map = {}
    spare_tiles_map = {}
    if unit_ids:
        units = await db.units.find({'id': {'$in': list(unit_ids)}}, {'_id': 0, 'id': 1, 'unit_no': 1, 'display_label': 1, 'spare_tiles': 1}).to_list(100000)
        unit_map = {u['id']: u.get('display_label') or u.get('unit_no', '') for u in units}
        spare_tiles_map = {u['id']: u.get('spare_tiles') for u in units if 'spare_tiles' in u}

    building_map = {}
    if building_ids:
        buildings = await db.buildings.find({'id': {'$in': list(building_ids)}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(100)
        building_map = {b['id']: b.get('name', '') for b in buildings}

    return floor_map, unit_map, building_map, spare_tiles_map


EXPORT_SPARE_TILES_BASE_TYPES = [
    'ריצוף יבש',
    'ריצוף מרפסות',
    'חיפוי אמבטיות',
    'ריצוף אמבטיות',
    'חיפוי מטבח',
]


def _spare_tiles_columns(spare_tiles_list):
    if not spare_tiles_list or not isinstance(spare_tiles_list, list):
        return [''] * 7
    by_type = {e['type']: e for e in spare_tiles_list if isinstance(e, dict)}
    base_counts = []
    for bt in EXPORT_SPARE_TILES_BASE_TYPES:
        entry = by_type.get(bt)
        base_counts.append(entry['count'] if entry else '')
    custom_types = [e for e in spare_tiles_list if isinstance(e, dict) and e.get('type') not in EXPORT_SPARE_TILES_BASE_TYPES and (e.get('count', 0) > 0 or e.get('notes'))]
    custom_str = ', '.join(f"{e['type']}: {e['count']}" for e in custom_types) if custom_types else ''
    all_notes = [f"{e['type']}: {e['notes']}" for e in spare_tiles_list if isinstance(e, dict) and e.get('notes')]
    notes_str = ', '.join(all_notes) if all_notes else ''
    return base_counts + [custom_str, notes_str]


def _generate_excel(tasks, project_name, user_map, company_map, floor_map, unit_map, building_map, spare_tiles_map=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if spare_tiles_map is None:
        spare_tiles_map = {}

    wb = Workbook()
    ws = wb.active
    ws.title = 'ליקויים'
    ws.sheet_view.rightToLeft = True

    headers = [
        'מספר ליקוי', 'פרויקט', 'בניין', 'קומה', 'דירה', 'תחום',
        'כותרת', 'תיאור', 'סטטוס', 'חברה/קבלן', 'תאריך יצירה',
        'תאריך עדכון', 'חוסם מסירה', 'מספר תמונות', 'קישורי תמונות',
        'ספייר: ריצוף יבש', 'ספייר: ריצוף מרפסות', 'ספייר: חיפוי אמבטיות',
        'ספייר: ריצוף אמבטיות', 'ספייר: חיפוי מטבח', 'ספייר: סוגים נוספים', 'ספייר: הערות',
    ]

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    header_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    cell_align = Alignment(horizontal='right', vertical='top', wrap_text=True)
    cell_font = Font(name='Arial', size=10)

    for row_idx, task in enumerate(tasks, 2):
        status = task.get('status', 'open')
        is_blocking = status in BLOCKING_STATUSES

        image_links = task.get('_resolved_image_links') or []
        image_count = task.get('attachments_count', 0) or len(image_links)
        links_text = ', '.join(f'תמונה {i+1}' for i in range(len(image_links))) if image_links else ''

        unit_spare = spare_tiles_map.get(task.get('unit_id', ''))
        spare_cols = _spare_tiles_columns(unit_spare)

        row_data = [
            task.get('display_number') or task.get('short_ref', ''),
            project_name,
            building_map.get(task.get('building_id', ''), ''),
            floor_map.get(task.get('floor_id', ''), ''),
            unit_map.get(task.get('unit_id', ''), ''),
            CATEGORY_LABEL.get(task.get('category', ''), task.get('category', '')),
            task.get('title', ''),
            task.get('description', ''),
            STATUS_LABEL.get(status, status),
            company_map.get(task.get('company_id', ''), ''),
            _format_datetime(task.get('created_at')),
            _format_datetime(task.get('updated_at')),
            'כן' if is_blocking else 'לא',
            image_count,
            links_text,
        ] + spare_cols

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        if image_links:
            link_cell = ws.cell(row=row_idx, column=15)
            link_font = Font(name='Arial', size=10, color='0563C1', underline='single')
            link_cell.value = 'תמונה 1'
            link_cell.hyperlink = image_links[0]
            link_cell.font = link_font
            if len(image_links) > 1:
                img_start_col = len(headers) + 1
                for extra_idx, img_url in enumerate(image_links[1:]):
                    col = img_start_col + extra_idx
                    label = f'תמונה {extra_idx + 2}'
                    c = ws.cell(row=row_idx, column=col, value=label)
                    c.hyperlink = img_url
                    c.font = link_font
                    c.alignment = cell_align
                    c.border = thin_border
                    if row_idx == 2 or not ws.cell(row=1, column=col).value:
                        hdr = ws.cell(row=1, column=col, value=label)
                        hdr.font = header_font
                        hdr.fill = header_fill
                        hdr.alignment = header_align
                        hdr.border = thin_border

    col_widths = [12, 15, 12, 10, 10, 14, 25, 30, 14, 18, 16, 16, 12, 10, 14, 14, 14, 14, 14, 14, 20, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tasks) + 1}"

    import os
    from openpyxl.drawing.image import Image as XlImage
    from PIL import Image as PILImage

    footer_row = len(tasks) + 4
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer_cell = ws.cell(row=footer_row, column=1, value='נוצר באמצעות BrikOps | brikops.com')
    footer_cell.font = Font(name='Arial', size=9, color='888888', italic=True)
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')

    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            pil = PILImage.open(logo_path)
            if pil.width > 200:
                ratio = 200 / pil.width
                pil = pil.resize((200, int(pil.height * ratio)), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format='PNG')
            buf.seek(0)
            img = XlImage(buf)
            img.width = 80
            img.height = 30
            ws.add_image(img, f'A{footer_row - 1}')
        except Exception as e:
            logger.warning(f"[EXCEL] Failed to embed logo: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


MAX_IMAGES_PER_DEFECT = 2
PDF_IMAGE_MAX_WIDTH_CM = 12
PDF_IMAGE_MAX_HEIGHT_CM = 7
PDF_IMAGE_REDUCED_WIDTH_CM = 7
PDF_IMAGE_REDUCED_HEIGHT_CM = 4.5


ALLOWED_IMAGE_HOSTS = {
    'brikops-prod-files.s3.eu-central-1.amazonaws.com',
    'brikops-prod-files.s3.amazonaws.com',
    's3.eu-central-1.amazonaws.com',
    's3.amazonaws.com',
}


def _is_safe_image_url(url):
    from urllib.parse import urlparse
    try:
        parsed = urlparse(url)
        host = parsed.hostname or ''
        if host in ALLOWED_IMAGE_HOSTS:
            return True
        if host.endswith('.amazonaws.com'):
            return True
        return False
    except Exception:
        return False


def _image_to_jpeg_buf(img_data_bytes):
    from PIL import Image as PILImage
    img = PILImage.open(io.BytesIO(img_data_bytes))
    img.verify()
    img = PILImage.open(io.BytesIO(img_data_bytes))
    if img.mode not in ('RGB',):
        img = img.convert('RGB')
    max_px = 800
    if img.width > max_px or img.height > max_px:
        img.thumbnail((max_px, max_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=75, optimize=True)
    buf.seek(0)
    return buf


_UPLOADS_BASE = None


def _get_uploads_base():
    global _UPLOADS_BASE
    if _UPLOADS_BASE is None:
        from pathlib import Path
        _UPLOADS_BASE = Path(__file__).parent.parent / 'uploads'
    return _UPLOADS_BASE


def _fetch_local_image(url):
    from pathlib import Path
    if not url.startswith('/api/uploads/'):
        return None
    rel = url[len('/api/uploads/'):]
    if not rel or '..' in rel or rel.startswith('/'):
        logger.warning(f'[PDF] Rejected suspicious local path: {url[:80]}')
        return None
    base = _get_uploads_base().resolve()
    file_path = (base / rel).resolve()
    try:
        file_path.relative_to(base)
    except ValueError:
        logger.warning(f'[PDF] Path traversal blocked: {url[:80]}')
        return None
    if not file_path.is_file():
        logger.warning(f'[PDF] Local file not found: {url[:80]}')
        return None
    raw = file_path.read_bytes()
    return _image_to_jpeg_buf(raw)


def _fetch_image_for_pdf(url):
    import requests as http_requests
    try:
        if url.startswith('/api/uploads/'):
            return _fetch_local_image(url)
        if not url.startswith('http'):
            logger.warning(f'[PDF] Skipping non-HTTP/non-local URL: {url[:80]}')
            return None
        if not _is_safe_image_url(url):
            logger.warning(f'[PDF] Blocked unsafe image URL: {url[:80]}')
            return None
        resp = http_requests.get(url, timeout=8)
        if resp.status_code != 200:
            logger.warning(f'[PDF] Image fetch failed: {resp.status_code} for {url[:80]}')
            return None
        return _image_to_jpeg_buf(resp.content)
    except Exception as e:
        logger.warning(f'[PDF] Image processing failed: {e} for {url[:80]}')
        return None


def _build_image_flowable(img_buf, max_w_cm, max_h_cm, caption_text, style_caption, heb_fn):
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Table, TableStyle, Image as RLImage, Paragraph
    from reportlab.lib import colors
    from PIL import Image as PILImage

    try:
        pil_img = PILImage.open(io.BytesIO(img_buf.getvalue()))
        iw, ih = pil_img.size
        max_w = max_w_cm * cm
        max_h = max_h_cm * cm
        scale = min(max_w / iw, max_h / ih, 1.0)
        draw_w = iw * scale
        draw_h = ih * scale
        img_buf.seek(0)
        rl_img = RLImage(img_buf, width=draw_w, height=draw_h)
        rl_img.hAlign = 'CENTER'
        caption_para = Paragraph(heb_fn(caption_text), style_caption)
        img_with_caption = Table(
            [[rl_img], [caption_para]],
            colWidths=[draw_w],
            rowHeights=[draw_h, None],
        )
        img_with_caption.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 0),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 1 * mm),
            ('TOPPADDING', (0, 1), (-1, 1), 0),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        img_with_caption.hAlign = 'CENTER'
        return img_with_caption
    except Exception as e:
        logger.warning(f'[PDF] Image render failed: {e}')
        return None


def _build_defect_card(task, doc_width, page_height, heb_fn, styles, lookup_maps,
                       image_size='normal', cached_images=None):
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        Paragraph, Spacer, Table, TableStyle, KeepTogether
    )

    user_map, company_map, floor_map, unit_map, building_map = lookup_maps
    AMBER_LIGHT = colors.HexColor('#FEF3C7')
    SLATE_200 = colors.HexColor('#E2E8F0')
    WHITE = colors.white
    CARD_PAD = 6 * mm

    if image_size == 'reduced':
        img_max_w = PDF_IMAGE_REDUCED_WIDTH_CM
        img_max_h = PDF_IMAGE_REDUCED_HEIGHT_CM
    else:
        img_max_w = PDF_IMAGE_MAX_WIDTH_CM
        img_max_h = PDF_IMAGE_MAX_HEIGHT_CM

    inner_width = doc_width - 2 * CARD_PAD
    card_content = []

    status = task.get('status', 'open')
    is_blocking = status in BLOCKING_STATUSES
    title_text = task.get('title', '')
    number = task.get('display_number') or task.get('short_ref', '')
    prefix = f'#{number} ' if number else ''

    status_color = colors.HexColor(STATUS_COLORS.get(status, '#64748B'))
    header_bg = colors.HexColor('#DC2626') if is_blocking else AMBER_LIGHT
    header_text_color = colors.white if is_blocking else colors.HexColor('#334155')

    style_header_dynamic = styles['defect_title'].__class__(
        'PDFDefectTitleDynamic',
        parent=styles['defect_title'],
        textColor=header_text_color,
    )

    blocking_badge = ''
    if is_blocking:
        blocking_badge = '  ◆ חוסם מסירה'

    header_para = Paragraph(heb_fn(f'{prefix}{title_text}{blocking_badge}'), style_header_dynamic)
    header_tbl = Table([[header_para]], colWidths=[inner_width])
    header_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), header_bg),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
        ('LEFTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('ROUNDEDCORNERS', [2 * mm, 2 * mm, 0, 0]),
    ]))
    card_content.append(header_tbl)

    status_label_text = STATUS_LABEL.get(status, status)
    status_badge_style = styles['field_value'].__class__(
        'PDFStatusBadge',
        parent=styles['field_value'],
        textColor=colors.white,
        fontSize=8,
        leading=11,
    )
    status_badge_para = Paragraph(heb_fn(status_label_text), status_badge_style)
    status_badge_tbl = Table([[status_badge_para]], colWidths=[None])
    status_badge_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), status_color),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 1.5 * mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5 * mm),
        ('LEFTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('ROUNDEDCORNERS', [1.5 * mm, 1.5 * mm, 1.5 * mm, 1.5 * mm]),
    ]))
    status_badge_tbl.hAlign = 'RIGHT'

    fields = [
        ('תחום', CATEGORY_LABEL.get(task.get('category', ''), task.get('category', ''))),
        ('בניין', building_map.get(task.get('building_id', ''), '')),
        ('קומה', floor_map.get(task.get('floor_id', ''), '')),
        ('דירה', unit_map.get(task.get('unit_id', ''), '')),
        ('חברה/קבלן', company_map.get(task.get('company_id', ''), '')),
        ('תאריך יצירה', _format_datetime(task.get('created_at'))),
        ('תאריך עדכון', _format_datetime(task.get('updated_at'))),
    ]
    desc = task.get('description', '')
    if desc:
        fields.insert(1, ('תיאור', desc))

    card_content.append(Spacer(1, 2 * mm))
    card_content.append(status_badge_tbl)

    field_rows = []
    for label, value in fields:
        if not value:
            continue
        field_rows.append([
            Paragraph(heb_fn(str(value)), styles['field_value']),
            Paragraph(heb_fn(label), styles['field_label']),
        ])

    if field_rows:
        details_tbl = Table(field_rows, colWidths=[inner_width * 0.7, inner_width * 0.3])
        details_tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        card_content.append(Spacer(1, 2 * mm))
        card_content.append(details_tbl)

    image_links = task.get('_resolved_image_links') or []
    rendered_images = 0
    fetched_bufs = cached_images if cached_images is not None else {}
    if image_links:
        card_content.append(Spacer(1, 3 * mm))
        for img_idx, img_url in enumerate(image_links[:MAX_IMAGES_PER_DEFECT]):
            if img_url in fetched_bufs:
                img_buf = fetched_bufs[img_url]
                if img_buf:
                    img_buf.seek(0)
            else:
                img_buf = _fetch_image_for_pdf(img_url)
                fetched_bufs[img_url] = img_buf
            if img_buf:
                caption = f'תמונה {img_idx + 1}'
                img_flowable = _build_image_flowable(
                    img_buf, img_max_w, img_max_h,
                    caption, styles['caption'], heb_fn
                )
                if img_flowable:
                    card_content.append(img_flowable)
                    if img_idx < MAX_IMAGES_PER_DEFECT - 1:
                        card_content.append(Spacer(1, 2 * mm))
                    rendered_images += 1

    total_available = len(image_links)
    not_shown = total_available - rendered_images
    if not_shown > 0:
        card_content.append(Spacer(1, 1 * mm))
        card_content.append(Paragraph(
            heb_fn(f'({not_shown} תמונות נוספות לא מוצגות)'), styles['count']))

    card_content.append(Spacer(1, 2 * mm))

    card_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, -1), WHITE),
        ('BOX', (0, 0), (-1, -1), 0.75, SLATE_200),
        ('ROUNDEDCORNERS', [2 * mm, 2 * mm, 2 * mm, 2 * mm]),
        ('TOPPADDING', (0, 0), (-1, -1), CARD_PAD),
        ('BOTTOMPADDING', (0, 0), (-1, -1), CARD_PAD),
        ('LEFTPADDING', (0, 0), (-1, -1), CARD_PAD),
        ('RIGHTPADDING', (0, 0), (-1, -1), CARD_PAD),
    ]
    if is_blocking:
        card_style_cmds.append(('LINEAFTER', (0, 0), (0, -1), 3, colors.HexColor('#DC2626')))

    outer_card = Table(
        [[card_content]],
        colWidths=[doc_width],
        style=TableStyle(card_style_cmds),
    )

    card_w, card_h = outer_card.wrap(doc_width, page_height)
    return outer_card, card_h, rendered_images, card_content, fetched_bufs


def _generate_pdf(tasks, project_name, scope_label, scope_type, user_map, company_map,
                  floor_map, unit_map, building_map, filters_desc=None):
    import os
    import arabic_reshaper
    from bidi.algorithm import get_display
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fonts_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'fonts')
    try:
        pdfmetrics.registerFont(TTFont('Rubik', os.path.join(fonts_dir, 'Rubik-Regular.ttf')))
    except Exception:
        pass

    def heb(text):
        if not text:
            return ''
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)

    SLATE_700 = colors.HexColor('#334155')
    SLATE_500 = colors.HexColor('#64748B')
    AMBER_DARK = colors.HexColor('#B45309')

    style_title = ParagraphStyle('PDFTitle', fontName='Rubik', fontSize=18,
                                  alignment=TA_CENTER, textColor=SLATE_700, leading=24)
    style_subtitle = ParagraphStyle('PDFSubtitle', fontName='Rubik', fontSize=11,
                                     alignment=TA_CENTER, textColor=SLATE_500, leading=15)
    style_filter = ParagraphStyle('PDFFilter', fontName='Rubik', fontSize=9,
                                   alignment=TA_CENTER, textColor=AMBER_DARK, leading=12)
    style_field_label = ParagraphStyle('PDFFieldLabel', fontName='Rubik', fontSize=9,
                                        alignment=TA_RIGHT, textColor=SLATE_500, leading=12)
    style_field_value = ParagraphStyle('PDFFieldValue', fontName='Rubik', fontSize=10,
                                        alignment=TA_RIGHT, textColor=SLATE_700, leading=13)
    style_defect_title = ParagraphStyle('PDFDefectTitle', fontName='Rubik', fontSize=12,
                                         alignment=TA_RIGHT, textColor=SLATE_700, leading=16,
                                         fontWeight='bold')
    style_count = ParagraphStyle('PDFCount', fontName='Rubik', fontSize=9,
                                  alignment=TA_CENTER, textColor=SLATE_500, leading=12)
    style_caption = ParagraphStyle('PDFCaption', fontName='Rubik', fontSize=8,
                                    alignment=TA_CENTER, textColor=SLATE_500, leading=10)

    styles = {
        'defect_title': style_defect_title,
        'field_label': style_field_label,
        'field_value': style_field_value,
        'count': style_count,
        'caption': style_caption,
    }

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    page_height = A4[1] - doc.topMargin - doc.bottomMargin
    elements = []

    BLUE_600 = colors.HexColor('#2563EB')
    GRAY_100 = colors.HexColor('#F3F4F6')
    GRAY_300 = colors.HexColor('#D1D5DB')

    scope_heb = heb('בניין') if scope_type == 'building' else heb('דירה')
    elements.append(Paragraph(heb('דו״ח ליקויים'), style_title))
    elements.append(Spacer(1, 2 * mm))

    blue_line = Table([['']], colWidths=[4 * cm], rowHeights=[0.8 * mm])
    blue_line.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BLUE_600),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    blue_line.hAlign = 'CENTER'
    elements.append(blue_line)
    elements.append(Spacer(1, 4 * mm))

    info_line = f'{project_name} · {scope_heb}: {scope_label}'
    elements.append(Paragraph(heb(info_line), style_subtitle))

    today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    elements.append(Paragraph(heb(f'תאריך: {today_str}'), style_subtitle))
    elements.append(Spacer(1, 3 * mm))

    if filters_desc:
        elements.append(Paragraph(heb(f'סינון: {filters_desc}'), style_filter))
        elements.append(Spacer(1, 2 * mm))

    status_counts = Counter(t.get('status', 'open') for t in tasks)
    blocking_count = sum(1 for t in tasks if t.get('status', 'open') in BLOCKING_STATUSES)

    style_summary_label = ParagraphStyle('PDFSummaryLabel', fontName='Rubik', fontSize=9,
                                          alignment=TA_RIGHT, textColor=SLATE_500, leading=12)
    style_summary_value = ParagraphStyle('PDFSummaryValue', fontName='Rubik', fontSize=11,
                                          alignment=TA_CENTER, textColor=SLATE_700, leading=14)

    summary_cells = []
    summary_labels = []
    total_count_para = Paragraph(heb(str(len(tasks))), style_summary_value)
    total_label_para = Paragraph(heb('סה״כ'), style_summary_label)
    summary_cells.append(total_count_para)
    summary_labels.append(total_label_para)

    open_count = status_counts.get('open', 0) + status_counts.get('reopened', 0)
    if open_count:
        summary_cells.append(Paragraph(heb(str(open_count)), style_summary_value))
        summary_labels.append(Paragraph(heb('פתוחים'), style_summary_label))

    in_progress = sum(status_counts.get(s, 0) for s in IN_PROGRESS_STATUSES)
    if in_progress:
        summary_cells.append(Paragraph(heb(str(in_progress)), style_summary_value))
        summary_labels.append(Paragraph(heb('בביצוע'), style_summary_label))

    closed_count = status_counts.get('closed', 0)
    if closed_count:
        summary_cells.append(Paragraph(heb(str(closed_count)), style_summary_value))
        summary_labels.append(Paragraph(heb('סגורים'), style_summary_label))

    if blocking_count:
        style_blocking_val = ParagraphStyle('PDFBlockingVal', fontName='Rubik', fontSize=11,
                                             alignment=TA_CENTER, textColor=colors.HexColor('#DC2626'), leading=14)
        summary_cells.append(Paragraph(heb(str(blocking_count)), style_blocking_val))
        summary_labels.append(Paragraph(heb('חוסמים'), style_summary_label))

    num_cols = len(summary_cells)
    col_w = doc.width / num_cols if num_cols else doc.width
    summary_tbl = Table([summary_cells, summary_labels], colWidths=[col_w] * num_cols)
    summary_tbl.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), GRAY_100),
        ('BOX', (0, 0), (-1, -1), 0.5, GRAY_300),
        ('ROUNDEDCORNERS', [2 * mm, 2 * mm, 2 * mm, 2 * mm]),
        ('TOPPADDING', (0, 0), (-1, 0), 3 * mm),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 1 * mm),
        ('TOPPADDING', (0, 1), (-1, 1), 0),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 3 * mm),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, GRAY_300),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1, 6 * mm))

    lookup_maps = (user_map, company_map, floor_map, unit_map, building_map)

    for idx, task in enumerate(tasks):
        card, card_h, rendered, card_content, img_cache = _build_defect_card(
            task, doc.width, page_height, heb, styles, lookup_maps,
            image_size='normal'
        )

        if card_h <= page_height:
            elements.append(KeepTogether([card]))
        else:
            card_reduced, card_h_reduced, _, content_reduced, _ = _build_defect_card(
                task, doc.width, page_height, heb, styles, lookup_maps,
                image_size='reduced', cached_images=img_cache
            )
            if card_h_reduced <= page_height:
                elements.append(KeepTogether([card_reduced]))
            else:
                elements.extend(content_reduced)

        if idx < len(tasks) - 1:
            elements.append(Spacer(1, 5 * mm))

    def _add_page_number(canvas_obj, doc_obj):
        from reportlab.lib.units import cm
        page_num = canvas_obj.getPageNumber()
        canvas_obj.saveState()
        canvas_obj.setFont('Rubik', 8)
        canvas_obj.setFillColor(colors.HexColor('#94A3B8'))
        canvas_obj.drawCentredString(doc_obj.pagesize[0] / 2, 1 * cm,
                                      heb(f'{page_num} עמוד'))
        canvas_obj.restoreState()

    doc.build(elements, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    output.seek(0)
    return output


def _build_filters_description(filters: ExportFilters):
    parts = []
    if filters.status and filters.status != 'all':
        status_labels = {'open': 'פתוחים', 'closed': 'סגורים', 'blocking': 'חוסמי מסירה', 'in_progress': 'בטיפול'}
        parts.append(f'סטטוס: {status_labels.get(filters.status, filters.status)}')
    if filters.category and filters.category != 'all':
        parts.append(f'תחום: {CATEGORY_LABEL.get(filters.category, filters.category)}')
    if filters.search:
        parts.append(f'חיפוש: {filters.search}')
    if filters.company and filters.company != 'all':
        parts.append('חברה: מסוננת')
    if filters.assignee and filters.assignee != 'all':
        parts.append('אחראי: מסונן')
    if filters.floor and filters.floor != 'all':
        parts.append('קומה: מסוננת')
    if filters.unit and filters.unit != 'all':
        parts.append('דירה: מסוננת')
    return ' · '.join(parts) if parts else None


@router.post("/defects/export")
async def export_defects(req: ExportRequest, user: dict = Depends(get_current_user)):
    db = get_db()

    if req.scope == 'unit':
        if not req.unit_id:
            raise HTTPException(status_code=400, detail='unit_id required for unit scope')
        unit = await db.units.find_one({'id': req.unit_id}, {'_id': 0})
        if not unit:
            raise HTTPException(status_code=404, detail='Unit not found')
        building = await db.buildings.find_one({'id': unit.get('building_id')}, {'_id': 0}) if unit.get('building_id') else None
        project_id = unit.get('project_id') or (building.get('project_id') if building else None)
        if not project_id:
            raise HTTPException(status_code=400, detail='Could not determine project')
        await _check_project_read_access(user, project_id)
        project = await db.projects.find_one({'id': project_id}, {'_id': 0})

        tasks = await db.tasks.find({'unit_id': req.unit_id, 'archived': {'$ne': True}}, {'_id': 0}).to_list(100000)
        scope_label = unit.get('display_label') or unit.get('unit_no', '')
        filename_scope = f"apartment_{scope_label}"

    elif req.scope == 'building':
        if not req.building_id:
            raise HTTPException(status_code=400, detail='building_id required for building scope')
        building = await db.buildings.find_one({'id': req.building_id, 'archived': {'$ne': True}}, {'_id': 0})
        if not building:
            raise HTTPException(status_code=404, detail='Building not found')
        project_id = building['project_id']
        await _check_project_read_access(user, project_id)
        project = await db.projects.find_one({'id': project_id}, {'_id': 0})

        tasks = await db.tasks.find({'building_id': req.building_id, 'archived': {'$ne': True}}, {'_id': 0}).to_list(100000)
        scope_label = building.get('name', '')
        filename_scope = f"building_{scope_label}"

    elif req.scope == 'project':
        if not req.project_id:
            raise HTTPException(status_code=400, detail='project_id required for project scope')
        project = await db.projects.find_one({'id': req.project_id, 'archived': {'$ne': True}}, {'_id': 0})
        if not project:
            raise HTTPException(status_code=404, detail='Project not found')

        role = await _get_project_role(user, req.project_id)
        if role not in ('project_manager', 'owner') and user.get('platform_role') != 'super_admin':
            raise HTTPException(status_code=403, detail='אין הרשאה לייצוא נתונים')

        tasks = await db.tasks.find(
            {'project_id': req.project_id, 'archived': {'$ne': True}},
            {'_id': 0}
        ).to_list(100000)
        scope_label = project.get('name', '')
        filename_scope = f"project_{scope_label}"

    else:
        raise HTTPException(status_code=400, detail='Invalid scope. Use "unit", "building", or "project".')

    project_name = project.get('name', '') if project else ''

    filtered = _apply_filters(tasks, req.filters, scope=req.scope)
    def _sort_key(t):
        v = t.get('created_at', '')
        if isinstance(v, datetime):
            return v.isoformat()
        return str(v)
    filtered.sort(key=_sort_key, reverse=True)

    user_map, company_map = await _build_lookup_maps(db, filtered)
    floor_map, unit_map, building_map_loc, spare_tiles_map = await _build_location_maps(db, filtered)

    for t in filtered:
        t['_resolved_image_links'] = await _resolve_image_links(db, t)

    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    safe_scope = filename_scope.replace(' ', '_').replace('/', '_')
    from urllib.parse import quote

    export_format = (req.format or 'excel').lower()
    if export_format not in VALID_EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail=f'Invalid format. Use "excel" or "pdf".')

    if export_format == 'pdf':
        filters_desc = _build_filters_description(req.filters)
        pdf_bytes = _generate_pdf(
            filtered, project_name, scope_label, req.scope,
            user_map, company_map, floor_map, unit_map, building_map_loc,
            filters_desc=filters_desc
        )
        filename = f"defects_{safe_scope}_{today}.pdf"
        ascii_filename = f"defects_{req.scope}_{today}.pdf"
        encoded_filename = quote(filename)
        disposition = f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}"
        logger.info(f"[EXPORT-PDF] user={user.get('id')} scope={req.scope} tasks={len(filtered)} filename={filename}")
        if req.scope == 'project':
            await _audit('project', req.project_id, 'data_export', user.get('id'), {
                'format': export_format,
                'task_count': len(filtered),
            })
        return StreamingResponse(
            pdf_bytes,
            media_type='application/pdf',
            headers={'Content-Disposition': disposition}
        )

    excel_bytes = _generate_excel(
        filtered, project_name, user_map, company_map,
        floor_map, unit_map, building_map_loc, spare_tiles_map
    )

    filename = f"defects_{safe_scope}_{today}.xlsx"
    ascii_filename = f"defects_{req.scope}_{today}.xlsx"
    encoded_filename = quote(filename)
    disposition = f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}"

    logger.info(f"[EXPORT] user={user.get('id')} scope={req.scope} tasks={len(filtered)} filename={filename}")

    if req.scope == 'project':
        await _audit('project', req.project_id, 'data_export', user.get('id'), {
            'format': export_format,
            'task_count': len(filtered),
        })

    return StreamingResponse(
        excel_bytes,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': disposition}
    )
