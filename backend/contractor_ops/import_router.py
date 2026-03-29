import uuid
import csv
import io
import re
import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel

from contractor_ops.router import get_db, get_current_user, _is_super_admin, _audit, _now
from contractor_ops.phone_utils import clean_phone_for_import, normalize_israeli_phone

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/projects/{project_id}/import/g4", tags=["import-g4"])

STATIC_DIR = Path(__file__).resolve().parent.parent / "static"
TEMPLATE_PATH = STATIC_DIR / "g4_template.xlsx"

MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_ROWS = 1000

HEADER_ALIASES = {
    'building': ['בניין', 'בנין', 'building', 'בניין מס', 'מס בניין'],
    'floor': ['קומה', 'floor', 'מס קומה'],
    'apartment': ['דירה', 'apartment', 'מס דירה', 'דירה מס', 'דירה מספר', 'יחידה'],
    'tenant_name': ['שם', 'שם חשבון', 'שם רוכש', 'שם דייר', 'tenant', 'name', 'שם מלא'],
    'tenant_id': ['ת"ז', 'תז', 'ת.ז', 'ת.ז.', 'ת״ז', 'id', 'מספר זהות', 'תעודת זהות'],
    'tenant_phone': ['טלפון', 'נייד', 'טלפון נייד', 'phone', 'mobile', 'סלולרי', 'פלאפון', 'טל', 'טל נייד'],
    'tenant_phone_2': ['טלפון נוסף', 'טלפון שני', 'נייד שני', 'טלפון נייד שני', 'phone2'],
    'tenant_email': ['אימייל', 'email', 'דואל', 'מייל', 'דוא"ל'],
    'tenant_2_name': ['שם נוסף', 'שם שני', 'שם בן/בת זוג', 'שם 2'],
    'tenant_2_id': ['ת"ז נוסף', 'ת"ז 2', 'תז שני', 'ת״ז נוסף'],
    'handover_date': ['תאריך מסירה', 'תאריך מסירת דירה', 'מסירה', 'handover date'],
    'unit_type': ['אפיון', 'תאור אפיון', 'סוג דירה', 'אפיון דירה', 'type'],
}

REQUIRED_COLUMNS = {"apartment", "tenant_name"}


def _normalize_header(h):
    if h is None:
        return ""
    return re.sub(r'\s+', ' ', str(h).strip()).lower()


def _detect_columns(header_cells: list) -> dict:
    col_map = {}
    unmatched = []
    for col_idx, raw in enumerate(header_cells):
        norm = _normalize_header(raw)
        if not norm:
            continue
        matched_key = None
        for key, aliases in HEADER_ALIASES.items():
            if key in col_map:
                continue
            for alias in aliases:
                if norm == alias.lower():
                    matched_key = key
                    break
            if matched_key:
                break
        if not matched_key:
            norm_tokens = set(re.split(r'[\s/\-_]+', norm))
            candidate_groups = []
            for key, aliases in HEADER_ALIASES.items():
                if key in col_map:
                    continue
                for alias in aliases:
                    alias_lower = alias.lower()
                    alias_tokens = set(re.split(r'[\s/\-_]+', alias_lower))
                    if len(alias_tokens) >= 2 and alias_tokens.issubset(norm_tokens):
                        candidate_groups.append(key)
                        break
            if len(candidate_groups) == 1:
                matched_key = candidate_groups[0]

        if matched_key:
            col_map[matched_key] = {"index": col_idx, "header": str(raw).strip()}
        else:
            unmatched.append(str(raw).strip())
    return {"columns": col_map, "unmatched": unmatched}


def _detect_header_row_xlsx(content: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    best_row = None
    best_result = None
    best_count = 0
    for row_idx in range(1, 4):
        cells = []
        for row in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
            cells = list(row)
            break
        if not cells:
            continue
        result = _detect_columns(cells)
        count = len(result["columns"])
        if count > best_count:
            best_count = count
            best_result = result
            best_row = row_idx
    wb.close()
    if best_count < 2:
        return None, None
    return best_row, best_result


def _detect_header_row_csv(text: str):
    reader = csv.reader(io.StringIO(text))
    rows_to_try = []
    for i, row in enumerate(reader):
        rows_to_try.append(row)
        if i >= 2:
            break
    best_row = None
    best_result = None
    best_count = 0
    for idx, cells in enumerate(rows_to_try):
        result = _detect_columns(cells)
        count = len(result["columns"])
        if count > best_count:
            best_count = count
            best_result = result
            best_row = idx + 1
    if best_count < 2:
        return None, None
    return best_row, best_result


def _parse_xlsx_smart(content: bytes, header_row: int, col_map: dict) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    data_start = header_row + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        if row_idx > MAX_ROWS + data_start:
            break
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        parsed = {}
        for key, info in col_map.items():
            ci = info["index"]
            val = row[ci] if ci < len(row) else None
            parsed[key] = val
        parsed["source_row"] = row_idx
        rows.append(parsed)
    wb.close()
    return rows


def _parse_csv_smart(text: str, header_row: int, col_map: dict) -> list:
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row_idx_0, row in enumerate(reader):
        row_idx = row_idx_0 + 1
        if row_idx <= header_row:
            continue
        if row_idx > MAX_ROWS + header_row + 1:
            break
        if not row or all(cell.strip() == "" for cell in row):
            continue
        parsed = {}
        for key, info in col_map.items():
            ci = info["index"]
            val = row[ci].strip() if ci < len(row) else ""
            parsed[key] = val
        parsed["source_row"] = row_idx
        rows.append(parsed)
    return rows


def _parse_date(raw_val) -> tuple:
    if raw_val is None or (isinstance(raw_val, str) and not raw_val.strip()):
        return None, None

    if isinstance(raw_val, datetime):
        return raw_val.strftime("%Y-%m-%d"), None

    if isinstance(raw_val, (int, float)):
        try:
            serial = int(raw_val)
            if 1 < serial < 200000:
                base = datetime(1899, 12, 30)
                dt = base + timedelta(days=serial)
                return dt.strftime("%Y-%m-%d"), None
        except (ValueError, OverflowError):
            pass
        return None, "תאריך לא תקין"

    s = str(raw_val).strip()
    if not s:
        return None, None

    for sep in ('.', '/'):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    dt = datetime(y, m, d)
                    return dt.strftime("%Y-%m-%d"), None
                except (ValueError, OverflowError):
                    return None, "תאריך לא תקין"

    if '-' in s:
        parts = s.split('-')
        if len(parts) == 3 and len(parts[0]) == 4:
            try:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                dt = datetime(y, m, d)
                return dt.strftime("%Y-%m-%d"), None
            except (ValueError, OverflowError):
                return None, "תאריך לא תקין"

    return None, "תאריך לא תקין"


def _str_val(val):
    if val is None:
        return ""
    return str(val).strip()


def _validate_row(row: dict, detected_keys: set) -> dict:
    errors = []
    warnings = []

    apt = _str_val(row.get("apartment"))
    if not apt:
        errors.append("דירה חסרה")

    name = _str_val(row.get("tenant_name"))
    if not name:
        errors.append("שם רוכש חסר")

    id_number = _str_val(row.get("tenant_id"))
    if id_number and not re.match(r'^\d{9}$', id_number):
        warnings.append("ת\"ז לא תקין (9 ספרות)")

    for phone_key in ('tenant_phone', 'tenant_phone_2'):
        if phone_key not in detected_keys:
            continue
        raw_phone = row.get(phone_key)
        cleaned = clean_phone_for_import(raw_phone)
        if cleaned:
            try:
                result = normalize_israeli_phone(cleaned, allow_landline=True)
                row[phone_key] = result["phone_e164"]
            except (ValueError, Exception):
                warnings.append(f"טלפון לא תקין ({phone_key})" if phone_key == 'tenant_phone_2' else "טלפון לא תקין")
                row[phone_key] = cleaned
        else:
            row[phone_key] = ""

    email = _str_val(row.get("tenant_email"))
    if email and "@" not in email:
        warnings.append("מייל לא תקין")

    t2_id = _str_val(row.get("tenant_2_id"))
    if t2_id and not re.match(r'^\d{9}$', t2_id):
        warnings.append("ת\"ז רוכש 2 לא תקין")

    if 'handover_date' in detected_keys:
        date_val, date_warn = _parse_date(row.get("handover_date"))
        row["handover_date"] = date_val
        if date_warn:
            warnings.append(date_warn)

    for key in row:
        if key in ('source_row', 'handover_date', 'tenant_phone', 'tenant_phone_2'):
            continue
        if isinstance(row[key], (int, float)):
            row[key] = str(row[key])
        elif row[key] is None:
            row[key] = ""
        else:
            row[key] = str(row[key]).strip()

    return {
        "row": row,
        "errors": errors,
        "warnings": warnings,
        "valid": len(errors) == 0,
    }


def _normalize_building_name(name: str) -> str:
    s = name.strip()
    s = re.sub(r'^בניין\s*', '', s)
    s = re.sub(r'^בנין\s*', '', s)
    return s.strip()


def _normalize_unit_no(val: str) -> str:
    return val.lstrip('0') or val


async def _cross_reference(rows: list, project_id: str, detected_keys: set):
    db = get_db()
    buildings = await db.buildings.find(
        {'project_id': project_id}, {'_id': 0}
    ).to_list(500)

    units = await db.units.find(
        {'project_id': project_id}, {'_id': 0}
    ).to_list(10000)

    floors = await db.floors.find(
        {'project_id': project_id}, {'_id': 0}
    ).to_list(5000)
    floor_map = {f['id']: f for f in floors}

    building_lookup = {}
    for b in buildings:
        name_norm = _normalize_building_name(b.get('name', ''))
        building_lookup[name_norm.lower()] = b
        code = (b.get('code') or '').strip()
        if code:
            building_lookup[code.lower()] = b

    units_by_building = {}
    for u in units:
        bid = u.get('building_id', '')
        units_by_building.setdefault(bid, []).append(u)

    has_building_col = 'building' in detected_keys
    all_units = units

    overwrite_count = 0

    for v in rows:
        if not v["valid"]:
            v["match"] = {"status": "error"}
            continue

        r = v["row"]
        matched_building = None

        if has_building_col:
            bname_raw = _str_val(r.get("building"))
            if bname_raw:
                bname_norm = _normalize_building_name(bname_raw).lower()
                matched_building = building_lookup.get(bname_norm)
                if not matched_building:
                    for b in buildings:
                        if b.get('name', '').strip().lower() == bname_raw.lower():
                            matched_building = b
                            break
                if not matched_building:
                    v["match"] = {"status": "building_not_found"}
                    v["errors"].append(f"בניין \"{bname_raw}\" לא נמצא בפרויקט")
                    v["valid"] = False
                    continue

        apt_raw = _str_val(r.get("apartment"))
        apt_norm = _normalize_unit_no(apt_raw)

        search_units = units_by_building.get(matched_building['id'], []) if matched_building else all_units

        matched_units = []
        for u in search_units:
            u_no = u.get('unit_no', '')
            if u_no == apt_raw or _normalize_unit_no(u_no) == apt_norm:
                matched_units.append(u)

        if len(matched_units) == 1:
            mu = matched_units[0]
            mu_building = None
            for b in buildings:
                if b['id'] == mu.get('building_id'):
                    mu_building = b
                    break
            mu_floor = floor_map.get(mu.get('floor_id'))

            match_info = {
                "status": "matched",
                "building_id": mu.get('building_id', ''),
                "building_name": mu_building['name'] if mu_building else '',
                "floor_id": mu.get('floor_id', ''),
                "unit_id": mu['id'],
                "unit_label": mu.get('display_label') or mu.get('unit_no', ''),
            }

            if 'floor' in detected_keys and mu_floor:
                row_floor = _str_val(r.get("floor"))
                floor_name = mu_floor.get('name', '')
                floor_num = str(mu_floor.get('floor_number', ''))
                if row_floor and row_floor != floor_name and row_floor != floor_num:
                    v["warnings"].append(f"קומה באקסל \"{row_floor}\" שונה ממה שבמערכת \"{floor_name}\"")

            existing = await db.unit_tenant_data.find_one({
                "project_id": project_id,
                "unit_id": mu['id'],
            })
            if not existing:
                existing = await db.unit_tenant_data.find_one({
                    "project_id": project_id,
                    "building_name": match_info["building_name"],
                    "apartment_number": apt_raw,
                })
            if existing:
                match_info["status"] = "overwrite"
                v["overwrite"] = True
                overwrite_count += 1
            else:
                v["overwrite"] = False

            v["match"] = match_info

        elif len(matched_units) > 1 and not matched_building:
            bids = set(u.get('building_id') for u in matched_units)
            if len(bids) > 1:
                amb_names = []
                for bid in bids:
                    for b in buildings:
                        if b['id'] == bid:
                            amb_names.append(b.get('name', bid))
                            break
                v["match"] = {
                    "status": "ambiguous_unit",
                    "ambiguous_buildings": amb_names,
                }
                v["errors"].append(f"דירה {apt_raw} נמצאה ב-{len(bids)} בניינים — יש לציין בניין")
                v["valid"] = False
            else:
                mu = matched_units[0]
                mu_building = None
                for b in buildings:
                    if b['id'] == mu.get('building_id'):
                        mu_building = b
                        break
                mu_floor = floor_map.get(mu.get('floor_id'))
                match_info = {
                    "status": "matched",
                    "building_id": mu.get('building_id', ''),
                    "building_name": mu_building['name'] if mu_building else '',
                    "floor_id": mu.get('floor_id', ''),
                    "unit_id": mu['id'],
                    "unit_label": mu.get('display_label') or mu.get('unit_no', ''),
                }
                existing = await db.unit_tenant_data.find_one({
                    "project_id": project_id,
                    "unit_id": mu['id'],
                })
                if existing:
                    match_info["status"] = "overwrite"
                    v["overwrite"] = True
                    overwrite_count += 1
                else:
                    v["overwrite"] = False
                v["match"] = match_info
        else:
            v["match"] = {"status": "unit_not_found"}
            v["warnings"].append(f"דירה {apt_raw} לא נמצאה" + (f" בבניין {matched_building['name']}" if matched_building else " בפרויקט"))
            v["overwrite"] = False

    return overwrite_count


async def _check_import_access(user: dict, project_id: str):
    if _is_super_admin(user):
        return
    db = get_db()
    membership = await db.project_memberships.find_one(
        {"user_id": user["id"], "project_id": project_id}
    )
    if not membership:
        raise HTTPException(status_code=403, detail="אין לך גישה לפרויקט זה")
    allowed = {"owner", "project_manager", "management_team"}
    if membership.get("role") not in allowed:
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")


@router.get("/template")
async def download_template(
    project_id: str,
    user: dict = Depends(get_current_user),
):
    await _check_import_access(user, project_id)

    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="קובץ תבנית לא נמצא")

    return FileResponse(
        path=str(TEMPLATE_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="g4_template.xlsx"',
        },
    )


@router.post("/preview")
async def preview_import(
    project_id: str,
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    await _check_import_access(user, project_id)

    if file.size and file.size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="קובץ גדול מדי (מקסימום 5MB)")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="קובץ גדול מדי (מקסימום 5MB)")

    filename = (file.filename or "").lower()
    is_xlsx = filename.endswith(".xlsx")
    is_csv = filename.endswith(".csv")

    if not is_xlsx and not is_csv:
        raise HTTPException(status_code=400, detail="פורמט לא נתמך. יש להעלות xlsx או csv")

    try:
        if is_xlsx:
            header_row, detection = _detect_header_row_xlsx(content)
        else:
            text = content.decode("utf-8-sig")
            header_row, detection = _detect_header_row_csv(text)
    except Exception as e:
        logger.error(f"Failed to detect headers: {e}")
        raise HTTPException(status_code=400, detail="שגיאה בקריאת כותרות הקובץ")

    if header_row is None or detection is None:
        raise HTTPException(status_code=400, detail="לא זוהו עמודות חובה: דירה ושם רוכש")

    col_map = detection["columns"]
    unmatched_columns = detection["unmatched"]
    detected_keys = set(col_map.keys())

    if not (REQUIRED_COLUMNS <= detected_keys):
        missing = []
        if "apartment" not in detected_keys:
            missing.append("דירה")
        if "tenant_name" not in detected_keys:
            missing.append("שם רוכש")
        raise HTTPException(status_code=400, detail=f"לא זוהו עמודות חובה: {', '.join(missing)}")

    try:
        if is_xlsx:
            raw_rows = _parse_xlsx_smart(content, header_row, col_map)
        else:
            raw_rows = _parse_csv_smart(text, header_row, col_map)
    except Exception as e:
        logger.error(f"Failed to parse file: {e}")
        raise HTTPException(status_code=400, detail="שגיאה בקריאת הקובץ")

    if not raw_rows:
        raise HTTPException(status_code=400, detail="הקובץ ריק או לא מכיל שורות נתונים")

    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"הקובץ מכיל יותר מ-{MAX_ROWS} שורות")

    validated = [_validate_row(r, detected_keys) for r in raw_rows]

    overwrite_count = await _cross_reference(validated, project_id, detected_keys)

    seen_keys = {}
    for v in validated:
        if not v["valid"]:
            continue
        r = v["row"]
        match = v.get("match", {})
        unit_id = match.get("unit_id")
        if unit_id:
            key = (project_id, unit_id)
        else:
            key = (_str_val(r.get("building", "")), _str_val(r.get("apartment", "")))
        if key in seen_keys:
            v["warnings"].append("כפילות בקובץ")
            seen_keys[key]["warnings"].append("כפילות בקובץ")
        else:
            seen_keys[key] = v

    valid_count = sum(1 for v in validated if v["valid"])
    error_count = sum(1 for v in validated if not v["valid"])
    matched_count = sum(1 for v in validated if v.get("match", {}).get("status") in ("matched", "overwrite"))
    not_found_count = sum(1 for v in validated if v.get("match", {}).get("status") in ("unit_not_found",))

    detected_columns_response = {}
    for key, info in col_map.items():
        detected_columns_response[key] = {"index": info["index"], "header": info["header"]}

    return {
        "rows": validated,
        "valid_count": valid_count,
        "error_count": error_count,
        "overwrite_count": overwrite_count,
        "matched_count": matched_count,
        "not_found_count": not_found_count,
        "total": len(validated),
        "detected_columns": detected_columns_response,
        "header_row": header_row,
        "unmatched_columns": unmatched_columns,
    }


class ImportRow(BaseModel):
    building_name: str = ""
    building_id: str = ""
    floor: str = ""
    floor_id: str = ""
    apartment_number: str
    unit_id: str = ""
    tenant_name: str
    tenant_id_number: str = ""
    tenant_phone: str = ""
    tenant_phone_2: str = ""
    tenant_email: str = ""
    tenant_2_name: str = ""
    tenant_2_id_number: str = ""
    handover_date: Optional[str] = None
    unit_type: str = ""
    source_row: Optional[int] = None


class ExecuteImportRequest(BaseModel):
    rows: List[ImportRow]


@router.post("/execute")
async def execute_import(
    project_id: str,
    body: ExecuteImportRequest,
    user: dict = Depends(get_current_user),
):
    await _check_import_access(user, project_id)

    if not body.rows:
        raise HTTPException(status_code=400, detail="אין שורות לייבוא")

    if len(body.rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"יותר מ-{MAX_ROWS} שורות")

    db = get_db()
    batch_id = str(uuid.uuid4())
    ts = _now()

    imported = 0
    skipped = 0
    errors = []

    submitted_unit_ids = set()
    for row in body.rows:
        uid = (row.unit_id or "").strip()
        if uid:
            submitted_unit_ids.add(uid)

    valid_unit_ids = {}
    if submitted_unit_ids:
        cursor = db.units.find(
            {"project_id": project_id, "id": {"$in": list(submitted_unit_ids)}},
            {"id": 1, "building_id": 1, "_id": 0}
        )
        async for u in cursor:
            valid_unit_ids[u["id"]] = u

    all_units = []
    cursor = db.units.find({"project_id": project_id}, {"id": 1, "unit_no": 1, "building_id": 1, "_id": 0})
    async for u in cursor:
        all_units.append(u)

    buildings = []
    cursor = db.buildings.find({"project_id": project_id}, {"id": 1, "name": 1, "_id": 0})
    async for b in cursor:
        buildings.append(b)
    building_by_id = {b["id"]: b for b in buildings}
    building_by_name = {}
    for b in buildings:
        building_by_name[b["name"].strip().lower()] = b
        building_by_name[_normalize_building_name(b["name"]).lower()] = b

    units_by_building = {}
    for u in all_units:
        bid = u.get("building_id", "")
        units_by_building.setdefault(bid, []).append(u)

    for row in body.rows:
        row_data = row.dict()
        row_errors = []

        apt_number = (row_data.get("apartment_number") or "").strip()
        if not apt_number:
            row_errors.append("דירה חסרה")
        if not (row_data.get("tenant_name") or "").strip():
            row_errors.append("שם רוכש חסר")

        submitted_uid = (row_data.get("unit_id") or "").strip()

        resolved_uid = None
        resolved_building_id = None

        if submitted_uid:
            if submitted_uid in valid_unit_ids:
                resolved_uid = submitted_uid
                resolved_building_id = valid_unit_ids[submitted_uid].get("building_id")
            else:
                row_errors.append("יחידה לא נמצאה בפרויקט")
        elif not row_errors:
            bname = (row_data.get("building_name") or "").strip()
            if bname:
                matched_building = building_by_name.get(bname.lower()) or building_by_name.get(_normalize_building_name(bname).lower())
                if matched_building:
                    search_units = units_by_building.get(matched_building["id"], [])
                    apt_norm = _normalize_unit_no(apt_number)
                    found = [u for u in search_units if u.get("unit_no") == apt_number or _normalize_unit_no(u.get("unit_no", "")) == apt_norm]
                    if len(found) == 1:
                        resolved_uid = found[0]["id"]
                        resolved_building_id = matched_building["id"]
                    elif len(found) > 1:
                        row_errors.append("כפילות יחידות — לא ניתן לשייך")
                    else:
                        row_errors.append("דירה לא נמצאה בפרויקט")
                else:
                    row_errors.append("בניין לא נמצא בפרויקט")
            else:
                apt_norm = _normalize_unit_no(apt_number)
                found = [u for u in all_units if u.get("unit_no") == apt_number or _normalize_unit_no(u.get("unit_no", "")) == apt_norm]
                if len(found) == 1:
                    resolved_uid = found[0]["id"]
                    resolved_building_id = found[0].get("building_id")
                elif len(found) > 1:
                    row_errors.append("כפילות יחידות בפרויקט — יש לציין בניין")
                else:
                    row_errors.append("דירה לא נמצאה בפרויקט")

        if row_errors:
            skipped += 1
            errors.append({
                "source_row": row_data.get("source_row"),
                "apartment": apt_number,
                "errors": row_errors,
            })
            continue

        phone_e164 = row_data.get("tenant_phone", "").strip()
        phone_2_e164 = row_data.get("tenant_phone_2", "").strip()

        tenant = {
            "name": row_data["tenant_name"].strip(),
            "id_number": row_data.get("tenant_id_number", "").strip(),
            "phone": phone_e164,
            "phone_2": phone_2_e164,
            "email": row_data.get("tenant_email", "").strip(),
        }

        tenant_2 = None
        t2_name = row_data.get("tenant_2_name", "").strip()
        if t2_name:
            tenant_2 = {
                "name": t2_name,
                "id_number": row_data.get("tenant_2_id_number", "").strip(),
            }

        resolved_building_name = (row_data.get("building_name") or "").strip()
        if resolved_building_id and resolved_building_id in building_by_id:
            resolved_building_name = building_by_id[resolved_building_id]["name"]

        doc = {
            "project_id": project_id,
            "building_name": resolved_building_name,
            "building_id": resolved_building_id or (row_data.get("building_id") or "").strip() or None,
            "floor": (row_data.get("floor") or "").strip(),
            "floor_id": (row_data.get("floor_id") or "").strip() or None,
            "apartment_number": apt_number,
            "unit_id": resolved_uid,
            "tenant": tenant,
            "tenant_2": tenant_2,
            "handover_date": row_data.get("handover_date") or None,
            "unit_type": (row_data.get("unit_type") or "").strip() or None,
            "source": "g4_import",
            "imported_at": ts,
            "imported_by": user["id"],
            "import_batch_id": batch_id,
        }

        upsert_filter = {"project_id": project_id, "unit_id": resolved_uid}

        try:
            await db.unit_tenant_data.update_one(
                upsert_filter,
                {"$set": doc},
                upsert=True,
            )
            imported += 1
        except Exception as e:
            logger.error(f"Failed to upsert tenant data: {e}")
            skipped += 1
            errors.append({
                "source_row": row_data.get("source_row"),
                "apartment": row_data.get("apartment_number", ""),
                "errors": [str(e)],
            })

    await _audit("g4_import", batch_id, "executed", user["id"], {
        "project_id": project_id,
        "imported": imported,
        "skipped": skipped,
        "total_rows": len(body.rows),
    })

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "batch_id": batch_id,
    }
