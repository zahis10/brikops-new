"""Execution Matrix .xlsx export builder (Batch #502, 2026-05-05).

Pure helper — no DB / auth / FastAPI. Receives already-resolved data
from the router endpoint and returns a BytesIO of the .xlsx workbook.

Single sheet "מטריצת ביצוע" with:
  - Header row (frozen) + first 3 location columns frozen (E2 anchor).
  - RTL view (sheet_view.rightToLeft).
  - Status cells: Hebrew label + light fill matching MATRIX_STATUSES ramp.
  - Tag cells: text_value as plain text.
  - Notes attached as Excel comments (hover popup) — pre-cleaned of
    XML control chars to avoid IllegalCharacterError on PM-pasted text
    from Word/Email/Excel/WhatsApp.

Status colors mirror frontend STATUS_CONFIG (Tailwind 100-tier ramp).
"""
from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contractor_ops.spare_tiles import SPARE_OVERALL_LABELS


STATUS_LABELS = {
    "completed":      "הושלם",
    "partial":        "בוצע חלקית",
    "ready_for_work": "מוכן לעבודה",  # #503-followup-3
    "in_progress":    "בתהליך",
    "pending_review": "ממתין לאישור",  # #503-followup-2
    "not_done":       "לא בוצע",
    "not_relevant":   "לא רלוונטי",
    "no_findings":    "אין חוסרים",
}

STATUS_FILLS = {
    "completed":      PatternFill("solid", fgColor="D1FAE5"),  # green-100
    "partial":        PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    "ready_for_work": PatternFill("solid", fgColor="CFFAFE"),  # cyan-100 (#503-followup-3)
    "in_progress":    PatternFill("solid", fgColor="DBEAFE"),  # blue-100
    "pending_review": PatternFill("solid", fgColor="FFEDD5"),  # orange-100 (#503-followup-2)
    "not_done":       PatternFill("solid", fgColor="FEE2E2"),  # red-100
    "not_relevant":   PatternFill("solid", fgColor="F1F5F9"),  # slate-100
    "no_findings":    PatternFill("solid", fgColor="CCFBF1"),  # teal-100
}

HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")  # slate-100
SPARE_FILLS = {
    "short": PatternFill("solid", fgColor="FEE2E2"),
    "borderline": PatternFill("solid", fgColor="FEF3C7"),
    "ok": PatternFill("solid", fgColor="D1FAE5"),
    "recorded": PatternFill("solid", fgColor="D1FAE5"),
    "not_entered": PatternFill("solid", fgColor="F1F5F9"),
    "no_target": PatternFill("solid", fgColor="F1F5F9"),
    "no_profile": PatternFill("solid", fgColor="F1F5F9"),
}

# XML 1.0 illegal control chars — openpyxl raises IllegalCharacterError
# on these inside Comment text. PMs paste from Word / WhatsApp / Email
# where these can sneak in (NULL, vertical tab, form feed, etc).
_CTRL_CHAR_RE = re.compile(r"[\x00-\x08\x0b-\x1f]")


def _clean_note(note_raw):
    """Strip + remove control chars + truncate to 500 chars."""
    note_clean = (note_raw or "").strip()
    note_clean = _CTRL_CHAR_RE.sub("", note_clean)
    return note_clean[:500]


def build_matrix_xlsx(project, units, stages, cells, buildings, floors, spare=None):
    """Build the .xlsx workbook.

    Args:
        project:   project doc (uses .name only).
        units:     pre-filtered + sorted list of unit docs.
        stages:    pre-filtered + sorted list of stage docs.
        cells:     cell summaries (must be _summarize_cell output —
                   needs last_actor_name for Comment author).
        buildings: dict {building_id: building_doc}.
        floors:    dict {floor_id: floor_doc}.

    Returns:
        BytesIO positioned at offset 0, ready to stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "מטריצת ביצוע"
    ws.sheet_view.rightToLeft = True

    # ---- Header row ----
    headers = ["בניין", "קומה", "דירה", "מס׳ חדרים"]
    for s in stages:
        headers.append(s.get("title") or "—")
    spare_enabled = bool(spare and spare.get("enabled"))
    if spare_enabled:
        headers.append("ריצוף ספייר")
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )
        cell.fill = HEADER_FILL

    # ---- Data rows ----
    cells_index = {(c["unit_id"], c["stage_id"]): c for c in cells}
    for row_idx, unit in enumerate(units, start=2):
        building = buildings.get(unit.get("building_id")) or {}
        floor = floors.get(unit.get("floor_id")) or {}

        ws.cell(row=row_idx, column=1, value=building.get("name") or "—")
        ws.cell(row=row_idx, column=2, value=floor.get("floor_number") or "—")
        ws.cell(row=row_idx, column=3, value=unit.get("unit_no") or "—")
        ws.cell(row=row_idx, column=4, value=unit.get("room_count") or "")

        for stage_idx, stage in enumerate(stages):
            col = 5 + stage_idx
            cell_xl = ws.cell(row=row_idx, column=col)
            c = cells_index.get((unit["id"], stage["id"]))
            if not c:
                continue
            if stage.get("type") == "tag":
                cell_xl.value = c.get("text_value") or ""
            else:
                status = c.get("status")
                if status:
                    cell_xl.value = STATUS_LABELS.get(status, status)
                    fill = STATUS_FILLS.get(status)
                    if fill is not None:
                        cell_xl.fill = fill

            note_clean = _clean_note(c.get("note"))
            if note_clean:
                author = c.get("last_actor_name") or "BrikOps"
                cell_xl.comment = Comment(note_clean, author)

        if spare_enabled:
            spare_col = 5 + len(stages)
            spare_summary = (spare.get("by_unit") or {}).get(unit["id"]) or {}
            overall = spare_summary.get("overall", "no_profile")
            label = SPARE_OVERALL_LABELS.get(overall, overall)
            borderline = spare_summary.get("borderline") or []
            unfilled = spare_summary.get("unfilled") or []
            entered_count = spare_summary.get("entered_count") or 0
            applicable_count = spare_summary.get("applicable_count") or 0
            progress = (
                "לא הוזן" if entered_count == 0
                else f"הוזן {entered_count}/{applicable_count}"
            ) if applicable_count and entered_count < applicable_count else None
            if overall == "no_profile":
                label = "אחר" + (
                    f" · הוזן {entered_count}/{applicable_count}"
                    if entered_count else ""
                )
            elif overall == "short":
                details = ", ".join(
                    f"{row['name']} {row['missing']}"
                    if row.get("missing") is not None
                    else f"{row['name']} (אין ספייר)"
                    for row in spare_summary.get("short") or []
                )
                if details:
                    label = f"{label}: {details}"
                if borderline:
                    label = f"{label} · גבולי: {', '.join(borderline)}"
                if progress:
                    label = f"{label} · {progress}"
            elif overall == "borderline":
                if borderline:
                    label = f"{label}: {', '.join(borderline)}"
                if progress:
                    label = f"{label} · {progress}"
            elif overall == "not_entered":
                details = ", ".join(unfilled)
                if entered_count > 0:
                    label = f"הוזן {entered_count}/{applicable_count}"
                    if details:
                        label = f"{label} · לא הוזן: {details}"
                elif details:
                    label = f"{label}: {details}"
            elif overall == "recorded":
                label = (
                    f"הוזן {applicable_count}/{applicable_count}"
                    if applicable_count else label
                )
            spare_cell = ws.cell(row=row_idx, column=spare_col, value=label)
            fill = SPARE_FILLS.get(overall)
            if fill is not None:
                spare_cell.fill = fill
            profile = spare_summary.get("profile")
            if profile:
                spare_cell.comment = Comment(profile, "BrikOps")

    # ---- Frozen panes: header row + first 3 location columns ----
    ws.freeze_panes = "E2"

    # ---- Auto-width per column (clamp 8..30) ----
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                vl = len(str(v))
                if vl > max_len:
                    max_len = vl
        ws.column_dimensions[col_letter].width = max(8, min(30, max_len + 2))

    # #502-followup — Excel-native auto-filter arrows on every header
    # column. Users get sort A-Z / value-filter / custom-filter UX for
    # free inside Excel itself. ws.dimensions = bounding range of
    # populated cells (e.g. "A1:O35").
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
