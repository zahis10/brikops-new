"""Batch refactor-safety-split — section moved verbatim from safety_router.py (lines 2483-3196). MOVE, never edit."""
from contractor_ops.safety._shared import (  # noqa: F401
    Depends,
    HTTPException,
    Optional,
    Query,
    SAFETY_WRITERS,
    SafetyCategory,
    SafetyDocumentKind,
    SafetyDocumentStatus,
    SafetySeverity,
    StreamingResponse,
    _audit,
    _check_project_access,
    _time,
    datetime,
    get_db,
    io,
    require_roles,
    router,
    timedelta,
    timezone,
)

# =====================================================================
# Score & Exports — Phase 1 Part 3 (Backend Advanced)
# =====================================================================
_SCORE_CACHE: dict = {}  # {project_id: (timestamp, payload)}
_SCORE_TTL_SECONDS = 300  # 5 minutes

# Capped penalty buckets (spec §Steps.1)
_SCORE_DOCS_MAX = 40
_SCORE_TASKS_MAX = 25
_SCORE_TRAINING_MAX = 20
_SCORE_INCIDENTS_MAX = 15

# Safety score — required training types per worker (spec §Steps.1)
REQUIRED_TRAINING_TYPES = frozenset({"safety_induction", "height_work", "electrical"})


async def _compute_safety_score(db, project_id: str) -> dict:
    """
    Compute 0-100 safety score from live state via four capped penalty buckets:

      - documents bucket  (≤ _SCORE_DOCS_MAX = 40)
            sev3*10 + sev2*5 + sev1*2 over open/in_progress documents
      - tasks bucket      (≤ _SCORE_TASKS_MAX = 25)
            overdue (due_at past, status not completed/cancelled): 5 each
      - training bucket   (≤ _SCORE_TRAINING_MAX = 20)
            untrained workers: 4 each (no in-force training)
      - incidents bucket  (≤ _SCORE_INCIDENTS_MAX = 15)
            incidents in last 90 days: 8 each

    score = max(0, 100 - sum_of_capped_penalties).

    All Mongo reads filter `deletedAt: None`. Documents and tasks use
    `$group/$cond` aggregation in the style of snapshot_cron.py.
    """
    now = datetime.now(timezone.utc)
    today_iso = now.date().isoformat()
    now_iso = now.isoformat()
    cutoff_90d = (now - timedelta(days=90)).isoformat()

    # Documents bucket — single $group with $cond per severity (unchanged)
    docs_agg = await db.safety_documents.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "kind": {"$ne": "observation"},
            "status": {"$in": ["open", "in_progress"]},
        }},
        {"$group": {
            "_id": None,
            "sev3": {"$sum": {"$cond": [{"$eq": ["$severity", "3"]}, 1, 0]}},
            "sev2": {"$sum": {"$cond": [{"$eq": ["$severity", "2"]}, 1, 0]}},
            "sev1": {"$sum": {"$cond": [{"$eq": ["$severity", "1"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    sev3 = docs_agg[0].get("sev3", 0) if docs_agg else 0
    sev2 = docs_agg[0].get("sev2", 0) if docs_agg else 0
    sev1 = docs_agg[0].get("sev1", 0) if docs_agg else 0
    docs_raw = sev3 * 10 + sev2 * 5 + sev1 * 2
    docs_penalty = min(docs_raw, _SCORE_DOCS_MAX)

    # Tasks bucket — overdue tasks grouped by severity (spec §Steps.1)
    # Weights: sev3=6, sev2=3, sev1=1.5. Sum capped at _SCORE_TASKS_MAX (25).
    tasks_agg = await db.safety_tasks.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "due_at": {"$ne": None, "$lt": now_iso},
            "status": {"$nin": ["completed", "cancelled"]},
        }},
        {"$group": {
            "_id": None,
            "sev3": {"$sum": {"$cond": [{"$eq": ["$severity", "3"]}, 1, 0]}},
            "sev2": {"$sum": {"$cond": [{"$eq": ["$severity", "2"]}, 1, 0]}},
            "sev1": {"$sum": {"$cond": [{"$eq": ["$severity", "1"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    t_sev3 = tasks_agg[0].get("sev3", 0) if tasks_agg else 0
    t_sev2 = tasks_agg[0].get("sev2", 0) if tasks_agg else 0
    t_sev1 = tasks_agg[0].get("sev1", 0) if tasks_agg else 0
    tasks_raw = t_sev3 * 6 + t_sev2 * 3 + t_sev1 * 1.5
    tasks_penalty = min(tasks_raw, _SCORE_TASKS_MAX)
    overdue_tasks_total = t_sev3 + t_sev2 + t_sev1

    # Incidents bucket — last 90d grouped by type (spec §Steps.1)
    # Weights: injury=5, property_damage=3, near_miss=1. Cap _SCORE_INCIDENTS_MAX (15).
    inc_agg = await db.safety_incidents.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "occurred_at": {"$gte": cutoff_90d},
        }},
        {"$group": {
            "_id": None,
            "injury":          {"$sum": {"$cond": [{"$eq": ["$incident_type", "injury"]}, 1, 0]}},
            "property_damage": {"$sum": {"$cond": [{"$eq": ["$incident_type", "property_damage"]}, 1, 0]}},
            "near_miss":       {"$sum": {"$cond": [{"$eq": ["$incident_type", "near_miss"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    inc_injury   = inc_agg[0].get("injury", 0) if inc_agg else 0
    inc_property = inc_agg[0].get("property_damage", 0) if inc_agg else 0
    inc_near     = inc_agg[0].get("near_miss", 0) if inc_agg else 0
    incidents_raw = inc_injury * 5 + inc_property * 3 + inc_near * 1
    incidents_penalty = min(incidents_raw, _SCORE_INCIDENTS_MAX)
    recent_incidents_total = inc_injury + inc_property + inc_near

    # Training bucket — per worker, check each REQUIRED_TRAINING_TYPE has in-force record
    # (spec §Steps.1: 4pts per expired-required-training, 5pts for workers with NO training record)
    worker_ids_docs = await db.safety_workers.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0, "id": 1}
    ).to_list(length=10000)
    worker_ids = [w["id"] for w in worker_ids_docs]

    workers_with_expired_training = 0
    workers_without_training = 0
    expired_required_count = 0

    if worker_ids:
        in_force = await db.safety_trainings.find(
            {
                "project_id": project_id,
                "deletedAt": None,
                "worker_id": {"$in": worker_ids},
                "training_type": {"$in": list(REQUIRED_TRAINING_TYPES)},
                "$or": [{"expires_at": None}, {"expires_at": {"$gte": today_iso}}],
            },
            {"_id": 0, "worker_id": 1, "training_type": 1},
        ).to_list(length=100000)

        any_training_worker_ids = set(
            r["worker_id"] for r in await db.safety_trainings.find(
                {"project_id": project_id, "deletedAt": None, "worker_id": {"$in": worker_ids}},
                {"_id": 0, "worker_id": 1},
            ).to_list(length=100000)
        )

        worker_inforce: dict = {wid: set() for wid in worker_ids}
        for r in in_force:
            worker_inforce[r["worker_id"]].add(r["training_type"])

        for wid in worker_ids:
            if wid not in any_training_worker_ids:
                workers_without_training += 1
                continue
            missing_types = REQUIRED_TRAINING_TYPES - worker_inforce[wid]
            if missing_types:
                workers_with_expired_training += 1
                expired_required_count += len(missing_types)

    training_raw = expired_required_count * 4 + workers_without_training * 5
    training_penalty = min(training_raw, _SCORE_TRAINING_MAX)

    total_penalty = docs_penalty + tasks_penalty + training_penalty + incidents_penalty
    score = max(0, int(round(100 - total_penalty)))

    return {
        "score": score,
        "breakdown": {
            "doc_penalty": round(docs_penalty, 2),
            "task_penalty": round(tasks_penalty, 2),
            "training_penalty": round(training_penalty, 2),
            "incident_penalty": round(incidents_penalty, 2),
            "doc_counts": {"sev3": sev3, "sev2": sev2, "sev1": sev1},
            "overdue_task_counts": {
                "sev3": t_sev3, "sev2": t_sev2, "sev1": t_sev1,
                "total": overdue_tasks_total,
            },
            "workers_with_expired_training": workers_with_expired_training,
            "workers_without_training": workers_without_training,
            "total_workers": len(worker_ids),
            "incidents_last_90d": {
                "injury": inc_injury,
                "property_damage": inc_property,
                "near_miss": inc_near,
                "total": recent_incidents_total,
            },
            "caps": {
                "doc_max": _SCORE_DOCS_MAX,
                "task_max": _SCORE_TASKS_MAX,
                "training_max": _SCORE_TRAINING_MAX,
                "incident_max": _SCORE_INCIDENTS_MAX,
            },
        },
        "computed_at": now_iso,
    }


def _strip_pii(records: list) -> list:
    """Remove PII fields from worker records before any export."""
    cleaned = []
    for r in records:
        c = {k: v for k, v in r.items() if k not in ("_id", "id_number", "id_number_hash")}
        cleaned.append(c)
    return cleaned


@router.get("/{project_id}/score")
async def get_safety_score(
    project_id: str,
    refresh: bool = Query(False),
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    Return safety score with spec-required response shape:
      {score, breakdown, computed_at, cache_age_seconds}

    5-min in-process cache. `?refresh=true` recomputes. Always audited
    as `safety_score:computed` with `{score, cache_hit}`.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    cached = _SCORE_CACHE.get(project_id)
    cache_hit = False
    if cached and not refresh and (_time.time() - cached[0]) < _SCORE_TTL_SECONDS:
        payload = cached[1]
        cache_age = int(_time.time() - cached[0])
        cache_hit = True
    else:
        payload = await _compute_safety_score(db, project_id)
        _SCORE_CACHE[project_id] = (_time.time(), payload)
        cache_age = 0

    await _audit("safety_score", project_id, "computed", user["id"], {
        "project_id": project_id,
        "score": payload["score"],
        "cache_hit": cache_hit,
    })
    return {**payload, "cache_age_seconds": cache_age}


async def _gather_export_data(db, project_id: str, doc_filter_extra: Optional[dict] = None):
    """Fetch all data needed for exports. Strips PII from workers."""
    workers = await db.safety_workers.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)
    workers = _strip_pii(workers)

    trainings = await db.safety_trainings.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    doc_q = {"project_id": project_id, "deletedAt": None}
    if doc_filter_extra:
        doc_q.update(doc_filter_extra)
    documents = await db.safety_documents.find(doc_q, {"_id": 0}).to_list(length=100000)

    tasks = await db.safety_tasks.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    incidents = await db.safety_incidents.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    company_ids = {r.get("company_id") for r in workers + documents + tasks if r.get("company_id")}
    company_map = {}
    if company_ids:
        companies = await db.project_companies.find(
            {"id": {"$in": list(company_ids)}, "deletedAt": None},
            {"_id": 0, "id": 1, "name": 1},
        ).to_list(length=10000)
        company_map = {c["id"]: c.get("name", "") for c in companies}

    user_ids = set()
    for src in (documents, tasks, incidents):
        for r in src:
            for f in ("assignee_id", "reporter_id", "created_by"):
                if r.get(f):
                    user_ids.add(r[f])
    user_map = {}
    if user_ids:
        users = await db.users.find(
            {"id": {"$in": list(user_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(length=10000)
        user_map = {u["id"]: u.get("name", "") for u in users}

    return workers, trainings, documents, tasks, incidents, company_map, user_map


_EXCEL_HEADER_FONT_KW = dict(name="Arial", bold=True, size=11, color="FFFFFF")
_EXCEL_HEADER_FILL_KW = dict(start_color="475569", end_color="475569", fill_type="solid")


def _excel_styles():
    """Cached style objects for our Excel sheets (header grey, RTL alignment)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        "header_font": Font(**_EXCEL_HEADER_FONT_KW),
        "header_fill": PatternFill(**_EXCEL_HEADER_FILL_KW),
        "header_align": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "cell_font": Font(name="Arial", size=10),
        "cell_align": Alignment(horizontal="right", vertical="top", wrap_text=True),
        "thin": Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        ),
    }


def _fmt_dt(s):
    if not s:
        return ""
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(s)[:16]


def _write_sheet(wb, title, headers, rows, widths):
    """
    Add an RTL Hebrew sheet. Bold grey header, frozen row 1, optional
    auto-filter when there are data rows.
    """
    from openpyxl.utils import get_column_letter
    s = _excel_styles()
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["header_align"]
        c.border = s["thin"]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = s["cell_font"]
            c.alignment = s["cell_align"]
            c.border = s["thin"]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Spec: frozen row 1
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def _build_safety_excel_3sheet(
    workers, trainings, documents, incidents, company_map, worker_name_map, user_map
):
    """
    Build the regulatory RTL Hebrew workbook (spec §Steps.2). NOTE: the
    function name is a legacy misnomer — it now emits FOUR sheets:
      Sheet 1: עובדים  (Workers — NO id_number / id_number_hash)
      Sheet 2: הדרכות  (Trainings)
      Sheet 3: ליקויים (Safety findings / documents)
      Sheet 4: אירועים (Incidents)
    Frozen row 1 on every sheet.
    """
    from openpyxl import Workbook
    from services.safety_pdf import (
        SEVERITY_HE, INCIDENT_TYPE_HE, CATEGORY_HE, DOC_STATUS_HE,
    )

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: workers (PII already stripped upstream)
    worker_rows = []
    for w in workers:
        worker_rows.append([
            w.get("full_name", ""),
            w.get("profession", "") or "",
            company_map.get(w.get("company_id", ""), ""),
            w.get("phone", "") or "",
            _fmt_dt(w.get("created_at")),
        ])
    _write_sheet(
        wb, "עובדים",
        ["שם מלא", "מקצוע", "חברה", "טלפון", "תאריך כניסה"],
        worker_rows,
        [22, 18, 20, 14, 18],
    )

    # Sheet 2: trainings
    training_rows = []
    for tr in trainings:
        training_rows.append([
            worker_name_map.get(tr.get("worker_id", ""), ""),
            tr.get("training_type", "") or "",
            tr.get("instructor_name", "") or "",
            _fmt_dt(tr.get("trained_at")),
            _fmt_dt(tr.get("expires_at")),
        ])
    _write_sheet(
        wb, "הדרכות",
        ["עובד", "סוג הדרכה", "מדריך", "תאריך הדרכה", "תוקף עד"],
        training_rows,
        [22, 22, 18, 18, 18],
    )

    # Split into defects (ליקויים) and observations (תיעוד) — one collection,
    # kind discriminator. Legacy docs (no kind) count as defects.
    defects = [d for d in documents if d.get("kind") != "observation"]
    observations = [d for d in documents if d.get("kind") == "observation"]

    # Sheet 3: ליקויים (safety findings / documents) — mirrors the filtered
    # documents export row logic. No photo_urls (S3 keys, not human-readable).
    document_rows = []
    for d in defects:
        document_rows.append([
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            SEVERITY_HE.get(d.get("severity", ""), ""),
            DOC_STATUS_HE.get(d.get("status", ""), d.get("status", "")),
            _fmt_dt(d.get("found_at")),
            d.get("location", "") or "",
            company_map.get(d.get("company_id", ""), ""),
            user_map.get(d.get("assignee_id", ""), ""),
            d.get("description", "") or "",
        ])
    _write_sheet(
        wb, "ליקויים",
        ["כותרת", "קטגוריה", "חומרה", "סטטוס", "תאריך גילוי",
         "מיקום", "חברה", "אחראי", "תיאור"],
        document_rows,
        [26, 16, 12, 12, 16, 20, 20, 18, 40],
    )

    # Sheet: תיעוד (observations — archival, no severity/status)
    observation_rows = []
    for d in observations:
        observation_rows.append([
            _fmt_dt(d.get("found_at")),
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            d.get("location", "") or "",
            d.get("description", "") or "",
            user_map.get(d.get("reporter_id", ""), ""),
        ])
    _write_sheet(
        wb, "תיעוד",
        ["תאריך", "כותרת", "קטגוריה", "מיקום", "תיאור", "דווח ע\"י"],
        observation_rows,
        [16, 26, 16, 20, 40, 18],
    )

    # Sheet 4: incidents
    incident_rows = []
    for inc in incidents:
        incident_rows.append([
            INCIDENT_TYPE_HE.get(inc.get("incident_type", ""), inc.get("incident_type", "")),
            SEVERITY_HE.get(inc.get("severity", ""), ""),
            _fmt_dt(inc.get("occurred_at")),
            inc.get("location", "") or "",
            worker_name_map.get(inc.get("injured_worker_id", ""), "") or "",
            "כן" if inc.get("reported_to_authority") else "לא",
            inc.get("description", "") or "",
        ])
    _write_sheet(
        wb, "אירועים",
        ["סוג אירוע", "חומרה", "תאריך אירוע", "מיקום",
         "עובד נפגע", "דווח לרשות", "תיאור"],
        incident_rows,
        [16, 12, 18, 20, 22, 14, 40],
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_filtered_documents_excel(documents, company_map, user_map):
    """
    Build single-sheet 'ממצאי בטיחות' workbook (spec §Steps.3) of
    documents matching the 7-dim filter. Empty list still produces a
    valid Excel with header only.
    """
    from openpyxl import Workbook
    from services.safety_pdf import CATEGORY_HE, SEVERITY_HE, DOC_STATUS_HE

    wb = Workbook()
    wb.remove(wb.active)

    rows = []
    for d in documents:
        rows.append([
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            SEVERITY_HE.get(d.get("severity", ""), ""),
            DOC_STATUS_HE.get(d.get("status", ""), d.get("status", "")),
            d.get("location", "") or "",
            company_map.get(d.get("company_id", ""), ""),
            user_map.get(d.get("assignee_id", ""), ""),
            user_map.get(d.get("reporter_id", ""), ""),
            _fmt_dt(d.get("found_at")),
            _fmt_dt(d.get("created_at")),
            _fmt_dt(d.get("resolved_at")),
            d.get("description", "") or "",
        ])
    _write_sheet(
        wb, "ממצאי בטיחות",
        ["כותרת", "קטגוריה", "חומרה", "סטטוס", "מיקום", "חברה",
         "אחראי", "מדווח", "נמצא בתאריך", "נוצר", "נפתר", "תיאור"],
        rows,
        [25, 16, 10, 12, 18, 18, 16, 16, 18, 18, 18, 35],
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/{project_id}/export/excel")
async def export_safety_excel(
    project_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    4-sheet Hebrew RTL Excel export (workers / trainings / ליקויים / incidents).
    Management-only. PII (id_number / id_number_hash) is stripped.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    workers, trainings, documents, tasks, incidents, company_map, user_map = \
        await _gather_export_data(db, project_id)
    worker_name_map = {w["id"]: w.get("full_name", "") for w in workers}

    buf = _build_safety_excel_3sheet(
        workers, trainings, documents, incidents, company_map, worker_name_map, user_map
    )

    await _audit("safety_export", project_id, "excel_exported", user["id"], {
        "project_id": project_id,
        "workers": len(workers),
        "trainings": len(trainings),
        "documents": len(documents),
        "incidents": len(incidents),
    })

    filename = (
        f"safety_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/filtered")
async def export_safety_filtered(
    project_id: str,
    kind: Optional[SafetyDocumentKind] = None,
    category: Optional[SafetyCategory] = None,
    severity: Optional[SafetySeverity] = None,
    status_: Optional[SafetyDocumentStatus] = Query(None, alias="status"),
    company_id: Optional[str] = None,
    assignee_id: Optional[str] = None,
    reporter_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    Single-sheet 'ממצאי בטיחות' Excel of documents matching the same
    7-dim filter as list_documents. Empty result returns a valid Excel
    (200) with the header row only.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    extra = {}
    if kind == SafetyDocumentKind.observation:
        extra["kind"] = "observation"
    elif kind == SafetyDocumentKind.defect:
        extra["kind"] = {"$ne": "observation"}
    if category:    extra["category"] = category.value
    if severity:    extra["severity"] = severity.value
    if status_:     extra["status"] = status_.value
    if company_id:  extra["company_id"] = company_id
    if assignee_id: extra["assignee_id"] = assignee_id
    if reporter_id: extra["reporter_id"] = reporter_id
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to
        extra["found_at"] = rng

    doc_q = {"project_id": project_id, "deletedAt": None, **extra}
    documents = await db.safety_documents.find(doc_q, {"_id": 0}).to_list(length=100000)

    company_ids = {d.get("company_id") for d in documents if d.get("company_id")}
    company_map = {}
    if company_ids:
        companies = await db.project_companies.find(
            {"id": {"$in": list(company_ids)}, "deletedAt": None},
            {"_id": 0, "id": 1, "name": 1},
        ).to_list(length=10000)
        company_map = {c["id"]: c.get("name", "") for c in companies}

    user_ids = set()
    for d in documents:
        for f in ("assignee_id", "reporter_id"):
            if d.get(f):
                user_ids.add(d[f])
    user_map = {}
    if user_ids:
        users = await db.users.find(
            {"id": {"$in": list(user_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(length=10000)
        user_map = {u["id"]: u.get("name", "") for u in users}

    buf = _build_filtered_documents_excel(documents, company_map, user_map)

    # Flatten $-operator values (found_at range, kind $ne) into plain keys so the
    # audit payload never contains $-prefixed field names (MongoDB rejects those).
    applied = {k: v for k, v in extra.items() if k not in ("found_at", "kind")}
    if kind:
        applied["kind"] = kind.value
    if date_from:
        applied["date_from"] = date_from
    if date_to:
        applied["date_to"] = date_to

    await _audit("safety_export", project_id, "filtered_exported", user["id"], {
        "project_id": project_id,
        "filters": applied,
        "matched_documents": len(documents),
    })

    filename = (
        f"safety_filtered_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/pdf-register")
async def export_safety_pdf_register(
    project_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """Generate the 9-section Hebrew 'פנקס כללי' PDF. Management-only."""
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    from services.safety_pdf import generate_safety_register
    pdf_bytes = await generate_safety_register(db, project_id)

    await _audit("safety_export", project_id, "pdf_register_exported", user["id"], {
        "project_id": project_id,
        "size_bytes": len(pdf_bytes),
    })

    filename = (
        f"safety_register_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/tours/{tour_id}/export/pdf")
async def export_tour_pdf(
    project_id: str,
    tour_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """Generate the per-tour safety report PDF. Any status is exportable."""
    db = get_db()
    await _check_project_access(user, project_id)

    tour = await db.safety_tours.find_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None}, {"_id": 0, "id": 1})
    if not tour:
        raise HTTPException(status_code=404, detail="tour not found")

    from services.safety_pdf import generate_tour_report
    pdf_bytes = await generate_tour_report(db, project_id, tour_id)

    await _audit("safety_export", tour_id, "tour_pdf_exported", user["id"], {
        "project_id": project_id,
        "tour_id": tour_id,
        "size_bytes": len(pdf_bytes),
    })

    filename = (
        f"safety_tour_{tour_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


