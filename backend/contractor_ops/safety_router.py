"""
Safety module router — Phase 1 Part 1 (Foundation) + Part 2 (Backend Core).

Part 1: feature flag wiring, healthz, ensure_safety_indexes (20 indices).
Part 2: 25 CRUD endpoints (Workers/Trainings/Documents/Tasks/Incidents) +
        7-dim Documents filter + photo/PDF upload helper + audit on every write.

Registration in server.py is GATED by ENABLE_SAFETY_MODULE env flag.
When flag is off, this module is never imported and endpoints 404.
"""
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Literal
import hashlib
import logging
import re
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, model_validator
import io
import time as _time

from services.object_storage import generate_url

from contractor_ops.router import (
    get_current_user, get_db, _now, _audit,
    _check_project_access, _get_project_role, _get_project_membership,
    _is_super_admin, require_roles,
)
from contractor_ops.schemas import (
    SafetyWorker, SafetyTraining, SafetyDocument, SafetyTask, SafetyIncident,
    SafetyCategory, SafetySeverity, SafetyDocumentStatus, SafetyDocumentKind,
    SafetyTaskStatus,
    SafetyTour, SafetyTourType, SafetyTourStatus, SafetyTourItem, SafetyTourSignature,
)
from contractor_ops.upload_rate_limit import (
    check_upload_rate_limit, check_upload_bytes, check_content_length,
)
from contractor_ops.upload_quota import check_storage_quota, record_upload
from contractor_ops.upload_safety import (
    validate_upload, ALLOWED_IMAGE_EXTENSIONS,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/safety", tags=["safety"])


# =====================================================================
# Constants
# =====================================================================
SAFETY_WRITERS = ("project_manager", "management_team")
SAFETY_DELETERS = ("project_manager",)
INCIDENT_RETENTION_YEARS = 7

# Upload — accepts images AND PDFs (qc is image-only)
SAFETY_ALLOWED_CONTENT_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif",
    "application/pdf",
}
SAFETY_ALLOWED_EXTENSIONS = ALLOWED_IMAGE_EXTENSIONS | {".pdf"}
MAX_SAFETY_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB

VALID_INCIDENT_TYPES = {"near_miss", "injury", "property_damage"}


# =====================================================================
# Helpers
# =====================================================================
def _new_id() -> str:
    """UUID4 as string. Matches project convention."""
    return str(uuid.uuid4())


def _hash_id_number(id_number: Optional[str]) -> Optional[str]:
    """SHA-256 of stripped Israeli ID / passport. Lookup aid; never returned to client."""
    if not id_number:
        return None
    return hashlib.sha256(id_number.strip().encode("utf-8")).hexdigest()


def _retention_date(from_iso: str) -> str:
    """from_iso + 7 years, ISO string. Used for incident soft-delete (regulatory clock)."""
    dt = datetime.fromisoformat(from_iso.replace("Z", "+00:00"))
    return (dt + timedelta(days=365 * INCIDENT_RETENTION_YEARS)).isoformat()


async def _ensure_company_or_placeholder(
    db, project_id: str, company_id: Optional[str], company_name: Optional[str], actor_id: str
) -> Optional[str]:
    """
    If company_id is provided and exists on this project (non-deleted) — return it.
    If only company_name is provided — create or reuse a placeholder project_companies doc.
    If neither — return None.
    """
    if company_id:
        doc = await db.project_companies.find_one({
            "id": company_id,
            "project_id": project_id,
            "deletedAt": None,
        })
        if doc:
            return company_id
        raise HTTPException(status_code=404, detail="company_id not found on this project")

    if not company_name:
        return None

    name = company_name.strip()
    existing = await db.project_companies.find_one({
        "project_id": project_id,
        "name": name,
        "deletedAt": None,
    })
    if existing:
        return existing["id"]

    new_id = _new_id()
    await db.project_companies.insert_one({
        "id": new_id,
        "project_id": project_id,
        "name": name,
        "is_placeholder": True,
        "created_at": _now(),
        "created_by": actor_id,
    })
    await _audit("project_company", new_id, "create_placeholder", actor_id, {
        "project_id": project_id, "name": name, "source": "safety",
    })
    return new_id


async def _strip_mongo_id(items: list) -> list:
    """Drop _id from list results (Mongo returns ObjectId which isn't JSON-serializable)."""
    for it in items:
        it.pop("_id", None)
    return items


def _resolve_include_deleted(include_deleted: bool, user: dict) -> bool:
    """
    Spec: include_deleted=true is restricted to super_admin.
    Non-super callers passing include_deleted=true get 403, not silent demotion.
    """
    if include_deleted and not _is_super_admin(user):
        raise HTTPException(status_code=403, detail="include_deleted requires super_admin")
    return include_deleted


# =====================================================================
# healthz (Part 1)
# =====================================================================
@router.get("/healthz")
async def healthz(user: dict = Depends(get_current_user)):
    """
    Liveness check for Safety module.
    Requires auth — never expose module existence to unauthenticated scanners.
    """
    return {"ok": True, "module": "safety", "enabled": True}


# =====================================================================
# Shared request bodies
# =====================================================================
class SoftDeleteBody(BaseModel):
    reason: Optional[str] = None


# =====================================================================
# Workers CRUD
# =====================================================================
class SafetyWorkerCreate(BaseModel):
    company_id: Optional[str] = None
    company_name: Optional[str] = None
    full_name: str = Field(..., min_length=2, max_length=120)
    id_number: Optional[str] = None
    profession: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None


class SafetyWorkerUpdate(BaseModel):
    full_name: Optional[str] = None
    id_number: Optional[str] = None
    profession: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None
    company_id: Optional[str] = None


@router.post("/{project_id}/workers", status_code=201, response_model=SafetyWorker)
async def create_worker(
    project_id: str,
    payload: SafetyWorkerCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    resolved_company_id = await _ensure_company_or_placeholder(
        db, project_id, payload.company_id, payload.company_name, user["id"]
    )

    # PII: persist only the SHA-256 hash; never store or return raw id_number.
    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "company_id": resolved_company_id,
        "full_name": payload.full_name.strip(),
        "id_number": None,
        "id_number_hash": _hash_id_number(payload.id_number),
        "profession": payload.profession,
        "phone": payload.phone,
        "notes": payload.notes,
        "created_at": _now(),
        "created_by": user["id"],
        "deletedAt": None,
        "deletedBy": None,
        "deletion_reason": None,
        "retention_until": None,
    }
    await db.safety_workers.insert_one(doc)
    audit_after = {k: v for k, v in doc.items() if k not in ("id_number", "id_number_hash")}
    await _audit("safety_worker", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": audit_after,
    })
    return SafetyWorker(**doc)


@router.get("/{project_id}/workers")
async def list_workers(
    project_id: str,
    profession: Optional[str] = None,
    company_id: Optional[str] = None,
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if profession and profession.strip():
        q["profession"] = {"$regex": re.escape(profession.strip()), "$options": "i"}
    if company_id:
        q["company_id"] = company_id

    total = await db.safety_workers.count_documents(q)
    cursor = db.safety_workers.find(q, {"_id": 0}).sort("created_at", -1).skip(offset).limit(limit)
    items = await cursor.to_list(length=limit)
    # strip the hash from list responses too
    for it in items:
        it.pop("id_number_hash", None)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{project_id}/workers/{worker_id}", response_model=SafetyWorker)
async def get_worker(
    project_id: str,
    worker_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_workers.find_one({"id": worker_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="worker not found")
    # PII: never expose hash or any raw id_number
    doc.pop("id_number_hash", None)
    doc["id_number"] = None
    return SafetyWorker(**doc)


@router.patch("/{project_id}/workers/{worker_id}", response_model=SafetyWorker)
async def update_worker(
    project_id: str,
    worker_id: str,
    payload: SafetyWorkerUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_workers.find_one({"id": worker_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="worker not found")
    if before.get("deletedAt"):
        raise HTTPException(status_code=410, detail="worker deleted")

    updates = payload.model_dump(exclude_unset=True)
    if "id_number" in updates:
        # PII: rotate hash only; never persist raw id_number on the document.
        updates["id_number_hash"] = _hash_id_number(updates["id_number"]) if updates["id_number"] else None
        updates["id_number"] = None
    if "company_id" in updates and updates["company_id"]:
        comp = await db.project_companies.find_one({
            "id": updates["company_id"], "project_id": project_id, "deletedAt": None,
        })
        if not comp:
            raise HTTPException(status_code=404, detail="company_id not found on this project")
    updates["updated_at"] = _now()
    updates["updated_by"] = user["id"]

    await db.safety_workers.update_one({"id": worker_id, "project_id": project_id, "deletedAt": None}, {"$set": updates})
    after = await db.safety_workers.find_one({"id": worker_id})
    _STRIP_PII = ("_id", "id_number", "id_number_hash")
    await _audit("safety_worker", worker_id, "updated", user["id"], {
        "project_id": project_id,
        "before": {k: v for k, v in before.items() if k not in _STRIP_PII},
        "after": {k: v for k, v in after.items() if k not in _STRIP_PII},
    })
    return SafetyWorker(**after)


@router.delete("/{project_id}/workers/{worker_id}", status_code=204)
async def delete_worker(
    project_id: str,
    worker_id: str,
    body: SoftDeleteBody = SoftDeleteBody(),
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_workers.find_one({"id": worker_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="worker not found")
    if before.get("deletedAt"):
        return

    now = _now()
    await db.safety_workers.update_one({"id": worker_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": now,
        "deletedBy": user["id"],
        "deletion_reason": body.reason,
        "retention_until": _retention_date(now),
    }})
    await _audit("safety_worker", worker_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason,
    })


# =====================================================================
# Trainings CRUD
# =====================================================================
class SafetyTrainingCreate(BaseModel):
    worker_id: str
    training_type: str
    trained_at: str
    instructor_name: Optional[str] = None
    duration_minutes: Optional[int] = None
    location: Optional[str] = None
    expires_at: Optional[str] = None
    certificate_url: Optional[str] = None


class SafetyTrainingUpdate(BaseModel):
    training_type: Optional[str] = None
    trained_at: Optional[str] = None
    instructor_name: Optional[str] = None
    duration_minutes: Optional[int] = None
    location: Optional[str] = None
    expires_at: Optional[str] = None
    certificate_url: Optional[str] = None


@router.post("/{project_id}/trainings", status_code=201, response_model=SafetyTraining)
async def create_training(
    project_id: str,
    payload: SafetyTrainingCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    worker = await db.safety_workers.find_one({
        "id": payload.worker_id,
        "project_id": project_id,
        "deletedAt": None,
    })
    if not worker:
        raise HTTPException(status_code=404, detail="worker not found or deleted")

    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "worker_id": payload.worker_id,
        "training_type": payload.training_type.strip(),
        "instructor_name": payload.instructor_name,
        "duration_minutes": payload.duration_minutes,
        "location": payload.location,
        "trained_at": payload.trained_at,
        "expires_at": payload.expires_at,
        "certificate_url": payload.certificate_url,
        "created_at": _now(),
        "created_by": user["id"],
        "deletedAt": None,
        "deletedBy": None,
    }
    await db.safety_trainings.insert_one(doc)
    await _audit("safety_training", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": {k: v for k, v in doc.items() if k != "_id"},
    })
    return SafetyTraining(**doc)


@router.get("/{project_id}/trainings")
async def list_trainings(
    project_id: str,
    worker_id: Optional[str] = None,
    training_type: Optional[str] = None,
    expires_before: Optional[str] = None,
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if worker_id:
        q["worker_id"] = worker_id
    if training_type and training_type.strip():
        q["training_type"] = {"$regex": re.escape(training_type.strip()), "$options": "i"}
    if expires_before:
        q["expires_at"] = {"$ne": None, "$lt": expires_before}

    total = await db.safety_trainings.count_documents(q)
    cursor = db.safety_trainings.find(q, {"_id": 0}).sort("trained_at", -1).skip(offset).limit(limit)
    items = await cursor.to_list(length=limit)
    for it in items:
        k = it.get("certificate_url")
        it["certificate_display_url"] = (generate_url(k) if k else None)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{project_id}/trainings/{training_id}", response_model=SafetyTraining)
async def get_training(
    project_id: str,
    training_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_trainings.find_one({"id": training_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="training not found")
    k = doc.get("certificate_url")
    doc["certificate_display_url"] = (generate_url(k) if k else None)
    return SafetyTraining(**doc)


@router.patch("/{project_id}/trainings/{training_id}", response_model=SafetyTraining)
async def update_training(
    project_id: str,
    training_id: str,
    payload: SafetyTrainingUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_trainings.find_one({"id": training_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="training not found")
    if before.get("deletedAt"):
        raise HTTPException(status_code=410, detail="training deleted")

    updates = payload.model_dump(exclude_unset=True)
    updates["updated_at"] = _now()
    updates["updated_by"] = user["id"]

    await db.safety_trainings.update_one({"id": training_id, "project_id": project_id, "deletedAt": None}, {"$set": updates})
    after = await db.safety_trainings.find_one({"id": training_id})
    await _audit("safety_training", training_id, "updated", user["id"], {
        "project_id": project_id,
        "before": {k: v for k, v in before.items() if k != "_id"},
        "after": {k: v for k, v in after.items() if k != "_id"},
    })
    return SafetyTraining(**after)


@router.delete("/{project_id}/trainings/{training_id}", status_code=204)
async def delete_training(
    project_id: str,
    training_id: str,
    body: SoftDeleteBody = SoftDeleteBody(),
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_trainings.find_one({"id": training_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="training not found")
    if before.get("deletedAt"):
        return
    now = _now()
    await db.safety_trainings.update_one({"id": training_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": now,
        "deletedBy": user["id"],
        "deletion_reason": body.reason,
        "retention_until": _retention_date(now),
    }})
    await _audit("safety_training", training_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason,
    })


# =====================================================================
# Documents CRUD + 7-dim filter
# =====================================================================
class SafetyDocumentCreate(BaseModel):
    kind: SafetyDocumentKind = SafetyDocumentKind.defect
    category: SafetyCategory
    severity: Optional[SafetySeverity] = None
    title: str = Field(..., min_length=2, max_length=200)
    found_at: str
    description: Optional[str] = None
    location: Optional[str] = None
    company_id: Optional[str] = None
    profession: Optional[str] = None
    assignee_id: Optional[str] = None
    photo_urls: List[str] = Field(default_factory=list)
    attachment_urls: List[str] = Field(default_factory=list)

    @model_validator(mode="after")
    def _validate_kind_severity(self):
        if self.kind == SafetyDocumentKind.defect and self.severity is None:
            raise ValueError("יש לבחור חומרה לליקוי")
        if self.kind == SafetyDocumentKind.observation and self.severity is not None:
            raise ValueError("לתיעוד אין חומרה")
        return self


class SafetyDocumentUpdate(BaseModel):
    category: Optional[SafetyCategory] = None
    severity: Optional[SafetySeverity] = None
    status: Optional[SafetyDocumentStatus] = None
    title: Optional[str] = None
    description: Optional[str] = None
    location: Optional[str] = None
    company_id: Optional[str] = None
    profession: Optional[str] = None
    assignee_id: Optional[str] = None
    photo_urls: Optional[List[str]] = None
    attachment_urls: Optional[List[str]] = None
    resolved_at: Optional[str] = None


@router.post("/{project_id}/documents", status_code=201, response_model=SafetyDocument)
async def create_document(
    project_id: str,
    payload: SafetyDocumentCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    if payload.company_id:
        comp = await db.project_companies.find_one({
            "id": payload.company_id, "project_id": project_id, "deletedAt": None,
        })
        if not comp:
            raise HTTPException(status_code=404, detail="company_id not found on this project")

    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "kind": payload.kind.value,
        "category": payload.category.value,
        "severity": payload.severity.value if payload.severity else None,
        "status": SafetyDocumentStatus.open.value,
        "title": payload.title.strip(),
        "description": payload.description,
        "location": payload.location,
        "company_id": payload.company_id,
        "profession": payload.profession,
        "assignee_id": payload.assignee_id,
        "reporter_id": user["id"],          # SERVER-SET, never trust client
        "photo_urls": payload.photo_urls,
        "attachment_urls": payload.attachment_urls,
        "found_at": payload.found_at,
        "resolved_at": None,
        "created_at": _now(),
        "created_by": user["id"],
        "updated_at": None,
        "deletedAt": None,
        "deletedBy": None,
    }
    await db.safety_documents.insert_one(doc)
    await _audit("safety_document", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": {k: v for k, v in doc.items() if k != "_id"},
    })
    return SafetyDocument(**doc)


@router.get("/{project_id}/documents")
async def list_documents(
    project_id: str,
    kind: Optional[SafetyDocumentKind] = None,
    category: Optional[SafetyCategory] = None,
    severity: Optional[SafetySeverity] = None,
    status_: Optional[SafetyDocumentStatus] = Query(None, alias="status"),
    company_id: Optional[str] = None,
    assignee_id: Optional[str] = None,
    reporter_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if kind == SafetyDocumentKind.observation:
        q["kind"] = "observation"
    elif kind == SafetyDocumentKind.defect:
        q["kind"] = {"$ne": "observation"}   # covers legacy pre-backfill docs
    if category:    q["category"] = category.value
    if severity:    q["severity"] = severity.value
    if status_:     q["status"] = status_.value
    if company_id:  q["company_id"] = company_id
    if assignee_id: q["assignee_id"] = assignee_id
    if reporter_id: q["reporter_id"] = reporter_id
    if date_from or date_to:
        q["found_at"] = {}
        if date_from: q["found_at"]["$gte"] = date_from
        if date_to:   q["found_at"]["$lte"] = date_to

    total = await db.safety_documents.count_documents(q)
    cursor = db.safety_documents.find(q, {"_id": 0}).sort("found_at", -1).skip(offset).limit(limit)
    items = await cursor.to_list(length=limit)
    for it in items:
        it["photo_display_urls"] = [
            (generate_url(k) if k else k) for k in (it.get("photo_urls") or [])
        ]
    filters_applied = {k: v for k, v in {
        "kind": kind.value if kind else None,
        "category": category.value if category else None,
        "severity": severity.value if severity else None,
        "status": status_.value if status_ else None,
        "company_id": company_id, "assignee_id": assignee_id, "reporter_id": reporter_id,
        "date_from": date_from, "date_to": date_to,
    }.items() if v}
    return {"items": items, "total": total, "limit": limit, "offset": offset,
            "filters_applied": filters_applied}


@router.get("/{project_id}/documents/{document_id}", response_model=SafetyDocument)
async def get_document(
    project_id: str,
    document_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_documents.find_one({"id": document_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="document not found")
    doc["photo_display_urls"] = [
        (generate_url(k) if k else k) for k in (doc.get("photo_urls") or [])
    ]
    return SafetyDocument(**doc)


@router.patch("/{project_id}/documents/{document_id}", response_model=SafetyDocument)
async def update_document(
    project_id: str,
    document_id: str,
    payload: SafetyDocumentUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_documents.find_one({"id": document_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="document not found")
    if before.get("deletedAt"):
        raise HTTPException(status_code=410, detail="document deleted")

    if before.get("kind", "defect") == "observation":
        blocked = {"status", "severity", "resolved_at"} & set(
            payload.model_dump(exclude_unset=True).keys()
        )
        if blocked:
            raise HTTPException(status_code=422, detail="שדות סטטוס/חומרה אינם רלוונטיים לתיעוד")

    updates = payload.model_dump(exclude_unset=True)

    # Coerce enum values to strings for Mongo
    if "category" in updates and updates["category"] is not None:
        updates["category"] = updates["category"].value if hasattr(updates["category"], "value") else updates["category"]
    if "severity" in updates and updates["severity"] is not None:
        updates["severity"] = updates["severity"].value if hasattr(updates["severity"], "value") else updates["severity"]
    if "status" in updates and updates["status"] is not None:
        updates["status"] = updates["status"].value if hasattr(updates["status"], "value") else updates["status"]

    if "company_id" in updates and updates["company_id"]:
        comp = await db.project_companies.find_one({
            "id": updates["company_id"], "project_id": project_id, "deletedAt": None,
        })
        if not comp:
            raise HTTPException(status_code=404, detail="company_id not found on this project")

    updates["updated_at"] = _now()
    updates["updated_by"] = user["id"]

    await db.safety_documents.update_one({"id": document_id, "project_id": project_id, "deletedAt": None}, {"$set": updates})
    after = await db.safety_documents.find_one({"id": document_id})
    await _audit("safety_document", document_id, "updated", user["id"], {
        "project_id": project_id,
        "before": {k: v for k, v in before.items() if k != "_id"},
        "after": {k: v for k, v in after.items() if k != "_id"},
    })
    return SafetyDocument(**after)


@router.delete("/{project_id}/documents/{document_id}", status_code=204)
async def delete_document(
    project_id: str,
    document_id: str,
    body: SoftDeleteBody = SoftDeleteBody(),
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_documents.find_one({"id": document_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="document not found")
    if before.get("deletedAt"):
        return
    now = _now()
    await db.safety_documents.update_one({"id": document_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": now,
        "deletedBy": user["id"],
        "deletion_reason": body.reason,
        "retention_until": _retention_date(now),
    }})
    await _audit("safety_document", document_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason,
    })


# =====================================================================
# Tasks CRUD + status transition guard
# =====================================================================
class SafetyTaskCreate(BaseModel):
    title: str = Field(..., min_length=2, max_length=200)
    severity: SafetySeverity
    document_id: Optional[str] = None
    description: Optional[str] = None
    assignee_id: Optional[str] = None
    company_id: Optional[str] = None
    due_at: Optional[str] = None


class SafetyTaskUpdate(BaseModel):
    title: Optional[str] = None
    severity: Optional[SafetySeverity] = None
    description: Optional[str] = None
    assignee_id: Optional[str] = None
    company_id: Optional[str] = None
    due_at: Optional[str] = None
    status: Optional[SafetyTaskStatus] = None
    corrective_action: Optional[str] = None
    verification_photo_urls: Optional[List[str]] = None


def _assert_transition_allowed(
    current: str, next_status: str, project_role: str, payload_dict: dict
) -> None:
    """Enforce status transition rules. Raises 409 on invalid transition."""
    if current == next_status:
        return  # idempotent
    # Spec: open → in_progress → completed; any → cancelled; reverse = 409
    allowed = {
        "open": {"in_progress", "cancelled"},
        "in_progress": {"completed", "cancelled"},
        "completed": {"cancelled"},
        "cancelled": set(),
    }
    if next_status not in allowed.get(current, set()):
        raise HTTPException(status_code=409, detail={
            "current_status": current,
            "requested": next_status,
            "reason": "transition not allowed",
        })
    if next_status == "completed":
        ca = payload_dict.get("corrective_action")
        if not ca or not str(ca).strip():
            raise HTTPException(status_code=409, detail={
                "current_status": current,
                "requested": next_status,
                "reason": "corrective_action is required to mark a task completed",
            })
    if next_status == "cancelled":
        if project_role != "project_manager":
            raise HTTPException(status_code=409, detail={
                "current_status": current,
                "requested": next_status,
                "reason": "only project_manager may cancel a task",
            })


@router.post("/{project_id}/tasks", status_code=201, response_model=SafetyTask)
async def create_task(
    project_id: str,
    payload: SafetyTaskCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    if payload.document_id:
        ref = await db.safety_documents.find_one({
            "id": payload.document_id, "project_id": project_id, "deletedAt": None,
        })
        if not ref:
            raise HTTPException(status_code=404, detail="document_id not found or deleted")
        if ref.get("kind", "defect") == "observation":
            raise HTTPException(status_code=422, detail="לא ניתן לפתוח משימה מתקנת על תיעוד")

    if payload.company_id:
        comp = await db.project_companies.find_one({
            "id": payload.company_id, "project_id": project_id, "deletedAt": None,
        })
        if not comp:
            raise HTTPException(status_code=404, detail="company_id not found on this project")

    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "document_id": payload.document_id,
        "title": payload.title.strip(),
        "description": payload.description,
        "status": SafetyTaskStatus.open.value,
        "severity": payload.severity.value,
        "assignee_id": payload.assignee_id,
        "company_id": payload.company_id,
        "due_at": payload.due_at,
        "completed_at": None,
        "corrective_action": None,
        "verification_photo_urls": [],
        "created_at": _now(),
        "created_by": user["id"],
        "updated_at": None,
        "deletedAt": None,
        "deletedBy": None,
    }
    await db.safety_tasks.insert_one(doc)
    await _audit("safety_task", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": {k: v for k, v in doc.items() if k != "_id"},
    })
    return SafetyTask(**doc)


@router.get("/{project_id}/tasks")
async def list_tasks(
    project_id: str,
    status_: Optional[SafetyTaskStatus] = Query(None, alias="status"),
    assignee_id: Optional[str] = None,
    document_id: Optional[str] = None,
    severity: Optional[SafetySeverity] = None,
    company_id: Optional[str] = None,
    overdue: Optional[bool] = None,
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    """List safety tasks.

    NOTE: when both overdue=true and status=<x> are sent, overdue WINS — its
    status $in (open/in_progress) overwrites the explicit status filter. The FE
    disables the status select while "באיחור" is on. overdue reuses the exact
    same semantics as the score's overdue bucket.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if status_:      q["status"] = status_.value
    if assignee_id:  q["assignee_id"] = assignee_id
    if document_id:  q["document_id"] = document_id
    if severity:     q["severity"] = severity.value
    if company_id:   q["company_id"] = company_id
    if overdue:
        now_iso = datetime.now(timezone.utc).isoformat()
        q["due_at"] = {"$ne": None, "$lt": now_iso}
        q["status"] = {"$in": ["open", "in_progress"]}

    total = await db.safety_tasks.count_documents(q)
    cursor = db.safety_tasks.find(q, {"_id": 0}).sort("created_at", -1).skip(offset).limit(limit)
    items = await cursor.to_list(length=limit)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{project_id}/tasks/{task_id}", response_model=SafetyTask)
async def get_task(
    project_id: str,
    task_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_tasks.find_one({"id": task_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="task not found")
    return SafetyTask(**doc)


@router.patch("/{project_id}/tasks/{task_id}", response_model=SafetyTask)
async def update_task(
    project_id: str,
    task_id: str,
    payload: SafetyTaskUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_tasks.find_one({"id": task_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="task not found")
    if before.get("deletedAt"):
        raise HTTPException(status_code=410, detail="task deleted")

    updates = payload.model_dump(exclude_unset=True)

    # Coerce enums
    if "severity" in updates and updates["severity"] is not None:
        updates["severity"] = updates["severity"].value if hasattr(updates["severity"], "value") else updates["severity"]

    next_status_raw = updates.get("status")
    if next_status_raw is not None:
        next_status = next_status_raw.value if hasattr(next_status_raw, "value") else next_status_raw
        project_role = await _get_project_role(user, project_id)
        _assert_transition_allowed(before.get("status", "open"), next_status, project_role, updates)
        updates["status"] = next_status
        if next_status == "completed":
            updates["completed_at"] = _now()

    if "company_id" in updates and updates["company_id"]:
        comp = await db.project_companies.find_one({
            "id": updates["company_id"], "project_id": project_id, "deletedAt": None,
        })
        if not comp:
            raise HTTPException(status_code=404, detail="company_id not found on this project")

    updates["updated_at"] = _now()
    updates["updated_by"] = user["id"]

    await db.safety_tasks.update_one({"id": task_id, "project_id": project_id, "deletedAt": None}, {"$set": updates})
    after = await db.safety_tasks.find_one({"id": task_id})
    await _audit("safety_task", task_id, "updated", user["id"], {
        "project_id": project_id,
        "before": {k: v for k, v in before.items() if k != "_id"},
        "after": {k: v for k, v in after.items() if k != "_id"},
    })
    return SafetyTask(**after)


@router.delete("/{project_id}/tasks/{task_id}", status_code=204)
async def delete_task(
    project_id: str,
    task_id: str,
    body: SoftDeleteBody = SoftDeleteBody(),
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_tasks.find_one({"id": task_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="task not found")
    if before.get("deletedAt"):
        return
    now = _now()
    await db.safety_tasks.update_one({"id": task_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": now,
        "deletedBy": user["id"],
        "deletion_reason": body.reason,
        "retention_until": _retention_date(now),
    }})
    await _audit("safety_task", task_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason,
    })


# =====================================================================
# Incidents CRUD + 7yr retention from occurred_at
# =====================================================================
class SafetyIncidentCreate(BaseModel):
    incident_type: str
    severity: SafetySeverity
    occurred_at: str
    description: str = Field(..., min_length=2)
    location: Optional[str] = None
    injured_worker_id: Optional[str] = None
    witnesses: List[str] = Field(default_factory=list)
    photo_urls: List[str] = Field(default_factory=list)
    medical_record_urls: List[str] = Field(default_factory=list)
    reported_to_authority: bool = False
    authority_report_ref: Optional[str] = None


class SafetyIncidentUpdate(BaseModel):
    incident_type: Optional[str] = None
    severity: Optional[SafetySeverity] = None
    occurred_at: Optional[str] = None
    description: Optional[str] = None
    location: Optional[str] = None
    injured_worker_id: Optional[str] = None
    witnesses: Optional[List[str]] = None
    photo_urls: Optional[List[str]] = None
    medical_record_urls: Optional[List[str]] = None
    reported_to_authority: Optional[bool] = None
    authority_report_ref: Optional[str] = None
    status: Optional[str] = None  # incident lifecycle: draft|reported|closed


@router.post("/{project_id}/incidents", status_code=201, response_model=SafetyIncident)
async def create_incident(
    project_id: str,
    payload: SafetyIncidentCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    if payload.incident_type not in VALID_INCIDENT_TYPES:
        raise HTTPException(status_code=400, detail=f"invalid incident_type; must be one of {sorted(VALID_INCIDENT_TYPES)}")

    if payload.injured_worker_id:
        worker = await db.safety_workers.find_one({
            "id": payload.injured_worker_id, "project_id": project_id, "deletedAt": None,
        })
        if not worker:
            raise HTTPException(status_code=404, detail="injured_worker_id not found or deleted")

    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "incident_type": payload.incident_type,
        "severity": payload.severity.value,
        "occurred_at": payload.occurred_at,
        "description": payload.description,
        "location": payload.location,
        "injured_worker_id": payload.injured_worker_id,
        "witnesses": payload.witnesses,
        "photo_urls": payload.photo_urls,
        "medical_record_urls": payload.medical_record_urls,
        "reported_to_authority": payload.reported_to_authority,
        "authority_report_ref": payload.authority_report_ref,
        "status": "draft",
        "created_at": _now(),
        "created_by": user["id"],
        "updated_at": None,
        "deletedAt": None,
        "deletedBy": None,
        "deletion_reason": None,
        "retention_until": None,
    }
    await db.safety_incidents.insert_one(doc)
    await _audit("safety_incident", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": {k: v for k, v in doc.items() if k != "_id"},
    })
    return SafetyIncident(**doc)


@router.get("/{project_id}/incidents")
async def list_incidents(
    project_id: str,
    incident_type: Optional[str] = None,
    severity: Optional[SafetySeverity] = None,
    reported_to_authority: Optional[bool] = None,
    injured_worker_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if incident_type: q["incident_type"] = incident_type
    if severity:      q["severity"] = severity.value
    if reported_to_authority is not None:
        q["reported_to_authority"] = reported_to_authority
    if injured_worker_id:
        q["injured_worker_id"] = injured_worker_id
    if date_from or date_to:
        rng = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            # pad a date-only "עד תאריך" to end-of-day so it is inclusive.
            # explicit parens: the concat belongs to the true-branch only.
            rng["$lte"] = (date_to + "T23:59:59") if len(date_to) == 10 else date_to
        q["occurred_at"] = rng

    total = await db.safety_incidents.count_documents(q)
    cursor = db.safety_incidents.find(q, {"_id": 0}).sort("occurred_at", -1).skip(offset).limit(limit)
    items = await cursor.to_list(length=limit)
    for it in items:
        it["photo_display_urls"] = [
            (generate_url(k) if k else k) for k in (it.get("photo_urls") or [])
        ]
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{project_id}/incidents/{incident_id}", response_model=SafetyIncident)
async def get_incident(
    project_id: str,
    incident_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_incidents.find_one({"id": incident_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="incident not found")
    doc["photo_display_urls"] = [
        (generate_url(k) if k else k) for k in (doc.get("photo_urls") or [])
    ]
    return SafetyIncident(**doc)


@router.patch("/{project_id}/incidents/{incident_id}", response_model=SafetyIncident)
async def update_incident(
    project_id: str,
    incident_id: str,
    payload: SafetyIncidentUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_incidents.find_one({"id": incident_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="incident not found")
    if before.get("deletedAt"):
        raise HTTPException(status_code=410, detail="incident deleted")

    # Once status != "draft", only project_manager OR safety_officer sub-role may edit
    if before.get("status", "draft") != "draft":
        membership = await _get_project_membership(user, project_id)
        is_pm = membership.get("role") == "project_manager"
        is_safety_officer = membership.get("sub_role") == "safety_officer"
        if not (is_pm or is_safety_officer or _is_super_admin(user)):
            raise HTTPException(
                status_code=403,
                detail="Only project_manager or safety_officer may edit a non-draft incident",
            )

    updates = payload.model_dump(exclude_unset=True)
    if "severity" in updates and updates["severity"] is not None:
        updates["severity"] = updates["severity"].value if hasattr(updates["severity"], "value") else updates["severity"]
    if "incident_type" in updates and updates["incident_type"] is not None:
        if updates["incident_type"] not in VALID_INCIDENT_TYPES:
            raise HTTPException(status_code=400, detail=f"invalid incident_type")
    if "injured_worker_id" in updates and updates["injured_worker_id"]:
        worker = await db.safety_workers.find_one({
            "id": updates["injured_worker_id"], "project_id": project_id, "deletedAt": None,
        })
        if not worker:
            raise HTTPException(status_code=404, detail="injured_worker_id not found or deleted")

    updates["updated_at"] = _now()
    updates["updated_by"] = user["id"]

    await db.safety_incidents.update_one({"id": incident_id, "project_id": project_id, "deletedAt": None}, {"$set": updates})
    after = await db.safety_incidents.find_one({"id": incident_id})
    await _audit("safety_incident", incident_id, "updated", user["id"], {
        "project_id": project_id,
        "before": {k: v for k, v in before.items() if k != "_id"},
        "after": {k: v for k, v in after.items() if k != "_id"},
    })
    return SafetyIncident(**after)


@router.delete("/{project_id}/incidents/{incident_id}", status_code=204)
async def delete_incident(
    project_id: str,
    incident_id: str,
    body: SoftDeleteBody,
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    if not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="deletion_reason is required for incidents")

    before = await db.safety_incidents.find_one({"id": incident_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="incident not found")
    if before.get("deletedAt"):
        return

    # Retention from OCCURRED_AT, not now (regulatory clock)
    retention = _retention_date(before["occurred_at"])
    await db.safety_incidents.update_one({"id": incident_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": _now(),
        "deletedBy": user["id"],
        "deletion_reason": body.reason.strip(),
        "retention_until": retention,
    }})
    await _audit("safety_incident", incident_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason.strip(),
        "retention_until": retention,
    })


# =====================================================================
# Tours CRUD (סיורי בטיחות) — Phase 2 step 4a
# A tour = a checklist walk of the site. Items are pass/fail/na; a FAILED
# item auto-opens a linked safety defect. Signatures are batch 4c (slots only).
# =====================================================================
DEFAULT_TOUR_CHECKLIST = [
    ("פיגומים — תקינות, משטחי עבודה ומאחזי יד", "scaffolding"),
    ("עבודה בגובה — רתמות, קווי חיים, אבטחת פתחים", "heights"),
    ("מעקות ופתחים ברצפות ובקירות", "heights"),
    ("חשמל — לוחות, כבלים מאריכים, הארקות ומפסקי פחת", "electrical_safety"),
    ("ציוד הרמה — עגורנים, אביזרי הרמה ותסקירים", "lifting"),
    ("חפירות ודיפון — יציבות, גישה וגידור", "excavation"),
    ("אש — מטפים, אחסון חומרים דליקים, עבודות ריתוך", "fire_safety"),
    ("ציוד מגן אישי — קסדות, נעלי בטיחות, אפודים זוהרים", "ppe"),
    ("סדר וניקיון — דרכי גישה ופינוי פסולת", "site_housekeeping"),
    ("חומרים מסוכנים — אחסון, שילוט וגיליונות בטיחות", "hazardous_materials"),
    ("שילוט ותמרור באתר", "other"),
    ("דרכי גישה ומילוט תקינות ופנויות", "other"),
    ("עזרה ראשונה — ערכות וזמינות מגיש עזרה ראשונה", "other"),
]


def _regen_tour_photo_urls(tour: dict) -> dict:
    """Regenerate per-GET presigned display URLs for every item's photos.
    Mirrors list_documents/list_incidents: permanent keys persisted, presigned
    URLs computed on read and never stored."""
    for it in (tour.get("items") or []):
        it["photo_display_urls"] = [
            (generate_url(k) if k else k) for k in (it.get("photo_urls") or [])
        ]
    return tour


class SafetyTourCreate(BaseModel):
    tour_type: SafetyTourType
    custom_name: Optional[str] = Field(None, max_length=120)
    tour_date: str                       # "YYYY-MM-DD"

    @model_validator(mode="after")
    def _validate(self):
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", self.tour_date or ""):
            raise ValueError("תאריך סיור לא תקין")
        if self.tour_type == SafetyTourType.custom and not (self.custom_name or "").strip():
            raise ValueError("יש להזין שם לסיור מותאם")
        return self


class SafetyTourMetaUpdate(BaseModel):    # draft-only meta edit
    tour_date: Optional[str] = None
    custom_name: Optional[str] = None

    @model_validator(mode="after")
    def _validate(self):
        if self.tour_date is not None and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", self.tour_date):
            raise ValueError("תאריך סיור לא תקין")
        return self


class SafetyTourItemUpdate(BaseModel):
    result: Literal["pass", "fail", "na"]
    note: Optional[str] = None
    photo_urls: Optional[List[str]] = None       # permanent keys from /upload
    severity: Optional[SafetySeverity] = None    # REQUIRED when result=fail
    defect_title: Optional[str] = Field(None, max_length=200)


class SafetyTourItemAdd(BaseModel):              # ad-hoc item during a walk
    label: str = Field(..., min_length=2, max_length=200)
    category: SafetyCategory = SafetyCategory.other


@router.post("/{project_id}/tours", status_code=201, response_model=SafetyTour)
async def create_tour(
    project_id: str,
    payload: SafetyTourCreate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    items = [
        {
            "id": _new_id(), "label": lbl, "category": cat, "result": None,
            "note": None, "photo_urls": [], "defect_id": None,
            "answered_at": None, "answered_by": None,
        }
        for (lbl, cat) in DEFAULT_TOUR_CHECKLIST
    ]
    custom_name = (
        payload.custom_name.strip()
        if payload.tour_type == SafetyTourType.custom and payload.custom_name
        else None
    )
    doc = {
        "id": _new_id(),
        "project_id": project_id,
        "tour_type": payload.tour_type.value,
        "custom_name": custom_name,
        "tour_date": payload.tour_date,
        "status": SafetyTourStatus.draft.value,
        "items": items,
        "work_manager_signature": None,
        "safety_assistant_signature": None,
        "safety_officer_signature": None,
        "submitted_at": None,
        "signed_at": None,
        "created_at": _now(),
        "created_by": user["id"],
        "updated_at": None,
        "deletedAt": None,
        "deletedBy": None,
        "deletion_reason": None,
        "retention_until": None,
    }
    await db.safety_tours.insert_one(doc)
    await _audit("safety_tour", doc["id"], "created", user["id"], {
        "project_id": project_id, "after": {k: v for k, v in doc.items() if k != "_id"},
    })
    return SafetyTour(**_regen_tour_photo_urls(doc))


@router.get("/{project_id}/tours")
async def list_tours(
    project_id: str,
    status_: Optional[SafetyTourStatus] = Query(None, alias="status"),
    tour_type: Optional[SafetyTourType] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = Query(0, ge=0),
    include_deleted: bool = False,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)

    q = {"project_id": project_id}
    include_deleted = _resolve_include_deleted(include_deleted, user)
    if not include_deleted:
        q["deletedAt"] = None
    if status_:    q["status"] = status_.value
    if tour_type:  q["tour_type"] = tour_type.value
    if date_from or date_to:
        rng = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            rng["$lte"] = (date_to + "T23:59:59") if len(date_to) == 10 else date_to
        q["tour_date"] = rng

    total = await db.safety_tours.count_documents(q)
    cursor = (
        db.safety_tours.find(q, {"_id": 0})
        .sort([("tour_date", -1), ("created_at", -1)])
        .skip(offset).limit(limit)
    )
    items = await cursor.to_list(length=limit)
    for t in items:
        _regen_tour_photo_urls(t)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{project_id}/tours/{tour_id}", response_model=SafetyTour)
async def get_tour(
    project_id: str,
    tour_id: str,
    user: dict = Depends(get_current_user),
):
    db = get_db()
    await _check_project_access(user, project_id)
    doc = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not doc:
        raise HTTPException(status_code=404, detail="tour not found")
    return SafetyTour(**_regen_tour_photo_urls(doc))


@router.patch("/{project_id}/tours/{tour_id}", response_model=SafetyTour)
async def update_tour_meta(
    project_id: str,
    tour_id: str,
    payload: SafetyTourMetaUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    before = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="tour not found")
    if before.get("status") != "draft":
        raise HTTPException(status_code=409, detail="לא ניתן לערוך סיור שהוגש")

    updates = payload.model_dump(exclude_unset=True)
    updates["updated_at"] = _now()
    await db.safety_tours.update_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None},
        {"$set": updates},
    )
    after = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id})
    await _audit("safety_tour", tour_id, "updated", user["id"], {
        "project_id": project_id, "changes": {k: v for k, v in updates.items() if k != "updated_at"},
    })
    return SafetyTour(**_regen_tour_photo_urls(after))


@router.patch("/{project_id}/tours/{tour_id}/items/{item_id}", response_model=SafetyTour)
async def update_tour_item(
    project_id: str,
    tour_id: str,
    item_id: str,
    payload: SafetyTourItemUpdate,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    tour = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not tour:
        raise HTTPException(status_code=404, detail="tour not found")
    if tour.get("status") != "draft":
        raise HTTPException(status_code=409, detail="לא ניתן לערוך סיור שהוגש")

    item = next((it for it in (tour.get("items") or []) if it.get("id") == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="item not found")

    if payload.result == "fail" and payload.severity is None:
        raise HTTPException(status_code=422, detail="יש לבחור חומרה לפריט שנכשל")

    # 1) Targeted arrayFilters update of the item fields (never a whole-items
    #    rewrite — two walkers editing different items must not clobber each other).
    item_set = {
        "items.$[elem].result": payload.result,
        "items.$[elem].answered_at": _now(),
        "items.$[elem].answered_by": user["id"],
        "updated_at": _now(),
    }
    if payload.note is not None:
        item_set["items.$[elem].note"] = payload.note
    if payload.photo_urls is not None:
        item_set["items.$[elem].photo_urls"] = payload.photo_urls
    upd = await db.safety_tours.update_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None, "status": "draft"},
        {"$set": item_set},
        array_filters=[{"elem.id": item_id}],
    )
    # The top-of-handler read saw draft, but a concurrent submit/delete could
    # have moved the tour out of draft between that read and this write. If the
    # guarded update matched nothing, refuse — never auto-defect on a tour that
    # is no longer editable.
    if upd.matched_count == 0:
        raise HTTPException(status_code=409, detail="לא ניתן לערוך סיור שהוגש")

    # 2) AUTO-DEFECT: a failed item opens exactly one linked safety defect.
    #    Order per spec: item fields first (above), THEN the atomic claim
    #    (pre-gen id + $elemMatch defect_id:None) guarded to the same draft
    #    lifecycle, then insert only if we won the claim (modified_count == 1)
    #    — no double defect on concurrent fails; fail→pass keeps the defect
    #    (human closes it later on the defects tab).
    if payload.result == "fail" and not item.get("defect_id"):
        defect_id = _new_id()
        claim = await db.safety_tours.update_one(
            {"id": tour_id, "project_id": project_id, "deletedAt": None, "status": "draft",
             "items": {"$elemMatch": {"id": item_id, "defect_id": None}}},
            {"$set": {"items.$.defect_id": defect_id}},
        )
        if claim.modified_count == 1:
            photo_keys = (
                payload.photo_urls if payload.photo_urls is not None
                else (item.get("photo_urls") or [])
            )
            title = (payload.defect_title or f"סיור בטיחות: {item['label']}")[:200]
            defect = {
                "id": defect_id,
                "project_id": project_id,
                "kind": "defect",
                "category": item["category"],
                "severity": payload.severity.value,
                "status": SafetyDocumentStatus.open.value,
                "title": title,
                "description": payload.note,
                "location": None,
                "company_id": None,
                "profession": None,
                "assignee_id": None,
                "reporter_id": user["id"],
                "photo_urls": list(photo_keys),
                "attachment_urls": [],
                "found_at": _now(),
                "resolved_at": None,
                "created_at": _now(),
                "created_by": user["id"],
                "updated_at": None,
                "deletedAt": None,
                "deletedBy": None,
                "tour_id": tour_id,
                "tour_item_id": item_id,
            }
            # Cross-collection compensation: we already claimed defect_id on the
            # item. If the defect insert fails, roll the claim back to None so a
            # retry can re-open the defect — otherwise the item would point at a
            # defect that does not exist ("ONE linked defect" must stay durable).
            try:
                await db.safety_documents.insert_one(defect)
            except Exception:
                await db.safety_tours.update_one(
                    {"id": tour_id, "project_id": project_id,
                     "items": {"$elemMatch": {"id": item_id, "defect_id": defect_id}}},
                    {"$set": {"items.$.defect_id": None}},
                )
                raise
            await _audit("safety_document", defect_id, "created", user["id"], {
                "project_id": project_id, "source": "safety_tour", "tour_id": tour_id,
                "after": {k: v for k, v in defect.items() if k != "_id"},
            })

    after = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id})
    after_item = next((it for it in (after.get("items") or []) if it.get("id") == item_id), {})
    await _audit("safety_tour", tour_id, "item_updated", user["id"], {
        "project_id": project_id, "item_id": item_id,
        "result": payload.result, "defect_id": after_item.get("defect_id"),
    })
    return SafetyTour(**_regen_tour_photo_urls(after))


@router.post("/{project_id}/tours/{tour_id}/items", response_model=SafetyTour)
async def add_tour_item(
    project_id: str,
    tour_id: str,
    payload: SafetyTourItemAdd,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    tour = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not tour:
        raise HTTPException(status_code=404, detail="tour not found")
    if tour.get("status") != "draft":
        raise HTTPException(status_code=409, detail="לא ניתן לערוך סיור שהוגש")

    new_item = {
        "id": _new_id(), "label": payload.label.strip(), "category": payload.category.value,
        "result": None, "note": None, "photo_urls": [], "defect_id": None,
        "answered_at": None, "answered_by": None,
    }
    upd = await db.safety_tours.update_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None, "status": "draft"},
        {"$push": {"items": new_item}, "$set": {"updated_at": _now()}},
    )
    # The top-of-handler read saw draft, but a concurrent submit could have moved
    # the tour out of draft before this $push landed — refuse rather than silently
    # no-op with a 200 + "item_added" audit (mirrors update_tour_item's guard).
    if upd.matched_count == 0:
        raise HTTPException(status_code=409, detail="לא ניתן לערוך סיור שהוגש")
    after = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id})
    await _audit("safety_tour", tour_id, "item_added", user["id"], {
        "project_id": project_id, "item_id": new_item["id"], "label": new_item["label"],
    })
    return SafetyTour(**_regen_tour_photo_urls(after))


@router.post("/{project_id}/tours/{tour_id}/submit", response_model=SafetyTour)
async def submit_tour(
    project_id: str,
    tour_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    tour = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not tour:
        raise HTTPException(status_code=404, detail="tour not found")
    if tour.get("status") != "draft":
        raise HTTPException(status_code=409, detail="הסיור כבר הוגש")
    if any(it.get("result") is None for it in (tour.get("items") or [])):
        raise HTTPException(status_code=422, detail="יש לענות על כל הפריטים לפני הגשה")

    now = _now()
    await db.safety_tours.update_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None},
        {"$set": {"status": "pending_signature", "submitted_at": now, "updated_at": now}},
    )
    after = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id})
    await _audit("safety_tour", tour_id, "submitted", user["id"], {"project_id": project_id})
    return SafetyTour(**_regen_tour_photo_urls(after))


@router.post("/{project_id}/tours/{tour_id}/reopen", response_model=SafetyTour)
async def reopen_tour(
    project_id: str,
    tour_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)
    tour = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not tour:
        raise HTTPException(status_code=404, detail="tour not found")
    if tour.get("status") != "pending_signature":
        raise HTTPException(status_code=409, detail="לא ניתן לפתוח מחדש")

    await db.safety_tours.update_one(
        {"id": tour_id, "project_id": project_id, "deletedAt": None},
        {"$set": {"status": "draft", "submitted_at": None, "updated_at": _now()}},
    )
    after = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id})
    await _audit("safety_tour", tour_id, "reopened", user["id"], {"project_id": project_id})
    return SafetyTour(**_regen_tour_photo_urls(after))


@router.delete("/{project_id}/tours/{tour_id}", status_code=204)
async def delete_tour(
    project_id: str,
    tour_id: str,
    body: SoftDeleteBody,
    user: dict = Depends(require_roles(*SAFETY_DELETERS)),
):
    db = get_db()
    await _check_project_access(user, project_id)

    if not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="deletion_reason is required for tours")

    before = await db.safety_tours.find_one({"id": tour_id, "project_id": project_id, "deletedAt": None})
    if not before:
        raise HTTPException(status_code=404, detail="tour not found")
    if before.get("deletedAt"):
        return

    now = _now()
    retention = _retention_date(now)
    await db.safety_tours.update_one({"id": tour_id, "project_id": project_id, "deletedAt": None}, {"$set": {
        "deletedAt": now,
        "deletedBy": user["id"],
        "deletion_reason": body.reason.strip(),
        "retention_until": retention,
    }})
    await _audit("safety_tour", tour_id, "deleted", user["id"], {
        "project_id": project_id, "deletion_reason": body.reason.strip(),
        "retention_until": retention,
    })


# =====================================================================
# Score & Exports — Phase 1 Part 3 (Backend Advanced)
# =====================================================================
_SCORE_CACHE: dict = {}  # {project_id: (timestamp, payload)}
_SCORE_TTL_SECONDS = 300  # 5 minutes

# Capped penalty buckets (spec §Steps.1)
_SCORE_DOCS_MAX = 40
_SCORE_TASKS_MAX = 25
_SCORE_TRAINING_MAX = 20
_SCORE_INCIDENTS_MAX = 15

# Safety score — required training types per worker (spec §Steps.1)
REQUIRED_TRAINING_TYPES = frozenset({"safety_induction", "height_work", "electrical"})


async def _compute_safety_score(db, project_id: str) -> dict:
    """
    Compute 0-100 safety score from live state via four capped penalty buckets:

      - documents bucket  (≤ _SCORE_DOCS_MAX = 40)
            sev3*10 + sev2*5 + sev1*2 over open/in_progress documents
      - tasks bucket      (≤ _SCORE_TASKS_MAX = 25)
            overdue (due_at past, status not completed/cancelled): 5 each
      - training bucket   (≤ _SCORE_TRAINING_MAX = 20)
            untrained workers: 4 each (no in-force training)
      - incidents bucket  (≤ _SCORE_INCIDENTS_MAX = 15)
            incidents in last 90 days: 8 each

    score = max(0, 100 - sum_of_capped_penalties).

    All Mongo reads filter `deletedAt: None`. Documents and tasks use
    `$group/$cond` aggregation in the style of snapshot_cron.py.
    """
    now = datetime.now(timezone.utc)
    today_iso = now.date().isoformat()
    now_iso = now.isoformat()
    cutoff_90d = (now - timedelta(days=90)).isoformat()

    # Documents bucket — single $group with $cond per severity (unchanged)
    docs_agg = await db.safety_documents.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "kind": {"$ne": "observation"},
            "status": {"$in": ["open", "in_progress"]},
        }},
        {"$group": {
            "_id": None,
            "sev3": {"$sum": {"$cond": [{"$eq": ["$severity", "3"]}, 1, 0]}},
            "sev2": {"$sum": {"$cond": [{"$eq": ["$severity", "2"]}, 1, 0]}},
            "sev1": {"$sum": {"$cond": [{"$eq": ["$severity", "1"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    sev3 = docs_agg[0].get("sev3", 0) if docs_agg else 0
    sev2 = docs_agg[0].get("sev2", 0) if docs_agg else 0
    sev1 = docs_agg[0].get("sev1", 0) if docs_agg else 0
    docs_raw = sev3 * 10 + sev2 * 5 + sev1 * 2
    docs_penalty = min(docs_raw, _SCORE_DOCS_MAX)

    # Tasks bucket — overdue tasks grouped by severity (spec §Steps.1)
    # Weights: sev3=6, sev2=3, sev1=1.5. Sum capped at _SCORE_TASKS_MAX (25).
    tasks_agg = await db.safety_tasks.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "due_at": {"$ne": None, "$lt": now_iso},
            "status": {"$nin": ["completed", "cancelled"]},
        }},
        {"$group": {
            "_id": None,
            "sev3": {"$sum": {"$cond": [{"$eq": ["$severity", "3"]}, 1, 0]}},
            "sev2": {"$sum": {"$cond": [{"$eq": ["$severity", "2"]}, 1, 0]}},
            "sev1": {"$sum": {"$cond": [{"$eq": ["$severity", "1"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    t_sev3 = tasks_agg[0].get("sev3", 0) if tasks_agg else 0
    t_sev2 = tasks_agg[0].get("sev2", 0) if tasks_agg else 0
    t_sev1 = tasks_agg[0].get("sev1", 0) if tasks_agg else 0
    tasks_raw = t_sev3 * 6 + t_sev2 * 3 + t_sev1 * 1.5
    tasks_penalty = min(tasks_raw, _SCORE_TASKS_MAX)
    overdue_tasks_total = t_sev3 + t_sev2 + t_sev1

    # Incidents bucket — last 90d grouped by type (spec §Steps.1)
    # Weights: injury=5, property_damage=3, near_miss=1. Cap _SCORE_INCIDENTS_MAX (15).
    inc_agg = await db.safety_incidents.aggregate([
        {"$match": {
            "project_id": project_id,
            "deletedAt": None,
            "occurred_at": {"$gte": cutoff_90d},
        }},
        {"$group": {
            "_id": None,
            "injury":          {"$sum": {"$cond": [{"$eq": ["$incident_type", "injury"]}, 1, 0]}},
            "property_damage": {"$sum": {"$cond": [{"$eq": ["$incident_type", "property_damage"]}, 1, 0]}},
            "near_miss":       {"$sum": {"$cond": [{"$eq": ["$incident_type", "near_miss"]}, 1, 0]}},
        }},
    ]).to_list(length=1)
    inc_injury   = inc_agg[0].get("injury", 0) if inc_agg else 0
    inc_property = inc_agg[0].get("property_damage", 0) if inc_agg else 0
    inc_near     = inc_agg[0].get("near_miss", 0) if inc_agg else 0
    incidents_raw = inc_injury * 5 + inc_property * 3 + inc_near * 1
    incidents_penalty = min(incidents_raw, _SCORE_INCIDENTS_MAX)
    recent_incidents_total = inc_injury + inc_property + inc_near

    # Training bucket — per worker, check each REQUIRED_TRAINING_TYPE has in-force record
    # (spec §Steps.1: 4pts per expired-required-training, 5pts for workers with NO training record)
    worker_ids_docs = await db.safety_workers.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0, "id": 1}
    ).to_list(length=10000)
    worker_ids = [w["id"] for w in worker_ids_docs]

    workers_with_expired_training = 0
    workers_without_training = 0
    expired_required_count = 0

    if worker_ids:
        in_force = await db.safety_trainings.find(
            {
                "project_id": project_id,
                "deletedAt": None,
                "worker_id": {"$in": worker_ids},
                "training_type": {"$in": list(REQUIRED_TRAINING_TYPES)},
                "$or": [{"expires_at": None}, {"expires_at": {"$gte": today_iso}}],
            },
            {"_id": 0, "worker_id": 1, "training_type": 1},
        ).to_list(length=100000)

        any_training_worker_ids = set(
            r["worker_id"] for r in await db.safety_trainings.find(
                {"project_id": project_id, "deletedAt": None, "worker_id": {"$in": worker_ids}},
                {"_id": 0, "worker_id": 1},
            ).to_list(length=100000)
        )

        worker_inforce: dict = {wid: set() for wid in worker_ids}
        for r in in_force:
            worker_inforce[r["worker_id"]].add(r["training_type"])

        for wid in worker_ids:
            if wid not in any_training_worker_ids:
                workers_without_training += 1
                continue
            missing_types = REQUIRED_TRAINING_TYPES - worker_inforce[wid]
            if missing_types:
                workers_with_expired_training += 1
                expired_required_count += len(missing_types)

    training_raw = expired_required_count * 4 + workers_without_training * 5
    training_penalty = min(training_raw, _SCORE_TRAINING_MAX)

    total_penalty = docs_penalty + tasks_penalty + training_penalty + incidents_penalty
    score = max(0, int(round(100 - total_penalty)))

    return {
        "score": score,
        "breakdown": {
            "doc_penalty": round(docs_penalty, 2),
            "task_penalty": round(tasks_penalty, 2),
            "training_penalty": round(training_penalty, 2),
            "incident_penalty": round(incidents_penalty, 2),
            "doc_counts": {"sev3": sev3, "sev2": sev2, "sev1": sev1},
            "overdue_task_counts": {
                "sev3": t_sev3, "sev2": t_sev2, "sev1": t_sev1,
                "total": overdue_tasks_total,
            },
            "workers_with_expired_training": workers_with_expired_training,
            "workers_without_training": workers_without_training,
            "total_workers": len(worker_ids),
            "incidents_last_90d": {
                "injury": inc_injury,
                "property_damage": inc_property,
                "near_miss": inc_near,
                "total": recent_incidents_total,
            },
            "caps": {
                "doc_max": _SCORE_DOCS_MAX,
                "task_max": _SCORE_TASKS_MAX,
                "training_max": _SCORE_TRAINING_MAX,
                "incident_max": _SCORE_INCIDENTS_MAX,
            },
        },
        "computed_at": now_iso,
    }


def _strip_pii(records: list) -> list:
    """Remove PII fields from worker records before any export."""
    cleaned = []
    for r in records:
        c = {k: v for k, v in r.items() if k not in ("_id", "id_number", "id_number_hash")}
        cleaned.append(c)
    return cleaned


@router.get("/{project_id}/score")
async def get_safety_score(
    project_id: str,
    refresh: bool = Query(False),
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    Return safety score with spec-required response shape:
      {score, breakdown, computed_at, cache_age_seconds}

    5-min in-process cache. `?refresh=true` recomputes. Always audited
    as `safety_score:computed` with `{score, cache_hit}`.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    cached = _SCORE_CACHE.get(project_id)
    cache_hit = False
    if cached and not refresh and (_time.time() - cached[0]) < _SCORE_TTL_SECONDS:
        payload = cached[1]
        cache_age = int(_time.time() - cached[0])
        cache_hit = True
    else:
        payload = await _compute_safety_score(db, project_id)
        _SCORE_CACHE[project_id] = (_time.time(), payload)
        cache_age = 0

    await _audit("safety_score", project_id, "computed", user["id"], {
        "project_id": project_id,
        "score": payload["score"],
        "cache_hit": cache_hit,
    })
    return {**payload, "cache_age_seconds": cache_age}


async def _gather_export_data(db, project_id: str, doc_filter_extra: Optional[dict] = None):
    """Fetch all data needed for exports. Strips PII from workers."""
    workers = await db.safety_workers.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)
    workers = _strip_pii(workers)

    trainings = await db.safety_trainings.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    doc_q = {"project_id": project_id, "deletedAt": None}
    if doc_filter_extra:
        doc_q.update(doc_filter_extra)
    documents = await db.safety_documents.find(doc_q, {"_id": 0}).to_list(length=100000)

    tasks = await db.safety_tasks.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    incidents = await db.safety_incidents.find(
        {"project_id": project_id, "deletedAt": None}, {"_id": 0}
    ).to_list(length=100000)

    company_ids = {r.get("company_id") for r in workers + documents + tasks if r.get("company_id")}
    company_map = {}
    if company_ids:
        companies = await db.project_companies.find(
            {"id": {"$in": list(company_ids)}, "deletedAt": None},
            {"_id": 0, "id": 1, "name": 1},
        ).to_list(length=10000)
        company_map = {c["id"]: c.get("name", "") for c in companies}

    user_ids = set()
    for src in (documents, tasks, incidents):
        for r in src:
            for f in ("assignee_id", "reporter_id", "created_by"):
                if r.get(f):
                    user_ids.add(r[f])
    user_map = {}
    if user_ids:
        users = await db.users.find(
            {"id": {"$in": list(user_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(length=10000)
        user_map = {u["id"]: u.get("name", "") for u in users}

    return workers, trainings, documents, tasks, incidents, company_map, user_map


_EXCEL_HEADER_FONT_KW = dict(name="Arial", bold=True, size=11, color="FFFFFF")
_EXCEL_HEADER_FILL_KW = dict(start_color="475569", end_color="475569", fill_type="solid")


def _excel_styles():
    """Cached style objects for our Excel sheets (header grey, RTL alignment)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        "header_font": Font(**_EXCEL_HEADER_FONT_KW),
        "header_fill": PatternFill(**_EXCEL_HEADER_FILL_KW),
        "header_align": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "cell_font": Font(name="Arial", size=10),
        "cell_align": Alignment(horizontal="right", vertical="top", wrap_text=True),
        "thin": Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        ),
    }


def _fmt_dt(s):
    if not s:
        return ""
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(s)[:16]


def _write_sheet(wb, title, headers, rows, widths):
    """
    Add an RTL Hebrew sheet. Bold grey header, frozen row 1, optional
    auto-filter when there are data rows.
    """
    from openpyxl.utils import get_column_letter
    s = _excel_styles()
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["header_align"]
        c.border = s["thin"]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = s["cell_font"]
            c.alignment = s["cell_align"]
            c.border = s["thin"]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Spec: frozen row 1
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def _build_safety_excel_3sheet(
    workers, trainings, documents, incidents, company_map, worker_name_map, user_map
):
    """
    Build the regulatory RTL Hebrew workbook (spec §Steps.2). NOTE: the
    function name is a legacy misnomer — it now emits FOUR sheets:
      Sheet 1: עובדים  (Workers — NO id_number / id_number_hash)
      Sheet 2: הדרכות  (Trainings)
      Sheet 3: ליקויים (Safety findings / documents)
      Sheet 4: אירועים (Incidents)
    Frozen row 1 on every sheet.
    """
    from openpyxl import Workbook
    from services.safety_pdf import (
        SEVERITY_HE, INCIDENT_TYPE_HE, CATEGORY_HE, DOC_STATUS_HE,
    )

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: workers (PII already stripped upstream)
    worker_rows = []
    for w in workers:
        worker_rows.append([
            w.get("full_name", ""),
            w.get("profession", "") or "",
            company_map.get(w.get("company_id", ""), ""),
            w.get("phone", "") or "",
            _fmt_dt(w.get("created_at")),
        ])
    _write_sheet(
        wb, "עובדים",
        ["שם מלא", "מקצוע", "חברה", "טלפון", "תאריך כניסה"],
        worker_rows,
        [22, 18, 20, 14, 18],
    )

    # Sheet 2: trainings
    training_rows = []
    for tr in trainings:
        training_rows.append([
            worker_name_map.get(tr.get("worker_id", ""), ""),
            tr.get("training_type", "") or "",
            tr.get("instructor_name", "") or "",
            _fmt_dt(tr.get("trained_at")),
            _fmt_dt(tr.get("expires_at")),
        ])
    _write_sheet(
        wb, "הדרכות",
        ["עובד", "סוג הדרכה", "מדריך", "תאריך הדרכה", "תוקף עד"],
        training_rows,
        [22, 22, 18, 18, 18],
    )

    # Split into defects (ליקויים) and observations (תיעוד) — one collection,
    # kind discriminator. Legacy docs (no kind) count as defects.
    defects = [d for d in documents if d.get("kind") != "observation"]
    observations = [d for d in documents if d.get("kind") == "observation"]

    # Sheet 3: ליקויים (safety findings / documents) — mirrors the filtered
    # documents export row logic. No photo_urls (S3 keys, not human-readable).
    document_rows = []
    for d in defects:
        document_rows.append([
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            SEVERITY_HE.get(d.get("severity", ""), ""),
            DOC_STATUS_HE.get(d.get("status", ""), d.get("status", "")),
            _fmt_dt(d.get("found_at")),
            d.get("location", "") or "",
            company_map.get(d.get("company_id", ""), ""),
            user_map.get(d.get("assignee_id", ""), ""),
            d.get("description", "") or "",
        ])
    _write_sheet(
        wb, "ליקויים",
        ["כותרת", "קטגוריה", "חומרה", "סטטוס", "תאריך גילוי",
         "מיקום", "חברה", "אחראי", "תיאור"],
        document_rows,
        [26, 16, 12, 12, 16, 20, 20, 18, 40],
    )

    # Sheet: תיעוד (observations — archival, no severity/status)
    observation_rows = []
    for d in observations:
        observation_rows.append([
            _fmt_dt(d.get("found_at")),
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            d.get("location", "") or "",
            d.get("description", "") or "",
            user_map.get(d.get("reporter_id", ""), ""),
        ])
    _write_sheet(
        wb, "תיעוד",
        ["תאריך", "כותרת", "קטגוריה", "מיקום", "תיאור", "דווח ע\"י"],
        observation_rows,
        [16, 26, 16, 20, 40, 18],
    )

    # Sheet 4: incidents
    incident_rows = []
    for inc in incidents:
        incident_rows.append([
            INCIDENT_TYPE_HE.get(inc.get("incident_type", ""), inc.get("incident_type", "")),
            SEVERITY_HE.get(inc.get("severity", ""), ""),
            _fmt_dt(inc.get("occurred_at")),
            inc.get("location", "") or "",
            worker_name_map.get(inc.get("injured_worker_id", ""), "") or "",
            "כן" if inc.get("reported_to_authority") else "לא",
            inc.get("description", "") or "",
        ])
    _write_sheet(
        wb, "אירועים",
        ["סוג אירוע", "חומרה", "תאריך אירוע", "מיקום",
         "עובד נפגע", "דווח לרשות", "תיאור"],
        incident_rows,
        [16, 12, 18, 20, 22, 14, 40],
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_filtered_documents_excel(documents, company_map, user_map):
    """
    Build single-sheet 'ממצאי בטיחות' workbook (spec §Steps.3) of
    documents matching the 7-dim filter. Empty list still produces a
    valid Excel with header only.
    """
    from openpyxl import Workbook
    from services.safety_pdf import CATEGORY_HE, SEVERITY_HE, DOC_STATUS_HE

    wb = Workbook()
    wb.remove(wb.active)

    rows = []
    for d in documents:
        rows.append([
            d.get("title", ""),
            CATEGORY_HE.get(d.get("category", ""), d.get("category", "")),
            SEVERITY_HE.get(d.get("severity", ""), ""),
            DOC_STATUS_HE.get(d.get("status", ""), d.get("status", "")),
            d.get("location", "") or "",
            company_map.get(d.get("company_id", ""), ""),
            user_map.get(d.get("assignee_id", ""), ""),
            user_map.get(d.get("reporter_id", ""), ""),
            _fmt_dt(d.get("found_at")),
            _fmt_dt(d.get("created_at")),
            _fmt_dt(d.get("resolved_at")),
            d.get("description", "") or "",
        ])
    _write_sheet(
        wb, "ממצאי בטיחות",
        ["כותרת", "קטגוריה", "חומרה", "סטטוס", "מיקום", "חברה",
         "אחראי", "מדווח", "נמצא בתאריך", "נוצר", "נפתר", "תיאור"],
        rows,
        [25, 16, 10, 12, 18, 18, 16, 16, 18, 18, 18, 35],
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/{project_id}/export/excel")
async def export_safety_excel(
    project_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    4-sheet Hebrew RTL Excel export (workers / trainings / ליקויים / incidents).
    Management-only. PII (id_number / id_number_hash) is stripped.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    workers, trainings, documents, tasks, incidents, company_map, user_map = \
        await _gather_export_data(db, project_id)
    worker_name_map = {w["id"]: w.get("full_name", "") for w in workers}

    buf = _build_safety_excel_3sheet(
        workers, trainings, documents, incidents, company_map, worker_name_map, user_map
    )

    await _audit("safety_export", project_id, "excel_exported", user["id"], {
        "project_id": project_id,
        "workers": len(workers),
        "trainings": len(trainings),
        "documents": len(documents),
        "incidents": len(incidents),
    })

    filename = (
        f"safety_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/filtered")
async def export_safety_filtered(
    project_id: str,
    kind: Optional[SafetyDocumentKind] = None,
    category: Optional[SafetyCategory] = None,
    severity: Optional[SafetySeverity] = None,
    status_: Optional[SafetyDocumentStatus] = Query(None, alias="status"),
    company_id: Optional[str] = None,
    assignee_id: Optional[str] = None,
    reporter_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """
    Single-sheet 'ממצאי בטיחות' Excel of documents matching the same
    7-dim filter as list_documents. Empty result returns a valid Excel
    (200) with the header row only.
    """
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    extra = {}
    if kind == SafetyDocumentKind.observation:
        extra["kind"] = "observation"
    elif kind == SafetyDocumentKind.defect:
        extra["kind"] = {"$ne": "observation"}
    if category:    extra["category"] = category.value
    if severity:    extra["severity"] = severity.value
    if status_:     extra["status"] = status_.value
    if company_id:  extra["company_id"] = company_id
    if assignee_id: extra["assignee_id"] = assignee_id
    if reporter_id: extra["reporter_id"] = reporter_id
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to
        extra["found_at"] = rng

    doc_q = {"project_id": project_id, "deletedAt": None, **extra}
    documents = await db.safety_documents.find(doc_q, {"_id": 0}).to_list(length=100000)

    company_ids = {d.get("company_id") for d in documents if d.get("company_id")}
    company_map = {}
    if company_ids:
        companies = await db.project_companies.find(
            {"id": {"$in": list(company_ids)}, "deletedAt": None},
            {"_id": 0, "id": 1, "name": 1},
        ).to_list(length=10000)
        company_map = {c["id"]: c.get("name", "") for c in companies}

    user_ids = set()
    for d in documents:
        for f in ("assignee_id", "reporter_id"):
            if d.get(f):
                user_ids.add(d[f])
    user_map = {}
    if user_ids:
        users = await db.users.find(
            {"id": {"$in": list(user_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(length=10000)
        user_map = {u["id"]: u.get("name", "") for u in users}

    buf = _build_filtered_documents_excel(documents, company_map, user_map)

    # Flatten $-operator values (found_at range, kind $ne) into plain keys so the
    # audit payload never contains $-prefixed field names (MongoDB rejects those).
    applied = {k: v for k, v in extra.items() if k not in ("found_at", "kind")}
    if kind:
        applied["kind"] = kind.value
    if date_from:
        applied["date_from"] = date_from
    if date_to:
        applied["date_to"] = date_to

    await _audit("safety_export", project_id, "filtered_exported", user["id"], {
        "project_id": project_id,
        "filters": applied,
        "matched_documents": len(documents),
    })

    filename = (
        f"safety_filtered_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export/pdf-register")
async def export_safety_pdf_register(
    project_id: str,
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    """Generate the 9-section Hebrew 'פנקס כללי' PDF. Management-only."""
    db = get_db()
    await _check_project_access(user, project_id)

    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="project not found")

    from services.safety_pdf import generate_safety_register
    pdf_bytes = await generate_safety_register(db, project_id)

    await _audit("safety_export", project_id, "pdf_register_exported", user["id"], {
        "project_id": project_id,
        "size_bytes": len(pdf_bytes),
    })

    filename = (
        f"safety_register_{project_id[:8]}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =====================================================================
# Photo / PDF upload helper (one endpoint for all resources)
# =====================================================================
@router.post("/{project_id}/upload")
async def upload_safety_file(
    project_id: str,
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_roles(*SAFETY_WRITERS)),
):
    # BATCH upload-hardening-phase3c-safety-imports (2026-05-24) —
    # per-user count limit + Content-Length pre-check (10MB cap =
    # module-level MAX_SAFETY_UPLOAD_SIZE).
    check_upload_rate_limit(user['id'])
    check_content_length(request.headers.get('content-length'), MAX_SAFETY_UPLOAD_SIZE)
    db = get_db()
    await _check_project_access(user, project_id)

    validate_upload(file, SAFETY_ALLOWED_EXTENSIONS, SAFETY_ALLOWED_CONTENT_TYPES)

    content = await file.read()
    if len(content) > MAX_SAFETY_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="הקובץ גדול מדי (מקסימום 10MB)")

    # BATCH upload-hardening-phase3c-safety-imports (2026-05-24) —
    # byte rate limit + per-org storage quota, BEFORE the file
    # reaches S3.
    check_upload_bytes(user['id'], len(content))
    _proj = await db.projects.find_one({'id': project_id}, {'_id': 0, 'org_id': 1})
    _org_id = (_proj or {}).get('org_id')
    await check_storage_quota(_org_id, len(content))

    content_type = file.content_type or "application/octet-stream"
    if content_type == "application/pdf":
        ext = "pdf"
    else:
        ext = content_type.split("/")[-1]
        if ext == "jpeg":
            ext = "jpg"

    file_id = _new_id()
    key = f"safety/{project_id}/{file_id}.{ext}"

    from services.object_storage import save_bytes as obj_save_bytes, generate_url as obj_generate_url
    stored_ref = obj_save_bytes(content, key, content_type)
    # BATCH upload-hardening-phase3c-safety-imports (2026-05-24) — ledger
    await record_upload(_org_id, len(content))
    url = obj_generate_url(stored_ref)

    await _audit("safety_upload", file_id, "upload", user["id"], {
        "project_id": project_id,
        "filename": file.filename,
        "key": key,
        "size": len(content),
        "content_type": content_type,
    })

    return {
        "id": file_id,
        "url": url,
        "stored_ref": stored_ref,
        "filename": file.filename,
        "content_type": content_type,
        "size": len(content),
    }


# =====================================================================
# Index management — called from server.py startup when module enabled.
# =====================================================================
async def ensure_safety_indexes(db) -> None:
    """
    Create MongoDB indices for Safety collections.
    Idempotent: `create_index` is a no-op if the same spec already exists.
    Safe to run on every startup.

    All indices use background=True to avoid locking hot collections.
    """
    # safety_workers — 3
    await db.safety_workers.create_index(
        [("project_id", 1), ("deletedAt", 1)],
        background=True, name="idx_sw_project_deleted",
    )
    await db.safety_workers.create_index(
        [("project_id", 1), ("company_id", 1)],
        background=True, name="idx_sw_project_company",
    )
    await db.safety_workers.create_index(
        [("project_id", 1), ("id_number", 1)],
        background=True, sparse=True, name="idx_sw_project_idnum",
    )
    # safety_trainings — 3
    await db.safety_trainings.create_index(
        [("project_id", 1), ("worker_id", 1)],
        background=True, name="idx_st_project_worker",
    )
    await db.safety_trainings.create_index(
        [("project_id", 1), ("expires_at", 1)],
        background=True, sparse=True, name="idx_st_project_expires",
    )
    await db.safety_trainings.create_index(
        [("project_id", 1), ("training_type", 1)],
        background=True, name="idx_st_project_type",
    )
    # safety_documents — 5
    await db.safety_documents.create_index(
        [("project_id", 1), ("deletedAt", 1), ("created_at", -1)],
        background=True, name="idx_sd_project_deleted_created",
    )
    await db.safety_documents.create_index(
        [("project_id", 1), ("status", 1)],
        background=True, name="idx_sd_project_status",
    )
    await db.safety_documents.create_index(
        [("project_id", 1), ("severity", 1)],
        background=True, name="idx_sd_project_severity",
    )
    await db.safety_documents.create_index(
        [("project_id", 1), ("category", 1)],
        background=True, name="idx_sd_project_category",
    )
    await db.safety_documents.create_index(
        [("project_id", 1), ("company_id", 1)],
        background=True, sparse=True, name="idx_sd_project_company",
    )
    # safety_tasks — 4
    await db.safety_tasks.create_index(
        [("project_id", 1), ("deletedAt", 1), ("due_at", 1)],
        background=True, name="idx_stk_project_deleted_due",
    )
    await db.safety_tasks.create_index(
        [("project_id", 1), ("status", 1)],
        background=True, name="idx_stk_project_status",
    )
    await db.safety_tasks.create_index(
        [("project_id", 1), ("assignee_id", 1)],
        background=True, sparse=True, name="idx_stk_project_assignee",
    )
    await db.safety_tasks.create_index(
        [("document_id", 1)],
        background=True, sparse=True, name="idx_stk_document",
    )
    # safety_incidents — 3 (7yr retention-critical)
    await db.safety_incidents.create_index(
        [("project_id", 1), ("occurred_at", -1)],
        background=True, name="idx_si_project_occurred",
    )
    await db.safety_incidents.create_index(
        [("project_id", 1), ("severity", 1)],
        background=True, name="idx_si_project_severity",
    )
    await db.safety_incidents.create_index(
        [("retention_until", 1)],
        background=True, sparse=True, name="idx_si_retention",
    )
    # project_companies — 2 safety-related (shared collection)
    await db.project_companies.create_index(
        [("project_id", 1), ("safety_contact_id", 1)],
        background=True, sparse=True, name="idx_pc_project_safety_contact",
    )
    await db.project_companies.create_index(
        [("project_id", 1), ("is_placeholder", 1)],
        background=True, sparse=True, name="idx_pc_project_placeholder",
    )
    # safety_tours — 3
    await db.safety_tours.create_index(
        [("project_id", 1), ("tour_date", -1)],
        background=True, name="idx_str_project_date",
    )
    await db.safety_tours.create_index(
        [("project_id", 1), ("status", 1)],
        background=True, name="idx_str_project_status",
    )
    await db.safety_tours.create_index(
        [("project_id", 1), ("deletedAt", 1)],
        background=True, name="idx_str_project_deleted",
    )

    logger.info("Safety indices ensured (23 total across 7 collections)")
